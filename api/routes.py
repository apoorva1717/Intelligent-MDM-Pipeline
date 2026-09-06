"""Route definitions for the enrichment API."""

from __future__ import annotations

import io
import logging
import time
from collections import Counter
from typing import Any, Literal, Optional

from fastapi import APIRouter, Body, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import AliasChoices, ValidationError

from api.models import (
    EnrichmentOptions,
    EnrichmentRecord,
    EnrichmentRequest,
    EnrichmentResponse,
    EnrichmentResult,
    HealthResponse,
    IssueDetectionRequest,
    IssueDetectionResponse,
    IssueDetectionResult,
    TierConfigResponse,
)
from api.output_columns import RESPONSE_COLUMNS
from config import Settings, get_settings
from dedup.adjudicator import cluster_blocks
from dedup.consolidate import (
    EXAMPLE_REQUEST as CONSOLIDATE_EXAMPLE_REQUEST,
    EXAMPLE_RESPONSE as CONSOLIDATE_EXAMPLE_RESPONSE,
    ConsolidateRequest,
    ConsolidateResponse,
    consolidate_rows,
)
from dedup.consolidate_xlsx import ConsolidateFileError, consolidate_workbook
from dedup.llm import DedupLLM
from dedup.flags import v2_blocking, v2_id_conflict, v2_name2
from dedup.models import DedupRequest, DedupResponse, DedupRow
from dedup.prompts import prompt_version
from dedup.scoring import (
    ApprovalRequest,
    ApprovalResponse,
    ClusterNotFoundError,
    DuplicateRowIdError,
    ScoringRequest,
    ScoringResponse,
    apply_approval,
    build_summary,
    coerce_weights,
    detect_issues as detect_dedup_issues,
    elect_golden_records,
    load_weights,
)
from dedup.scoring_xlsx import ScoringFileError, score_workbook
from enrichment.issue_detection import (
    ISSUE_CATALOGUE,
    PERSISTENT_GROUP,
    REDUCIBLE_GROUPS,
    VERIFICATION_GROUPS,
    detect_issues,
    flag_for_review_is_set,
    provenance_is_low,
    split_flag_codes,
    DERIVED_LOW_FLAG_CODE,
)
from enrichment.flags import name2_needs_no_verification
from enrichment.orchestrator import Orchestrator

logger = logging.getLogger(__name__)

router = APIRouter()


def _get_orchestrator(settings: Settings | None = None) -> Orchestrator:
    """Build an Orchestrator with real or mock clients based on config."""
    if settings is None:
        settings = get_settings()

    mock_clients = None
    if settings.mock_external_calls:
        from tests.mocks.openai_mock import MockOpenAIClient
        from tests.mocks.page_mock import MockPageFetcher
        from tests.mocks.ror_mock import MockRORClient
        from tests.mocks.serp_mock import MockSearchClient

        mock_clients = {
            "ror": MockRORClient(settings),
            "search": MockSearchClient(),
            "page_fetcher": MockPageFetcher(),
            "llm": MockOpenAIClient(),
        }
        logger.info("Mock mode enabled — using mock clients")

    return Orchestrator(settings, mock_clients=mock_clients)


@router.get("/health", response_model=HealthResponse)
async def health_check() -> HealthResponse:
    """Health check endpoint for ADF and monitoring."""
    settings = get_settings()
    return HealthResponse(
        status="healthy",
        version="1.0.0",
        env=settings.env,
        mock_mode=settings.mock_external_calls,
        tiers_available=[1, 2, 3],
    )


@router.post("/enrich", response_model=EnrichmentResponse)
async def enrich_records(request: EnrichmentRequest) -> EnrichmentResponse:
    """Main enrichment endpoint — processes a batch of customer master data records.

    Always returns HTTP 200 for valid requests.  Per-record errors are
    reported in each result's `error` field.  HTTP 422 for validation
    errors (Pydantic).  HTTP 400 only for empty records array (handled
    by Pydantic min_length=1).
    """
    settings = get_settings()
    orchestrator = _get_orchestrator(settings)

    logger.info(
        "Enrichment request received: %d records, concurrency=%d",
        len(request.records),
        request.options.max_concurrency,
    )

    response = await orchestrator.enrich_batch(request.records, request.options)
    return response


# ---------------------------------------------------------------------------
# XLSX <-> records helpers for the /enrich/file endpoint
# ---------------------------------------------------------------------------


def _norm_header(name: str) -> str:
    """Normalise a column header for tolerant matching.

    Real spreadsheet exports vary in case, surrounding/internal whitespace,
    and punctuation (e.g. "Name 1" vs "name1" vs "NAME 1 "). Collapsing
    those differences lets enriched values land back on the right original
    column instead of being appended as a duplicate.
    """
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _input_alias_to_field() -> dict[str, str]:
    """Reverse-map every accepted input header/alias (normalised) to its
    model field name."""
    alias_to_field: dict[str, str] = {}
    for field_name, field in EnrichmentRecord.model_fields.items():
        alias = field.validation_alias
        if isinstance(alias, AliasChoices):
            names = list(alias.choices)
        elif isinstance(alias, str):
            names = [alias]
        else:
            names = []
        names.append(field_name)  # populate_by_name is enabled
        for name in names:
            alias_to_field.setdefault(_norm_header(name), field_name)
    return alias_to_field


def _present_fields(headers: list[str]) -> set[str]:
    """Map a file's header row onto the set of EnrichmentRecord field names it
    actually carries.

    Used by the issue endpoints so the required-field rules only judge columns
    that exist in the file (an enriched export that drops, say, Postal Code is
    not reported as "missing" it).
    """
    alias_to_field = _input_alias_to_field()
    present: set[str] = set()
    for header in headers:
        field = alias_to_field.get(_norm_header(header))
        if field is not None:
            present.add(field)
    return present


def _flag_for_review(row: dict[str, str], headers: list[str]) -> bool | None:
    """The row's ``Flag for Review`` value, or ``None`` when the file has no
    such column.

    ``None`` and ``False`` are not interchangeable here: ``None`` means "this
    is a raw input file, the question does not apply", while ``False`` means
    "this is an enriched file and the pipeline did not flag this record".
    G7-VERIFY-001, the code this used to drive, is withdrawn; the value is
    still read and still passed to ``detect_issues``, which no longer acts
    on it.
    """
    for header in headers:
        if _norm_header(header) == _norm_header("Flag for Review"):
            return flag_for_review_is_set(row.get(header))
    return None


#: The provenance columns that carry the doubt ``low-confidence-unchanged``
#: used to name as a flag code. Name 1 and Name 2 only — the same two fields
#: the pipeline derives its own review flag from (``flags.CORE_PROVENANCE_
#: FIELDS``): the flag asks a human to check a NAME, and a record whose domain
#: or type could not be settled is not a record with a wrong name in it.
_CORE_PROVENANCE_HEADERS: tuple[str, ...] = ("Name 1 Provenance", "Name 2 Provenance")


def _name2_has_no_canonical_form(record: EnrichmentRecord) -> bool:
    """Whether this record's Name 2 is a phrase nothing could confirm.

    The pipeline's own exemption, asked of a parsed row instead of a record in
    flight: an administrative desk ("Accounts Payable", "Central Receiving")
    or a phrase built only of facility functions ("Central Warehouse") has no
    institutional spelling for a reviewer to establish, so ``compute_flags``
    withholds the derived doubt and ships ``input:low`` on the field anyway —
    the column states what happened, and only the review request is withheld.

    An audit that read the provenance column and stopped there would put the
    request back, and ask a steward to canonicalise a phrase the pipeline
    declined to search for. The predicate is
    :func:`enrichment.flags.name2_needs_no_verification`, the same one the
    flag uses; it is handed a mapping rather than the model because it reads
    the enriched-record field names and the address tokens its
    ``identifies_nothing`` half compares against.
    """
    return name2_needs_no_verification({
        "name2_enriched": record.name2,
        "city": record.city,
        "region": record.region,
        "country_region_key": record.country_region_key,
    })


def _flag_codes(
    row: dict[str, str],
    headers: list[str],
    record: EnrichmentRecord | None = None,
) -> list[str] | None:
    """The row's enrichment flag codes, or ``None`` when this is a raw file.

    ``None`` and ``[]`` carry the same distinction as in :func:`_flag_for_review`:
    ``None`` means "no enrichment output here, the flag-derived codes cannot
    apply", ``[]`` means "enrichment ran and flagged nothing". Both suppress
    G6-RESOLVE-001 / G7-CONFIRM-001 / G8-VERIFY-001; only ``None`` says the
    question was never asked.

    ``low-confidence-unchanged`` is added from the provenance columns as well
    as read from ``Flag Codes``. The pipeline emits the token again, so a
    current export names it; an export taken while it was withdrawn has the
    state in ``Name 1 / Name 2 Provenance`` and nowhere else, and reading only
    ``Flag Codes`` would leave G8-VERIFY-001 dark for the largest population
    it exists to describe — the rows the pipeline left exactly as supplied.
    Both paths reach the same code, and the ``not in`` check is what keeps a
    row that carries it twice from saying so.

    *record* is the parsed row, and is consulted for one thing: the Name 2
    exemption the pipeline itself applies (see
    :func:`_name2_has_no_canonical_form`). Without it the two halves of this
    function contradict each other — the ``Flag Codes`` half honours the
    exemption because ``compute_flags`` made it, and the provenance half would
    re-raise the doubt from a column that is `input:low` precisely because
    nothing could confirm a phrase that has nothing to confirm.
    """
    codes: list[str] | None = None
    for header in headers:
        if _norm_header(header) == _norm_header("Flag Codes"):
            codes = split_flag_codes(row.get(header))
            break

    for header in headers:
        norm = _norm_header(header)
        if not any(norm == _norm_header(c) for c in _CORE_PROVENANCE_HEADERS):
            continue
        if codes is None:
            codes = []
        if not provenance_is_low(row.get(header)):
            continue
        if (
            norm == _norm_header("Name 2 Provenance")
            and record is not None
            and _name2_has_no_canonical_form(record)
        ):
            # The pipeline looked at this slot and decided there was nothing
            # to ask. Keep looking at the other column rather than stopping:
            # a Name 1 doubt on the same row still stands.
            continue
        if DERIVED_LOW_FLAG_CODE not in codes:
            codes.append(DERIVED_LOW_FLAG_CODE)
        break
    return codes


def _parse_xlsx(contents: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse an uploaded XLSX into (headers, row dicts).

    The first non-empty row is the header row; its cells map onto the SAP
    column names accepted by EnrichmentRecord (e.g. "Name 1", "Customer").
    Each subsequent non-blank row becomes a dict keyed by header.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required to parse XLSX uploads but is not installed.",
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any parse failure is a bad upload
        raise HTTPException(
            status_code=400,
            detail=f"Could not read uploaded file as XLSX: {exc}",
        ) from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    # Locate the header row (skip leading fully-empty rows).
    headers: list[str] = []
    for raw_header in rows:
        if raw_header is None:
            continue
        candidate = [
            str(cell).strip() if cell is not None else "" for cell in raw_header
        ]
        if any(candidate):
            headers = candidate
            break

    if not headers:
        workbook.close()
        raise HTTPException(status_code=400, detail="Uploaded XLSX has no header row.")

    row_dicts: list[dict[str, str]] = []
    for raw_row in rows:
        if raw_row is None:
            continue
        row_dict: dict[str, str] = {}
        for header, cell in zip(headers, raw_row):
            if not header or cell is None:
                continue
            value = str(cell).strip()
            if value:
                row_dict[header] = value
        if row_dict:  # skip entirely blank rows
            row_dicts.append(row_dict)

    workbook.close()

    if not row_dicts:
        raise HTTPException(
            status_code=400, detail="No data rows found in uploaded XLSX."
        )

    return headers, row_dicts


def _rows_to_records(row_dicts: list[dict[str, str]]) -> list[EnrichmentRecord]:
    """Validate row dicts into EnrichmentRecords using the shared model.

    Headers are normalised to their canonical model field name first, so a
    column written as "NAME 1", "Name1", or " name 1 " is still recognised
    as the same input field (model_validate alone only honours the exact
    declared aliases).
    """
    alias_to_field = _input_alias_to_field()
    records: list[EnrichmentRecord] = []
    errors: list[str] = []
    for index, row_dict in enumerate(row_dicts, start=1):
        # Map each raw header onto the model field it populates; keep the
        # first non-empty value when several headers resolve to one field.
        normalised: dict[str, str] = {}
        for header, value in row_dict.items():
            field = alias_to_field.get(_norm_header(header))
            key = field if field is not None else header
            if key not in normalised:
                normalised[key] = value
        try:
            records.append(EnrichmentRecord.model_validate(normalised))
        except ValidationError as exc:
            errors.append(f"row {index}: {exc.errors()[0].get('msg', str(exc))}")

    if errors:
        raise HTTPException(
            status_code=422,
            detail={"message": "Some rows failed validation", "errors": errors},
        )
    return records


def _copy_extra_sheets(contents: bytes, wb) -> None:
    """Copy every sheet except the data sheet (values only) from the uploaded
    workbook into ``wb``.

    The pipeline is /enrich/file -> /api/dedup/file -> /api/dedup/score/file,
    and the scoring workbook ships a "Weights" sheet the last stage reads.
    Rebuilding the output workbook must not silently drop such sheets on the
    way through.
    """
    from openpyxl import load_workbook

    try:
        src = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - the upload already parsed once; be lenient
        return
    active_title = src.active.title if src.active else None
    for sheet in src.worksheets:
        if sheet.title == active_title:
            continue
        out = wb.create_sheet(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            out.append(list(row))
    src.close()


def _passthrough_headers(input_headers: list[str], output_headers: list[str]) -> list[str]:
    """Input columns that enrichment neither consumes nor re-emits.

    These are carried through verbatim (e.g. the CRM/scoring columns —
    Sales_Order_*, SleepingCustomer, SF_ID_*, score_* — that
    /api/dedup/score/file needs at the end of the pipeline) so no field is
    lost before it reaches the stage that uses it.
    """
    known = {_norm_header(h) for h in output_headers}
    known |= set(_input_alias_to_field())
    passthrough: list[str] = []
    seen: set[str] = set()
    for header in input_headers:
        key = _norm_header(header)
        if not key or key in known or key in seen:
            continue
        seen.add(key)
        passthrough.append(header)
    return passthrough


def _cell(value: object) -> object:
    """Render one response value as an XLSX cell.

    openpyxl accepts only scalars, so a list-valued response field
    (``flag_codes``, ``flagged_fields``) is joined into the semicolon-separated
    form the other multi-value columns already use. Every other value is
    written unchanged — this is a serialisation adapter, not a transform.
    """
    if isinstance(value, (list, tuple)):
        return "; ".join(str(v) for v in value) or None
    return value


def _build_output_xlsx(
    results: list[EnrichmentResult],
    input_headers: list[str] | None = None,
    row_dicts: list[dict[str, str]] | None = None,
    source_contents: bytes | None = None,
) -> bytes:
    """Build the result workbook.

    One column per field in the /enrich response body (see
    api.output_columns.RESPONSE_COLUMNS), one row per enriched record.
    Unconsumed input columns are appended verbatim (results arrive in input
    order — see enrich_batch's asyncio.gather — so passthrough values join
    by position) and extra sheets such as "Weights" are copied over, so the
    downstream dedup and scoring stages never lose fields.
    """
    from openpyxl import Workbook

    fields = list(RESPONSE_COLUMNS.keys())
    headers = [RESPONSE_COLUMNS[f] for f in fields]
    passthrough = (
        _passthrough_headers(input_headers, headers)
        if input_headers and row_dicts is not None
        else []
    )

    wb = Workbook()
    ws = wb.active
    ws.append([*headers, *passthrough])

    for index, result in enumerate(results):
        data = result.model_dump()
        row_dict = row_dicts[index] if row_dicts is not None else {}
        ws.append(
            [_cell(data.get(field)) for field in fields]
            + [row_dict.get(header) for header in passthrough]
        )

    if source_contents is not None:
        _copy_extra_sheets(source_contents, wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_issues_xlsx(
    headers: list[str],
    row_dicts: list[dict[str, str]],
    issues_per_row: list[list[str]],
) -> bytes:
    """Echo the uploaded sheet with one appended ``Issues`` column.

    The original header row and cell values are preserved verbatim (blanks
    filled where a cell was empty); a trailing ``Issues`` column holds the
    semicolon-joined issue codes detected for that row (empty when clean).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([*headers, "Issues"])

    for row_dict, codes in zip(row_dicts, issues_per_row):
        values = [row_dict.get(header, "") for header in headers]
        ws.append([*values, "; ".join(codes)])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def _audit_upload(file: UploadFile) -> dict[str, list[str]]:
    """Validate, parse, and audit an XLSX upload into ``{record_id: [codes]}``.

    Detection is column-aware (see ``_present_fields``). Rows without a record
    id cannot be joined across files and are excluded (logged, not silently
    dropped). The first occurrence wins for a duplicated id.
    """
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail=f"{filename or 'Uploaded file'} must be an .xlsx (or .xlsm) workbook.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail=f"{filename or 'Uploaded file'} is empty.")

    headers, row_dicts = _parse_xlsx(contents)
    records = _rows_to_records(row_dicts)
    present = _present_fields(headers)

    issue_map: dict[str, list[str]] = {}
    excluded = 0
    for record, row in zip(records, row_dicts):
        rid = record.record_id
        if not rid:
            excluded += 1
            continue
        issue_map.setdefault(
            rid,
            detect_issues(
                record, present,
                flag_for_review=_flag_for_review(row, headers),
                flag_codes=_flag_codes(row, headers, record),
            ),
        )

    if excluded:
        logger.info(
            "issues/compare: %s — %d row(s) without a record id excluded from join",
            filename,
            excluded,
        )
    return issue_map


def _build_comparison_xlsx(
    before_map: dict[str, list[str]],
    after_map: dict[str, list[str]],
) -> bytes:
    """Build the before/after issue-reduction report.

    The report is **segmented by catalogue group**, because a single "issues
    remaining" total conflates three things that mean different things:

    * **Reduced** (G1-G5) — the actual story. These codes have a remediation
      path, so a code present before and absent after is work the pipeline did.
      The headline reduction percentage is computed over these groups alone.
    * **Expected to persist** (G6) — "Not Resolvable by Enrichment": no
      automated path exists to supply the value, so these codes are *supposed*
      to survive to the enriched file and be routed to a steward. Folding them
      into the reduction total counts the pipeline down for defects it was
      never able to fix.
    * **Verification** (G7) — raised *by* successful enrichment, not by a
      defect. Counting it would inflate the post-pipeline total in proportion
      to how well enrichment performed, inverting the meaning of the delta. It
      is reported on its own and never enters any reduction figure.

    Sheet 1 (Summary): the three blocks above + a per-code Before/After/Delta
    table carrying each code's group and segment.
    Sheet 2 (Per Record): every matched record's before codes, after codes,
    which issues were resolved, and which were newly introduced.
    Sheet 3 (Remaining Issues): for every issue still present after
    enrichment, the customer id of each record that still has it (one row
    per code/customer pairing), labelled with its segment.
    Comparison is over records present in BOTH files (joined by record id).
    """
    from openpyxl import Workbook

    matched_ids = [rid for rid in before_map if rid in after_map]
    only_before = [rid for rid in before_map if rid not in after_map]
    only_after = [rid for rid in after_map if rid not in before_map]

    def segment(code: str) -> str:
        """Which of the three report blocks *code* belongs to."""
        group = ISSUE_CATALOGUE[code].group
        if group in VERIFICATION_GROUPS:
            return "Verification"
        if group == PERSISTENT_GROUP:
            return "Expected to persist"
        return "Reduced"

    code_before: Counter = Counter()
    code_after: Counter = Counter()
    per_record_rows: list[list[str]] = []
    # Per-segment tallies. Only "Reduced" feeds the headline percentage.
    seg_before: Counter = Counter()
    seg_after: Counter = Counter()
    seg_resolved: Counter = Counter()
    seg_introduced: Counter = Counter()

    for rid in matched_ids:
        before = before_map[rid]
        after = after_map[rid]
        bset, aset = set(before), set(after)
        # Order resolved/introduced by catalogue order for stable output.
        resolved = [c for c in ISSUE_CATALOGUE if c in bset - aset]
        introduced = [c for c in ISSUE_CATALOGUE if c in aset - bset]

        for code in bset:
            seg_before[segment(code)] += 1
        for code in aset:
            seg_after[segment(code)] += 1
        for code in resolved:
            seg_resolved[segment(code)] += 1
        for code in introduced:
            seg_introduced[segment(code)] += 1
        code_before.update(bset)
        code_after.update(aset)

        per_record_rows.append([
            rid,
            "; ".join(before),
            "; ".join(after),
            "; ".join(resolved),
            "; ".join(introduced),
        ])

    reduced_before = seg_before["Reduced"]
    reduced_after = seg_after["Reduced"]
    net = reduced_before - reduced_after
    pct = (net / reduced_before * 100) if reduced_before else 0.0

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Issue Reduction Summary"])
    summary.append([])
    summary.append(["Records matched (joined by id)", len(matched_ids)])
    summary.append(["Records only in original", len(only_before)])
    summary.append(["Records only in enriched", len(only_after)])
    summary.append([])

    # --- Block 1: the reduction metric (G1-G5 only) ---
    # Labels are block-prefixed and unique across the sheet: three blocks each
    # carrying a bare "Issues before" would be ambiguous to a reader and would
    # collide for anything reading the sheet as label -> value pairs.
    summary.append([f"Reduced — {', '.join(REDUCIBLE_GROUPS)} (the reduction metric)"])
    summary.append(["Reduced: issues before", reduced_before])
    summary.append(["Reduced: issues after", reduced_after])
    summary.append(["Reduced: issues resolved", seg_resolved["Reduced"]])
    summary.append(["Reduced: issues introduced", seg_introduced["Reduced"]])
    summary.append(["Reduced: net reduction", net])
    summary.append(["Reduction %", round(pct, 1)])
    summary.append([])

    # --- Block 2: G6, where persistence is the correct outcome ---
    summary.append([
        f"Expected to persist — {PERSISTENT_GROUP} (Not Resolvable by Enrichment)"
    ])
    summary.append([
        "These codes have no automated remediation path. They are routed to a "
        "data steward, and their survival to the enriched file is correct "
        "behaviour — not an unreduced defect. Excluded from the reduction %."
    ])
    summary.append(["Expected to persist: issues before", seg_before["Expected to persist"]])
    summary.append(["Expected to persist: issues after", seg_after["Expected to persist"]])
    summary.append([
        "Expected to persist: persisted as expected",
        seg_before["Expected to persist"] - seg_resolved["Expected to persist"],
    ])
    summary.append([
        "Expected to persist: newly raised after enrichment",
        seg_introduced["Expected to persist"],
    ])
    summary.append([])

    # --- Block 3: G7, never part of any reduction figure ---
    summary.append([
        f"Verification — {', '.join(VERIFICATION_GROUPS)} (reported separately)"
    ])
    summary.append([
        "Raised BY enrichment so DATAshaper can assign the record to a "
        "steward — G7 to confirm a value the pipeline wrote, G8 where it "
        "could not establish one. Not quality issues and never counted in "
        "the reduction metric; see the Flag Reason column for the "
        "per-record trigger."
    ])
    summary.append(["Verification: records requiring verification", seg_after["Verification"]])
    summary.append([])

    summary.append(["Code", "Name", "Group", "Segment", "Before", "After", "Delta"])
    for code, entry in ISSUE_CATALOGUE.items():
        before_count = code_before.get(code, 0)
        after_count = code_after.get(code, 0)
        if before_count or after_count:
            summary.append([
                code, entry.name, entry.group, segment(code),
                before_count, after_count, after_count - before_count,
            ])

    per_record = wb.create_sheet("Per Record")
    per_record.append(
        ["record_id", "Issues Before", "Issues After", "Resolved", "Introduced"]
    )
    for row in per_record_rows:
        per_record.append(row)

    # Sheet 3 — Remaining Issues. For every issue still present after
    # enrichment, list the customer id of each record that still has it.
    # One row per (code, customer) so the sheet can be filtered or pivoted
    # by either column. Ordered by catalogue order, then customer id.
    # Covers every record in the enriched file (matched + enriched-only),
    # so no remaining issue is omitted. The Segment column separates a real
    # unreduced defect from an expected-persistence one and from a
    # verification disposition.
    remaining = wb.create_sheet("Remaining Issues")
    remaining.append(["Code", "Name", "Group", "Segment", "Customer"])
    for code, entry in ISSUE_CATALOGUE.items():
        ids = sorted(rid for rid, codes in after_map.items() if code in codes)
        for rid in ids:
            remaining.append([code, entry.name, entry.group, segment(code), rid])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.post("/enrich/file")
async def enrich_file(
    file: UploadFile = File(..., description="XLSX file of customer master data records"),
    max_concurrency: int = Query(default=5, ge=1, le=20),
    serp_provider: Literal["serpapi", "duckduckgo"] = Query(default="serpapi"),
    skip_tier: Optional[int] = Query(default=None),
) -> StreamingResponse:
    """Enrichment endpoint that takes an XLSX upload and returns an XLSX.

    The spreadsheet uses the same input fields as the /enrich request
    records (SAP column headers). Each data row is turned into one
    EnrichmentRecord, the records are enriched exactly as the JSON /enrich
    endpoint does, and the results are written to a workbook with one column
    per field in the /enrich response body (see api.output_columns).
    """
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file must be an .xlsx (or .xlsm) workbook.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    headers, row_dicts = _parse_xlsx(contents)
    records = _rows_to_records(row_dicts)
    options = EnrichmentOptions(
        max_concurrency=max_concurrency,
        serp_provider=serp_provider,
        skip_tier=skip_tier,
    )
    request = EnrichmentRequest(records=records, options=options)

    settings = get_settings()
    orchestrator = _get_orchestrator(settings)

    logger.info(
        "Enrichment file request received: %s, %d records, concurrency=%d",
        filename,
        len(request.records),
        request.options.max_concurrency,
    )

    response = await orchestrator.enrich_batch(request.records, request.options)

    output_bytes = _build_output_xlsx(
        response.results, headers, row_dicts, source_contents=contents
    )

    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "records"
    out_name = f"{stem}_enriched.xlsx"
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


def _audit_rows(
    headers: list[str], row_dicts: list[dict[str, str]]
) -> tuple[list[EnrichmentRecord], list[list[str]]]:
    """Validate a parsed sheet and detect the audit issue codes for each row.

    The one detection path behind both ``/issues`` and ``/issues/json``: same
    column-awareness, same ``Flag for Review`` and ``Flag Codes`` handling, so
    the two representations of the audit cannot drift apart. The validated
    records come back with the codes because the JSON response keys on
    ``record.record_id``.

    Nothing is withheld from the result. A suppression set used to drop four
    codes from the audit column; every code it hid on the grounds that no
    automated path resolves it is now withdrawn from the catalogue outright,
    and the one survivor — G2-NAME-012 — is reported like any other.
    """
    records = _rows_to_records(row_dicts)
    present = _present_fields(headers)
    issues_per_row = [
        detect_issues(
            record, present,
            flag_for_review=_flag_for_review(row, headers),
            flag_codes=_flag_codes(row, headers, record),
        )
        for record, row in zip(records, row_dicts)
    ]
    return records, issues_per_row


def _json_rows(
    records: list[dict[str, Any]]
) -> tuple[list[str], list[dict[str, str]]]:
    """Turn JSON record objects into the (headers, row dicts) shape the audit
    helpers already speak.

    Mirrors ``_parse_xlsx`` cell-for-cell: values are stringified and stripped,
    and empty ones are dropped so a blank string and an absent key mean the
    same thing they do in a workbook. The header list is the union of the keys
    across all records, in first-seen order — the JSON equivalent of the sheet's
    header row, which is what makes detection column-aware (a payload where no
    record carries ``Postal Code`` is not reported as missing it).

    An integral JSON float is rendered without its ``.0``: openpyxl hands back
    a whole-numbered cell as an ``int``, so a postal code sent as ``12345.0``
    has to reach the detector as ``"12345"`` and not as ``"12345.0"`` — which
    would otherwise be read as a malformed postal code the workbook never saw.

    Unlike ``_parse_xlsx`` this keeps an all-empty record instead of skipping
    it. A blank spreadsheet row is a parsing artefact; an empty JSON object was
    deliberately sent, and the response is positional — one result per request
    record — so dropping it would silently misalign the caller's join.
    """
    headers: list[str] = []
    seen: set[str] = set()
    row_dicts: list[dict[str, str]] = []
    for record in records:
        row: dict[str, str] = {}
        for key, value in record.items():
            header = str(key).strip()
            if not header:
                continue
            if header not in seen:
                seen.add(header)
                headers.append(header)
            if value is None:
                continue
            if isinstance(value, float) and not isinstance(value, bool) and value.is_integer():
                value = int(value)
            text = str(value).strip()
            if text:
                row[header] = text
        row_dicts.append(row)
    return headers, row_dicts


@router.post("/issues")
async def detect_file_issues(
    file: UploadFile = File(..., description="XLSX file of customer master data records"),
) -> StreamingResponse:
    """Audit an XLSX upload against the Issue Catalogue and return it annotated.

    Each data row is checked against the deterministic issue-detection rules
    (see ``enrichment.issue_detection``). The uploaded sheet is returned
    unchanged with a single appended ``Issues`` column listing every issue
    code found on that row (semicolon-separated, codes only). This is a pure
    audit: no enrichment, LLM, or external call is made.
    """
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file must be an .xlsx (or .xlsm) workbook.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    headers, row_dicts = _parse_xlsx(contents)
    records, issues_per_row = _audit_rows(headers, row_dicts)

    logger.info(
        "Issues file request received: %s, %d records, %d with issues",
        filename,
        len(records),
        sum(1 for issues in issues_per_row if issues),
    )

    output_bytes = _build_issues_xlsx(headers, row_dicts, issues_per_row)

    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "records"
    out_name = f"{stem}_issues.xlsx"
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@router.post("/issues/json", response_model=IssueDetectionResponse)
async def detect_json_issues(request: IssueDetectionRequest) -> IssueDetectionResponse:
    """JSON twin of ``POST /issues`` — audit records against the Issue Catalogue.

    Takes the same content the file endpoint takes, as JSON: one object per
    row carrying the file's columns keyed by header (``"Customer"``,
    ``"Name 1"``, ``"Postal Code"`` …). Returns one ``{record_id, issues}``
    entry per record, in request order, where ``issues`` holds the codes
    detected for that record — the same list the file endpoint writes into
    its ``Issues`` column.

    Detection is column-aware over the union of keys present across the
    request's records, so send every column the row has (including empty
    ones) to get the same verdict the workbook would produce. This is a pure
    audit: no enrichment, LLM, or external call is made.
    """
    headers, row_dicts = _json_rows(request.records)
    records, issues_per_row = _audit_rows(headers, row_dicts)

    logger.info(
        "Issues JSON request received: %d records, %d with issues",
        len(records),
        sum(1 for issues in issues_per_row if issues),
    )

    return IssueDetectionResponse(
        results=[
            IssueDetectionResult(record_id=record.record_id, issues=issues)
            for record, issues in zip(records, issues_per_row)
        ]
    )


@router.post("/issues/compare")
async def compare_file_issues(
    original: UploadFile = File(..., description="Raw (pre-enrichment) XLSX"),
    enriched: UploadFile = File(..., description="Enriched (post-pipeline) XLSX"),
) -> StreamingResponse:
    """Audit two files and report how many issues the pipeline reduced.

    Runs the deterministic issue detector over both the ``original`` and the
    ``enriched`` workbook, joins their rows by record id, and returns an XLSX
    report: a summary (totals + per-code Before/After/Delta) and a per-record
    breakdown of which issues were resolved vs newly introduced.

    Detection is column-aware, so a required-field rule never counts a column
    as "missing" in a file that simply doesn't carry it (e.g. the enriched
    export dropping Postal Code) — the before/after counts stay apples-to-apples.
    """
    before_map = await _audit_upload(original)
    after_map = await _audit_upload(enriched)

    logger.info(
        "Issues compare request: original=%s (%d records), enriched=%s (%d records)",
        original.filename,
        len(before_map),
        enriched.filename,
        len(after_map),
    )

    output_bytes = _build_comparison_xlsx(before_map, after_map)
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="issue_reduction_report.xlsx"'
        },
    )


# ---------------------------------------------------------------------------
# Preprocess — row grain -> customer grain (runs BEFORE clustering/scoring)
# ---------------------------------------------------------------------------


@router.post(
    "/api/preprocess/consolidate",
    response_model=ConsolidateResponse,
    # The example is also on the models, but Swagger UI renders a media-type
    # example unconditionally — and an endpoint whose body is a list of
    # free-form dicts is exactly the one a caller cannot guess from the schema.
    responses={
        200: {
            "content": {
                "application/json": {"example": CONSOLIDATE_EXAMPLE_RESPONSE}
            }
        }
    },
)
async def preprocess_consolidate(
    request: ConsolidateRequest = Body(..., openapi_examples={
        "one_customer_three_rows": {
            "summary": "One customer over three row-grain rows",
            "description": (
                "Shows the whole contract in three rows: '0013119338' and "
                "'13119338' group as ONE customer; all three company codes "
                "land on EVERY row; the two blank sales orgs are dropped "
                "rather than padded, so the sales-org list has one entry, not "
                "three; and 'Name 1' — a column this stage never reads — is "
                "echoed back untouched."
            ),
            "value": CONSOLIDATE_EXAMPLE_REQUEST,
        },
    }),
) -> ConsolidateResponse:
    """Consolidate company codes and sales orgs to customer grain (JSON).

    The SAP extract is at ROW grain — one row per customer per company-code /
    sales-area assignment. This appends the two CUSTOMER-level columns the
    scorer needs, ``Company_Code_Consolidated`` and ``Sales_Org_Consolidated``,
    onto every row of each customer. Rows in == rows out, same order, no row
    dropped, merged or otherwise altered; only ``Customer``, ``Company Code``
    and ``Sales Organization`` are read (snake_case aliases accepted).

    BATCHING INVARIANT: consolidation is only correct when EVERY row of a
    customer is in the same request. This endpoint cannot verify that from
    inside one request, so it emits a HEURISTIC warning naming the customers
    at the first and last row positions — the only places a split would show.
    That is a warning, not a guarantee: prefer
    POST /api/preprocess/consolidate/file, which always sees the whole file.

    A blank Customer is an error counted in ``summary.errors``; the row is
    still returned, unchanged, with both columns empty. Never a 500.
    """
    request_start = time.perf_counter()
    logger.info("Consolidate request received: %d rows", len(request.rows))

    rows, summary = consolidate_rows(request.rows, warn_batch_boundary=True)
    for warning in summary.warnings:
        logger.warning("consolidate request: %s", warning)

    logger.info(
        "consolidate_request",
        extra={
            "summary": summary.model_dump(),
            "total_latency_ms": int((time.perf_counter() - request_start) * 1000),
        },
    )
    return ConsolidateResponse(rows=rows, summary=summary)


@router.post("/api/preprocess/consolidate/file")
async def preprocess_consolidate_file(
    file: UploadFile = File(..., description="SAP row-grain extract (.xlsx/.xlsm)"),
    sheet: Optional[str] = Query(
        default=None,
        description="Worksheet to consolidate. Defaults to the first worksheet.",
    ),
) -> StreamingResponse:
    """Same consolidation as /api/preprocess/consolidate, XLSX in / XLSX out.

    The workbook is edited IN PLACE with openpyxl: every other sheet and every
    original column survive untouched, and the two consolidated columns are
    located or appended by header name (so a re-run overwrites rather than
    duplicating them). Identical column names to the JSON transport.

    This endpoint processes the WHOLE workbook and is therefore always safe
    with respect to the batching invariant — run it once over the complete
    extract, before any batching.
    """
    request_start = time.perf_counter()
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file must be an .xlsx (or .xlsm) workbook.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        output_bytes, summary = consolidate_workbook(contents, sheet)
    except ConsolidateFileError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    logger.info(
        "consolidate_request",
        extra={
            "summary": summary.model_dump(),
            # "filename" is a reserved LogRecord attribute — use upload_name.
            "upload_name": filename,
            "total_latency_ms": int((time.perf_counter() - request_start) * 1000),
        },
    )

    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "records"
    out_name = f"{stem}_consolidated.xlsx"
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


def _get_dedup_llm(settings: Settings):
    """Build the dedup adjudicator LLM, real or mock based on config.

    Mirrors ``_get_orchestrator`` — when ``MOCK_EXTERNAL_CALLS`` is set the
    mock client is used so the endpoint runs offline.
    """
    from dedup.cache import wrap_if_enabled

    if settings.mock_external_calls:
        from tests.mocks.dedup_mock import MockDedupLLM
        logger.info("Mock mode enabled — using mock dedup LLM")
        return MockDedupLLM()
    # Unwrapped unless DEDUP_FIXTURE_CACHE_DIR is set, so the shipped path is
    # unchanged; set it to record or replay a run (see dedup/cache.py).
    return wrap_if_enabled(DedupLLM(settings))


# ---------------------------------------------------------------------------
# XLSX <-> DedupRow helpers for the /api/dedup/file endpoint
# ---------------------------------------------------------------------------

# Normalised header (alnum, lowercase — see ``_norm_header``) → DedupRow
# field. Accepts both the human SAP headers used by the enriched export
# ("Customer", "Name 1", "Street 1", …) and the snake_case DedupRow field
# names, so the file produced by /enrich/file can be fed straight in.
_DEDUP_HEADER_ALIASES: dict[str, str] = {
    "rowid": "row_id",
    "recordid": "row_id",
    "customer": "row_id",
    "blockid": "block_id",
    "name1": "name1",
    "name2": "name2",
    "name3": "name3",
    "name4": "name4",
    "name5": "name5",
    "street": "street",
    "street1": "street",
    "streetcleaned": "street",
    "houseno": "house_no",
    "housenumber": "house_no",
    "postalcode": "postal_code",
    "zip": "postal_code",
    "city": "city",
    "country": "country",
    "countryregionkey": "country",
    "rorid": "ror_id",
    # "LEI ID" is what /enrich/file writes (api.output_columns.RESPONSE_COLUMNS),
    # and it normalises to the same key as "lei id" / "lei_id" / "LEI_ID".
    # Without this the column bound to nothing and the adjudicator ran without
    # the company legal-entity signal the JSON route carries.
    "leiid": "lei_id",
    "lei": "lei_id",
    "enrichedname": "enriched_name",
    # v2 (C.4). Six columns /enrich/file writes that adjudication never saw.
    # The first three are alternate names and the institution/company verdict —
    # the signals the slot classifier and the acronym/cross-slot candidate
    # rules read. The two provenance columns let an id conflict say WHY it is
    # a conflict rather than only that it is one. "Building" is bound as a
    # hint: it is shown to the model and kept out of blocking and the
    # signature key, because a building is not an entity.
    "operatingname": "operating_name",
    "suggestedname": "suggested_name",
    "recordtype": "record_type",
    "roridprovenance": "ror_id_provenance",
    "leiidprovenance": "lei_id_provenance",
    "building": "building",
}


def _rows_to_dedup_rows(row_dicts: list[dict[str, str]]) -> list[DedupRow]:
    """Validate parsed XLSX rows into DedupRows.

    Headers are normalised to their DedupRow field name first (see
    ``_DEDUP_HEADER_ALIASES``) so a column written as "Customer", "row_id",
    or "Name 1" is recognised. ``row_id`` is required by the model; a row
    missing it surfaces as a validation error rather than being dropped.

    A header the alias table does not know binds to nothing. That is correct
    for the CRM/scoring columns this stage only passes through, but it is also
    how a missing alias hides: the column is simply discarded and adjudication
    runs one signal short, with no error. Every unrecognised header is
    therefore named once at WARNING level.
    """
    rows: list[DedupRow] = []
    errors: list[str] = []
    unrecognised: dict[str, None] = {}  # insertion-ordered set of raw headers
    for index, row_dict in enumerate(row_dicts, start=1):
        normalised: dict[str, str] = {}
        for header, value in row_dict.items():
            field = _DEDUP_HEADER_ALIASES.get(_norm_header(header))
            if field is None:
                unrecognised.setdefault(header, None)
                continue
            if field in normalised:
                continue
            normalised[field] = value
        try:
            rows.append(DedupRow.model_validate(normalised))
        except ValidationError as exc:
            errors.append(f"row {index}: {exc.errors()[0].get('msg', str(exc))}")

    if unrecognised:
        logger.warning(
            "dedup file: %d column header(s) matched no DedupRow field and "
            "were not used for adjudication (passed through to the output "
            "unchanged): %s",
            len(unrecognised), ", ".join(repr(h) for h in unrecognised),
        )

    if errors:
        raise HTTPException(
            status_code=422,
            detail={"message": "Some rows failed validation", "errors": errors},
        )
    return rows


# Main-sheet result columns. Exactly ONE cluster key is exposed here —
# "Cluster ID" (the stable content hash). Internal keys (Block ID, Signature
# ID) live on the "Dedup Debug" sheet so an approver sees a single cluster
# column and can never confuse which one is authoritative.
_DEDUP_RESULT_COLUMNS = ["Cluster ID", "Routing", "LLM Flag", "Confidence", "Reasoning"]
_DEDUP_DEBUG_SHEET = "Dedup Debug"
_DEDUP_DEBUG_COLUMNS = ["row_id", "Cluster ID", "Block ID", "Signature ID"]

# v2 adds one column, and only while a v2 flag is on: with all three off the
# workbook must have exactly v1's columns. "Link ID" sits beside "Cluster ID"
# because a reader comparing the two is asking the question the pair exists to
# answer — same record, or merely same organisation?
_DEDUP_RESULT_COLUMNS_V2 = [
    "Cluster ID", "Link ID", "Routing", "LLM Flag", "Confidence", "Reasoning",
]
_DEDUP_DEBUG_COLUMNS_V2 = [
    "row_id", "Cluster ID", "Link ID", "Block ID", "Signature ID",
]

#: Written to every dedup workbook, v1 and v2 alike.
_DEDUP_RUN_SHEET = "Run"


def _dedup_run_metadata(response: DedupResponse) -> list[tuple[str, str]]:
    """What produced this workbook.

    Until now the output carried no record of the configuration behind it: no
    prompt version, no flag values, no model, no cache mode. Two workbooks
    could differ in every cluster and look identical in provenance, and the
    only way to tell a v1 run from a v2 one was to notice a missing column —
    which is exactly how a run with the flags unset gets mistaken for a run
    with them set.

    Read off the response rather than off the environment where it can be, so
    the sheet reports what the run DID rather than what the process was
    configured to do afterwards.
    """
    from dedup.cache import current_mode
    from dedup.flags import BLOCKING, ID_CONFLICT, NAME2, v2_any

    first = response.rows[0] if response.rows else None
    return [
        ("prompt_version", first.prompt_version if first else prompt_version()),
        ("model", first.model if first else ""),
        ("model_version", first.model_version if first else ""),
        (BLOCKING, str(v2_blocking()).lower()),
        (NAME2, str(v2_name2()).lower()),
        (ID_CONFLICT, str(v2_id_conflict()).lower()),
        ("dedup_v2_active", str(v2_any()).lower()),
        ("fixture_cache", current_mode()),
        ("rows_in", str(response.summary.rows_in)),
        ("blocks", str(response.summary.blocks)),
        ("llm_calls", str(response.summary.llm_calls)),
        ("rows_clustered", str(response.summary.rows_clustered)),
        ("rows_unique", str(response.summary.rows_unique)),
        ("rows_manual_review", str(response.summary.rows_manual_review)),
    ]


def _build_dedup_xlsx(
    headers: list[str],
    row_dicts: list[dict[str, str]],
    rows: list[DedupRow],
    response: DedupResponse,
    source_contents: bytes | None = None,
) -> bytes:
    """Echo the uploaded sheet with cluster-assignment columns appended.

    The original header row and cell values are preserved verbatim, then the
    result columns in ``_DEDUP_RESULT_COLUMNS`` are appended and joined onto
    each row by its ``row_id`` — with a SINGLE cluster key ("Cluster ID", the
    stable content hash). Internal keys (Block ID, Signature ID) are written to
    a separate "Dedup Debug" sheet. Rows the adjudicator did not return (should
    not happen) get blank result cells rather than being dropped. Extra sheets
    (e.g. the scoring "Weights" sheet) are copied over so the downstream
    /api/dedup/score/file stage still sees them.
    """
    from openpyxl import Workbook

    from dedup.flags import v2_any

    linked = v2_any()
    result_columns = _DEDUP_RESULT_COLUMNS_V2 if linked else _DEDUP_RESULT_COLUMNS
    debug_columns = _DEDUP_DEBUG_COLUMNS_V2 if linked else _DEDUP_DEBUG_COLUMNS

    result_by_id = {r.row_id: r for r in response.rows}

    wb = Workbook()
    ws = wb.active
    ws.append([*headers, *result_columns])

    debug_ws = wb.create_sheet(_DEDUP_DEBUG_SHEET)
    debug_ws.append(debug_columns)

    # Additive in both directions: v1's data sheets are untouched, and a v1
    # workbook gains the same sheet saying so.
    run_ws = wb.create_sheet(_DEDUP_RUN_SHEET)
    run_ws.append(["setting", "value"])
    for setting, value in _dedup_run_metadata(response):
        run_ws.append([setting, value])

    for row_dict, parsed in zip(row_dicts, rows):
        values = [row_dict.get(header, "") for header in headers]
        res = result_by_id.get(parsed.row_id)
        if res is None:
            ws.append([*values, *([""] * len(result_columns))])
            continue
        ws.append([
            *values,
            res.cluster_id,
            *([res.link_id] if linked else []),
            res.routing,
            res.llm_flag,
            res.confidence,
            res.reasoning,
        ])
        debug_ws.append([
            res.row_id, res.cluster_id,
            *([res.link_id] if linked else []),
            res.block_id, res.signature_id,
        ])

    if source_contents is not None:
        _copy_extra_sheets(source_contents, wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.post("/api/dedup/cluster-block", response_model=DedupResponse)
async def dedup_cluster_block(request: DedupRequest) -> DedupResponse:
    """Phase 2 "Pass 2" deduplication adjudicator.

    Takes address-gated candidate rows (already cleared by DATAshaper's
    address gates — same country + postal code + street) grouped into one or
    more blocks, decides which rows refer to the same (institution,
    department) entity, and returns cluster assignments. Auth is inherited
    from the Azure Function App (same key/function-auth pattern as the other
    endpoints); the function is JSON in / JSON out and never reads files.
    """
    settings = get_settings()
    llm = _get_dedup_llm(settings)

    logger.info("Dedup request received: %d rows", len(request.rows))

    try:
        return await cluster_blocks(request.rows, llm, settings=settings)
    finally:
        # Release the cached AsyncAzureOpenAI HTTP client cleanly, matching
        # the orchestrator's lifecycle handling. Mocks expose aclose only
        # when present.
        aclose = getattr(llm, "aclose", None)
        if callable(aclose):
            try:
                await aclose()
            except Exception as exc:  # noqa: BLE001
                logger.info("Dedup LLM aclose failed (non-fatal): %s", exc)


@router.post("/api/dedup/file")
async def dedup_file(
    file: UploadFile = File(..., description="XLSX file of address-gated candidate rows"),
) -> StreamingResponse:
    """Dedup adjudicator that takes an XLSX upload and returns an XLSX.

    The spreadsheet uses the same fields as the /api/dedup/cluster-block
    request rows — it accepts the human SAP headers emitted by /enrich/file
    ("Customer", "Name 1", "Street 1", …) or the snake_case DedupRow field
    names. Each data row becomes one DedupRow, the rows are clustered exactly
    as the JSON endpoint does, and the uploaded sheet is returned with the
    cluster-assignment columns appended (one row per input row).
    """
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file must be an .xlsx (or .xlsm) workbook.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    headers, row_dicts = _parse_xlsx(contents)
    rows = _rows_to_dedup_rows(row_dicts)

    settings = get_settings()
    llm = _get_dedup_llm(settings)

    logger.info(
        "Dedup file request received: %s, %d rows", filename, len(rows)
    )

    try:
        response = await cluster_blocks(rows, llm, settings=settings)
    finally:
        aclose = getattr(llm, "aclose", None)
        if callable(aclose):
            try:
                await aclose()
            except Exception as exc:  # noqa: BLE001
                logger.info("Dedup LLM aclose failed (non-fatal): %s", exc)

    output_bytes = _build_dedup_xlsx(
        headers, row_dicts, rows, response, source_contents=contents
    )

    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "records"
    out_name = f"{stem}_dedup.xlsx"

    # The same facts as the "Run" sheet, on the response itself, so a caller
    # piping the body to a file can tell which configuration produced it
    # without opening the workbook.
    run = dict(_dedup_run_metadata(response))
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{out_name}"',
            "X-Dedup-Prompt-Version": run["prompt_version"],
            "X-Dedup-Model": run["model"],
            "X-Dedup-V2-Blocking": run["DEDUP_V2_BLOCKING"],
            "X-Dedup-V2-Name2": run["DEDUP_V2_NAME2"],
            "X-Dedup-V2-Id-Conflict": run["DEDUP_V2_ID_CONFLICT"],
            "X-Dedup-Fixture-Cache": run["fixture_cache"],
        },
    )


# ---------------------------------------------------------------------------
# Scoring + golden-record election (Phase 2 "Pass 3")
# ---------------------------------------------------------------------------


@router.post("/api/dedup/score", response_model=ScoringResponse)
async def dedup_score(request: ScoringRequest) -> ScoringResponse:
    """Deterministic scoring + golden-record election.

    Separate from /api/dedup/cluster-block on purpose: clustering and
    election have different inputs, cadences, and cost profiles — election
    is pure arithmetic over dedup/weights.json and can be re-run on retuned
    weights without paying for LLM adjudication again. No LLM is involved.

    An empty rows list returns 200 with an empty result and zeroed summary;
    a duplicated row_id returns 400 (a duplicated BP number means a broken
    upstream join; scoring it would double-elect).
    """
    request_start = time.perf_counter()
    logger.info("Scoring request received: %d rows", len(request.rows))

    # Optional weights override — same all-or-nothing semantics as the
    # /api/dedup/score/file Weights sheet: a valid override applies wholesale,
    # a malformed one is ignored wholesale with a warning (never a hard error).
    weights = None
    request_warnings: list[str] = []
    if request.weights is not None:
        weights, reason = coerce_weights(
            request.weights, load_weights(), source="Weights payload"
        )
        if weights is None:
            request_warnings.append(reason)
            logger.warning("scoring request: %s", reason)

    try:
        results = elect_golden_records(request.rows, weights)
    except DuplicateRowIdError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Duplicate row_id(s) in request: {', '.join(exc.row_ids)}",
        ) from exc

    summary = build_summary(results, warnings=request_warnings)
    issues = detect_dedup_issues(request.rows, results)
    logger.info(
        "scoring_request",
        extra={
            "summary": summary.model_dump(),
            "issues": len(issues),
            "total_latency_ms": int((time.perf_counter() - request_start) * 1000),
        },
    )
    return ScoringResponse(rows=results, summary=summary, issues=issues)


@router.post("/api/dedup/approve", response_model=ApprovalResponse)
async def dedup_approve(request: ApprovalRequest) -> ApprovalResponse:
    """Record a human's approve/reject decision on one proposed cluster.

    Stateless: the caller submits the scored rows, the decision is applied to
    the named cluster_id (approval_status set; on "approved" the proposed winner
    is promoted into the golden fields), and the updated rows are echoed back.
    Persistence is intentionally out of scope — a durable approval store is a
    future step. Phase 3 consumes ONLY rows with approval_status="approved" or
    election_status="unique".
    """
    logger.info(
        "dedup_approve: cluster=%s decision=%s approver=%s rows=%d",
        request.cluster_id, request.decision, request.approver, len(request.rows),
    )
    try:
        rows, updated = apply_approval(
            request.rows, request.cluster_id, request.decision
        )
    except ClusterNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    return ApprovalResponse(
        cluster_id=request.cluster_id,
        decision=request.decision,
        approver=request.approver,
        updated_row_ids=updated,
        rows=rows,
    )


@router.post("/api/dedup/score/file")
async def dedup_score_file(
    file: UploadFile = File(..., description="Scoring XLSX (dedup test data with score_* columns)"),
) -> StreamingResponse:
    """Scoring + election endpoint that takes an XLSX upload and returns it filled.

    The uploaded workbook is edited IN PLACE with openpyxl — the "Weights"
    sheet and every original column survive untouched; only the empty
    score_*/derived/election columns are written (located by header name,
    never by index). A parseable Weights sheet overrides dedup/weights.json
    wholesale; a broken one is ignored wholesale with a warning in the
    logged summary. Rows with a blank Customer cell are skipped and counted
    in summary.errors.
    """
    request_start = time.perf_counter()
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file must be an .xlsx (or .xlsm) workbook.",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        output_bytes, summary = score_workbook(contents)
    except DuplicateRowIdError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Duplicate Customer number(s) in file: {', '.join(exc.row_ids)}",
        ) from exc
    except ScoringFileError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    logger.info(
        "scoring_request",
        extra={
            "summary": summary.model_dump(),
            # "filename" is a reserved LogRecord attribute — use upload_name.
            "upload_name": filename,
            "total_latency_ms": int((time.perf_counter() - request_start) * 1000),
        },
    )

    stem = filename.rsplit("/", 1)[-1].rsplit(".", 1)[0] or "records"
    out_name = f"{stem}_scored.xlsx"
    return StreamingResponse(
        io.BytesIO(output_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@router.get("/diag/llm")
async def diag_llm() -> dict:
    """Diagnostic: make one LLM call and return the raw outcome.

    Use this on Azure when you can't see logs — the actual exception
    string is returned in the HTTP response body.
    """
    import os

    env_snapshot = {
        "AZURE_OPENAI_ENDPOINT": os.getenv("AZURE_OPENAI_ENDPOINT", "<unset>"),
        "AZURE_OPENAI_DEPLOYMENT": os.getenv("AZURE_OPENAI_DEPLOYMENT", "<unset>"),
        "AZURE_OPENAI_API_KEY_present": bool(os.getenv("AZURE_OPENAI_API_KEY")),
        "AZURE_OPENAI_API_KEY_length": len(os.getenv("AZURE_OPENAI_API_KEY", "")),
    }
    try:
        from llm.openai_client import call_openai
        raw = await call_openai(
            system_prompt="Return valid JSON only.",
            user_prompt='Return {"ok": true}',
            max_tokens=50,
        )
        return {"status": "ok", "raw": raw, "env": env_snapshot}
    except Exception as exc:  # noqa: BLE001
        return {
            "status": "failed",
            "error_type": type(exc).__name__,
            "error_message": str(exc),
            "env": env_snapshot,
        }


@router.get("/diag/dedup-llm")
async def diag_dedup_llm() -> dict:
    """Diagnostic: make one real dedup adjudication call and return the raw
    outcome. Use this when /api/dedup/cluster-block returns everything as
    manual_review with errors>0 — it surfaces the actual Azure error string
    (e.g. an unsupported api-version or reasoning_effort rejection) that the
    per-block handler swallows to keep the batch alive.
    """
    import os

    env_snapshot = {
        "AZURE_OPENAI_ENDPOINT": os.getenv("AZURE_OPENAI_ENDPOINT", "<unset>"),
        "AOAI_DEPLOYMENT_DEDUP": os.getenv("AOAI_DEPLOYMENT_DEDUP", "<unset>"),
        "AOAI_API_VERSION_DEDUP": os.getenv("AOAI_API_VERSION_DEDUP", "<unset>"),
        "DEDUP_REASONING_EFFORT": os.getenv("DEDUP_REASONING_EFFORT", "<unset>"),
        "AZURE_OPENAI_API_KEY_present": bool(os.getenv("AZURE_OPENAI_API_KEY")),
    }
    llm = DedupLLM(get_settings())
    try:
        result = await llm.adjudicate(
            "Return valid JSON only.",
            'Return STRICT JSON: {"ok": true}',
            max_tokens=200,
        )
        return {
            "status": "ok" if result.error is None else "failed",
            "raw": result.raw,
            "error": result.error,
            "model_version": result.model_version,
            "prompt_tokens": result.prompt_tokens,
            "completion_tokens": result.completion_tokens,
            "api_version": llm._api_version,
            "reasoning_effort_used": llm._use_reasoning_effort,
            "env": env_snapshot,
        }
    finally:
        await llm.aclose()


@router.get("/tiers", response_model=TierConfigResponse)
async def get_tier_config() -> TierConfigResponse:
    """Return current tier thresholds and configuration."""
    settings = get_settings()
    # FIX(Bug 1): single ROR threshold for all record types
    return TierConfigResponse(
        ror_confidence_threshold=settings.ror_confidence_threshold,
        fuzzy_match_threshold=settings.fuzzy_match_threshold,
        max_page_content_chars=settings.max_page_content_chars,
        page_fetch_timeout_seconds=settings.page_fetch_timeout_seconds,
        default_max_concurrency=settings.default_max_concurrency,
        serp_provider="serpapi" if settings.serpapi_key else "duckduckgo",
        mock_mode=settings.mock_external_calls,
    )
