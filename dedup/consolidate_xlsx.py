"""XLSX transport for the company-code / sales-org consolidation stage.

The uploaded workbook is edited IN PLACE with openpyxl — every other sheet
and every original column survive untouched (it never round-trips through
pandas.to_excel, which drops other sheets). The two consolidated columns are
located or appended BY HEADER NAME via the same ``_ensure_column`` the
scoring writeback uses, so a re-run overwrites in place instead of appending
a second copy.

Unlike the JSON transport this one always sees the whole file, so it is
always safe with respect to the batching invariant: it is the endpoint to run
once over the complete extract, before any batching.
"""

from __future__ import annotations

import io
import logging
from typing import Dict, List, Optional, Tuple

from dedup.consolidate import (
    COMPANY_CODE_CONSOLIDATED_HEADER,
    COMPANY_CODE_HEADER,
    CUSTOMER_HEADER,
    Consolidation,
    ConsolidationSummary,
    SALES_ORG_CONSOLIDATED_HEADER,
    SALES_ORG_HEADER,
    blank_customer_warning,
    missing_column_warnings,
)
# Reused verbatim from the scoring writeback rather than re-implemented: the
# column this stage appends is the column that stage looks for.
from dedup.scoring_xlsx import _ensure_column, _header_columns, _is_blank, _norm

logger = logging.getLogger(__name__)


class ConsolidateFileError(ValueError):
    """The uploaded workbook cannot be consolidated (maps to HTTP 400)."""


def _select_sheet(wb, sheet: Optional[str]):
    """The named sheet, or the first worksheet."""
    if sheet:
        for ws in wb.worksheets:
            if ws.title == sheet:
                return ws
        raise ConsolidateFileError(
            f"No sheet named {sheet!r}. Sheets: {', '.join(wb.sheetnames)}."
        )
    if not wb.worksheets:
        raise ConsolidateFileError("The uploaded workbook has no worksheets.")
    return wb.worksheets[0]


def consolidate_workbook(
    contents: bytes, sheet: Optional[str] = None
) -> Tuple[bytes, ConsolidationSummary]:
    """Append the two consolidated columns to one sheet of a workbook.

    Returns (workbook bytes, summary). Row count is preserved: a data row is
    only skipped when EVERY cell in it is blank (openpyxl's ``max_row`` runs
    past the data on files that have been opened in Excel), and such a row is
    counted neither in ``rows_in`` nor in ``rows_out``.

    Raises ConsolidateFileError for a workbook that cannot be read or has no
    Customer column on the selected sheet.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ConsolidateFileError(
            "openpyxl is required to consolidate XLSX uploads."
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(contents))
    except Exception as exc:  # noqa: BLE001 - any parse failure is a bad upload
        raise ConsolidateFileError(
            f"Could not read uploaded file as XLSX: {exc}"
        ) from exc

    ws = _select_sheet(wb, sheet)
    columns: Dict[str, int] = _header_columns(ws)

    customer_col = columns.get(_norm(CUSTOMER_HEADER))
    if customer_col is None:
        raise ConsolidateFileError(
            f"Sheet {ws.title!r} has no '{CUSTOMER_HEADER}' column."
        )
    company_code_col = columns.get(_norm(COMPANY_CODE_HEADER))
    sales_org_col = columns.get(_norm(SALES_ORG_HEADER))

    # Pass 1 — fold every data row into the per-customer value sets.
    state = Consolidation()
    row_keys: List[Tuple[int, str]] = []
    for ws_row, values in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(_is_blank(v) for v in values):
            continue  # fully blank (or formula-only) trailing row
        cell = lambda col: values[col - 1] if col and col <= len(values) else None
        row_keys.append((ws_row, state.observe(
            cell(customer_col), cell(company_code_col), cell(sales_org_col)
        )))

    # Pass 2 — write both strings onto every row of every group.
    cc_out = _ensure_column(ws, columns, COMPANY_CODE_CONSOLIDATED_HEADER)
    so_out = _ensure_column(ws, columns, SALES_ORG_CONSOLIDATED_HEADER)
    # An empty group writes "" — which is what XLSX calls a blank cell, and
    # what openpyxl therefore reads back as None. That is the format's own
    # normalisation, not a different value: ``split_consolidated`` counts a
    # blank cell and an empty string identically (zero values).
    for ws_row, key in row_keys:
        company_codes, sales_orgs = state.resolve(key)
        ws.cell(row=ws_row, column=cc_out, value=company_codes)
        ws.cell(row=ws_row, column=so_out, value=sales_orgs)

    warnings = missing_column_warnings(
        has_company_code=company_code_col is not None,
        has_sales_org=sales_org_col is not None,
    )
    warnings += blank_customer_warning(state.errors)
    for warning in warnings:
        logger.warning("consolidate file: %s", warning)

    summary = state.summary(rows_out=len(row_keys), warnings=warnings)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), summary
