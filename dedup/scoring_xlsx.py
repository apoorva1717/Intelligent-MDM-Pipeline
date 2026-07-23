"""XLSX in/out for the scoring + golden-record election endpoint.

Fills the empty score_*/election columns of the scoring workbook IN PLACE
with openpyxl — the "Weights" sheet and every original column survive
untouched (never round-trips through pandas.to_excel, which drops other
sheets). All columns are located by header name, never by index.
"""

from __future__ import annotations

import io
import logging
from typing import Dict, List, Optional, Tuple

from dedup.scoring import (
    ScoringRow,
    ScoringSummary,
    build_summary,
    derived_counts,
    detect_issues,
    elect_golden_records,
    load_weights,
)

logger = logging.getLogger(__name__)

WEIGHTS_SHEET_NAME = "Weights"
ISSUES_SHEET_NAME = "Issues"
ISSUES_COLUMNS = ("row_id", "cluster_id", "issue_type", "detail")

# Input column header -> ScoringRow field (matched via _norm, so case /
# whitespace / punctuation variants of these headers still resolve).
INPUT_HEADERS: Dict[str, str] = {
    "Customer": "row_id",
    "Account group": "account_group",  # note the space in the column name
    "Sales_Order_Last_Used": "last_order_year",
    # G1: despite the "_Total_" header, this is interpreted as the count of
    # sales orders WITHIN the last-used year (Bernd's year-priority rule). The
    # count only differentiates records sharing the most-recent year in a
    # cluster. OPEN ITEM P2-21: confirm the click-report column layout actually
    # supplies a within-year count here (not a lifetime total) before go-live.
    "Sales_Order_Total_Count": "orders_in_last_used_year",
    "Sales_Order_Partner_Last_Used": "partner_last_order_year",
    # G1: within-year partner count (same P2-21 caveat as above).
    "Sales_Order_Partner_Total_Count": "partner_orders_in_last_used_year",
    "Equipment_Total_Count": "equipment_count",
    "SleepingCustomer": "sleeping_band",
    "CustomerStatus": "customer_status",
    "Company_Code_Consolidated": "company_code_consolidated",
    "Sales_Org_Consolidated": "sales_org_consolidated",
}

# The 8 Salesforce id slots, in order.
SF_ID_HEADERS: List[str] = [
    "SF_ID_Biosystems", "SF_ID_AXS",
    "SF_ID_3", "SF_ID_4", "SF_ID_5", "SF_ID_6", "SF_ID_7", "SF_ID_8",
]

# score_breakdown key -> output column header.
BREAKDOWN_COLUMNS: Dict[str, str] = {
    "sales_order_last_used": "score_SalesOrderLastUsed",
    "sales_order_count": "score_SalesOrderCount",
    "sales_order_partner_last_used": "score_SalesOrderPartnerLastUsed",
    "sales_order_partner_count": "score_SalesOrderPartnerCount",
    "equipment_count": "score_EquipmentCount",
    "sleeping_customer": "score_SleepingCustomer",
    "customer_status": "score_CustomerStatus",
    "account_group": "score_AccountGroup",
    "company_code_count": "score_CompanyCodeCount",
    "combined_presence_bonus": "score_CombinedPresence",
    "salesforce_instance_count": "score_SalesforceInstances",
}

DERIVED_COLUMNS = ("Company_Code_Count", "Sales_Org_Count", "Salesforce_Instance_Count")
ELECTION_COLUMNS = (
    "is_golden_record", "golden_record_id", "proposed_golden_id",
    "election_status", "approval_status",
)


class ScoringFileError(ValueError):
    """The uploaded workbook cannot be scored (maps to HTTP 400)."""


def _norm(name: object) -> str:
    """Normalise a header for tolerant by-name matching (same collapsing as
    api.routes._norm_header, duplicated here to avoid a circular import)."""
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _is_blank(value: object) -> bool:
    return value is None or not str(value).strip()


# ---------------------------------------------------------------------------
# Weights sheet override
# ---------------------------------------------------------------------------

def _parse_weights_sheet(ws, expected: dict) -> Tuple[Optional[dict], Optional[str]]:
    """Parse a Weights sheet into a weights dict, or explain why not.

    Matched by EXACT (Criterion, Band) label pairs against ``expected``
    (dedup/weights.json). All-or-nothing: any missing pair or non-numeric
    Points cell rejects the WHOLE sheet — a half-applied retune is worse
    than none. Returns (weights, None) or (None, reason).
    """
    cells: Dict[Tuple[str, str], object] = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        criterion = raw[0] if len(raw) > 0 else None
        band = raw[1] if len(raw) > 1 else None
        points = raw[2] if len(raw) > 2 else None
        if _is_blank(criterion) and _is_blank(band):
            continue
        cells[(str(criterion).strip(), str(band).strip())] = points

    parsed: dict = {}
    for criterion, bands in expected.items():
        parsed[criterion] = {}
        for band in bands:
            key = (criterion, band)
            if key not in cells:
                return None, (
                    f"Weights sheet ignored: missing (criterion, band) pair "
                    f"({criterion!r}, {band!r}); using dedup/weights.json"
                )
            points = cells[key]
            if isinstance(points, bool) or not isinstance(points, (int, float)):
                try:
                    points = float(str(points).strip())
                except (TypeError, ValueError):
                    return None, (
                        f"Weights sheet ignored: non-numeric Points for "
                        f"({criterion!r}, {band!r}); using dedup/weights.json"
                    )
            parsed[criterion][band] = int(points)
    return parsed, None


# ---------------------------------------------------------------------------
# Workbook scoring
# ---------------------------------------------------------------------------

def _find_data_sheet(wb):
    """First sheet whose header row carries a Customer column."""
    for ws in wb.worksheets:
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if any(_norm(h) == "customer" for h in headers if h is not None):
            return ws
    raise ScoringFileError("No sheet with a 'Customer' header column found.")


def _header_columns(ws) -> Dict[str, int]:
    """Normalised header -> 1-based column index (first occurrence wins)."""
    columns: Dict[str, int] = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, header in enumerate(header_row, start=1):
        if header is None:
            continue
        key = _norm(header)
        if key and key not in columns:
            columns[key] = idx
    return columns


def _ensure_column(ws, columns: Dict[str, int], header: str) -> int:
    """Column index for ``header``, appending a new header cell if absent."""
    key = _norm(header)
    if key in columns:
        return columns[key]
    idx = max(ws.max_column, max(columns.values(), default=0)) + 1
    ws.cell(row=1, column=idx, value=header)
    columns[key] = idx
    return idx


def _cluster_columns(columns: Dict[str, int]) -> Tuple[Optional[int], Optional[int]]:
    """(routing column, cluster-id column) for this workbook.

    The production pipeline is /enrich/file -> /api/dedup/file -> here, and
    the dedup stage appends "Routing" + "Cluster ID". Prefer that pair; fall
    back to the test fixture's expected_routing/expected_cluster. The pairs
    are never mixed — a routing value must be interpreted against its own
    cluster key.
    """
    if _norm("Routing") in columns and _norm("Cluster ID") in columns:
        return columns[_norm("Routing")], columns[_norm("Cluster ID")]
    return columns.get(_norm("expected_routing")), columns.get(_norm("expected_cluster"))


def _cluster_id_from_cells(routing: object, cluster: object) -> Optional[str]:
    """Map a (routing, cluster key) cell pair onto the cluster_id semantics.

    "cluster" and "manual_review" both use the cluster key (adjudicator
    semantics: for manual_review, membership is certain — only a merge was
    uncertain). "unique" or anything unrecognised -> no cluster. The dedup
    stage emits integer cluster ids (possibly 3.0 after an Excel round-trip);
    normalise those so cluster grouping keys stay stable.
    """
    routing_text = str(routing).strip().casefold() if routing is not None else ""
    if routing_text not in ("cluster", "manual_review"):
        return None
    if _is_blank(cluster):
        return None
    if isinstance(cluster, float) and cluster.is_integer():
        cluster = int(cluster)
    return str(cluster).strip()


def score_workbook(contents: bytes) -> Tuple[bytes, ScoringSummary]:
    """Score + elect over an uploaded scoring workbook, filled in place.

    Returns (workbook bytes, summary). Raises ScoringFileError for an
    unusable file and DuplicateRowIdError for a repeated Customer number.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ScoringFileError("openpyxl is required to score XLSX uploads.") from exc

    try:
        wb = load_workbook(io.BytesIO(contents))
    except Exception as exc:  # noqa: BLE001 - any parse failure is a bad upload
        raise ScoringFileError(f"Could not read uploaded file as XLSX: {exc}") from exc

    ws = _find_data_sheet(wb)
    columns = _header_columns(ws)

    # Weights: a parseable Weights sheet overrides weights.json wholesale;
    # a broken one is ignored wholesale with a warning in the summary.
    default_weights = load_weights()
    weights = default_weights
    request_warnings: List[str] = []
    weights_ws = next(
        (s for s in wb.worksheets if s.title.strip().casefold() == "weights"), None
    )
    if weights_ws is not None:
        sheet_weights, reason = _parse_weights_sheet(weights_ws, default_weights)
        if sheet_weights is not None:
            weights = sheet_weights
        else:
            request_warnings.append(reason)
            logger.warning("scoring file: %s", reason)

    input_cols = {
        field: columns.get(_norm(header))
        for header, field in INPUT_HEADERS.items()
    }
    if input_cols["row_id"] is None:
        raise ScoringFileError("The data sheet has no 'Customer' column.")
    sf_cols = [columns.get(_norm(h)) for h in SF_ID_HEADERS]
    routing_col, cluster_col = _cluster_columns(columns)
    # Adjudication merge confidence persisted by clustering — gates
    # low-confidence merges to manual_review at election time (no LLM re-run).
    confidence_col = columns.get(_norm("Confidence"))
    # Adjudication reasoning — surfaced as a verdict_contradiction issue.
    reasoning_col = columns.get(_norm("Reasoning"))

    read_cols = (
        [c for c in input_cols.values() if c]
        + [c for c in sf_cols if c]
        + [c for c in (routing_col, cluster_col, confidence_col, reasoning_col) if c]
    )

    rows: List[ScoringRow] = []
    row_indices: List[int] = []  # worksheet row number per parsed ScoringRow
    errors = 0
    for ws_row in range(2, ws.max_row + 1):
        cell = lambda col: ws.cell(row=ws_row, column=col).value if col else None
        if all(_is_blank(cell(c)) for c in read_cols):
            continue  # fully blank (or formula-only) trailing row
        customer = cell(input_cols["row_id"])
        if _is_blank(customer):
            errors += 1
            logger.warning("scoring file: row %d has a blank Customer — skipped", ws_row)
            continue
        payload = {
            field: cell(col)
            for field, col in input_cols.items()
            if field != "row_id"
        }
        rows.append(ScoringRow(
            row_id=str(customer).strip(),
            cluster_id=_cluster_id_from_cells(cell(routing_col), cell(cluster_col)),
            routing=cell(routing_col),
            confidence=cell(confidence_col),
            reasoning=cell(reasoning_col),
            salesforce_ids=[cell(c) for c in sf_cols],
            **payload,
        ))
        row_indices.append(ws_row)

    results = elect_golden_records(rows, weights)
    summary = build_summary(results, errors=errors, warnings=request_warnings)
    issues = detect_issues(rows, results)

    # Writeback — locate/append every output column by header name.
    breakdown_cols = {
        key: _ensure_column(ws, columns, header)
        for key, header in BREAKDOWN_COLUMNS.items()
    }
    final_col = _ensure_column(ws, columns, "score_final")
    cc_col, so_col, sf_count_col = (
        _ensure_column(ws, columns, h) for h in DERIVED_COLUMNS
    )
    golden_col, golden_id_col, proposed_col, status_col, approval_col = (
        _ensure_column(ws, columns, h) for h in ELECTION_COLUMNS
    )
    weights_ver_col = _ensure_column(ws, columns, "scored_with_weights_version")

    for ws_row, row, result in zip(row_indices, rows, results):
        for key, col in breakdown_cols.items():
            ws.cell(row=ws_row, column=col, value=result.score_breakdown[key])
        # Plain value; equals the sum of the written score_* cells by
        # construction (score = sum(score_breakdown.values())).
        ws.cell(row=ws_row, column=final_col, value=result.score)
        company_codes, sales_orgs, sf_instances = derived_counts(row)
        ws.cell(row=ws_row, column=cc_col, value=company_codes)
        ws.cell(row=ws_row, column=so_col, value=sales_orgs)
        ws.cell(row=ws_row, column=sf_count_col, value=sf_instances)
        # A manual_review row leaves is_golden_record / golden_record_id EMPTY —
        # nobody filtering is_golden_record alone may act on an unreviewed row.
        # The computed winner survives in proposed_golden_id.
        is_mr = result.election_status == "manual_review"
        ws.cell(row=ws_row, column=golden_col,
                value=None if is_mr else result.is_golden_record)
        ws.cell(row=ws_row, column=golden_id_col,
                value=None if is_mr else result.golden_record_id)
        ws.cell(row=ws_row, column=proposed_col, value=result.proposed_golden_id)
        ws.cell(row=ws_row, column=status_col, value=result.election_status)
        ws.cell(row=ws_row, column=approval_col, value=result.approval_status)
        ws.cell(row=ws_row, column=weights_ver_col,
                value=result.scored_with_weights_version)

    # Second sheet: the potential-inconsistency list (Bernd's feedback loop).
    # Rebuilt fresh each run — never round-tripped through pandas — so the
    # Weights sheet and every original column survive untouched.
    if ISSUES_SHEET_NAME in wb.sheetnames:
        del wb[ISSUES_SHEET_NAME]
    issues_ws = wb.create_sheet(ISSUES_SHEET_NAME)
    issues_ws.append(list(ISSUES_COLUMNS))
    for issue in issues:
        issues_ws.append(
            [issue.row_id, issue.cluster_id, issue.issue_type, issue.detail]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), summary
