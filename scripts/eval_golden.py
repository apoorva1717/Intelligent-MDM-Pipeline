"""Run the golden set through the pipeline and grade it against the reference.

The golden set is `docs/SAMPLE_DATA/testall100_SOLVED_REFERENCE_v1*.xlsx`: 99
records, each raw row paired with the enriched row a reviewer certified for it,
plus the rules for grading each column (`tools/golden_eval`).

    python scripts/eval_golden.py --out-dir logs/golden
    python scripts/eval_golden.py --frozen          # no network; cache only
    python scripts/eval_golden.py --limit 10        # smoke run

Four artefacts, all under ``--out-dir``:

* ``golden_input.xlsx``   the INPUT rows, as fed to the pipeline
* ``golden_enriched.xlsx``the run's own output workbook
* ``golden_eval.json``    every graded cell plus the summary
* ``golden_eval.md``      the report: what failed, and on which column

**The INPUT rows come from the reference workbook, not from
``test-all-100-original.xlsx``.** That file is column-shifted on 75 of its 99
rows — City holds an issue-code list and a date, Region holds the eval-set
label — and the reference's Method sheet says its INPUT rows were rebuilt from
the source corpus for exactly that reason. ``--input-workbook`` overrides this
if the original is ever repaired; ``scripts/eval_golden.py --check-original``
re-runs the corruption check that established it.
"""

from __future__ import annotations

import argparse
import asyncio
import io
import json
import os
import re
import sys
import time
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

from tools.golden_eval import (  # noqa: E402
    MISMATCH,
    by_column,
    compare,
    load_reference,
    normalise,
)

DEFAULT_REFERENCE = (
    "docs/SAMPLE_DATA/testall100_SOLVED_REFERENCE_v1 (1) (1).xlsx"
)
DEFAULT_ORIGINAL = "docs/SAMPLE_DATA/test-all-100-original.xlsx"


def check_original(path: Path) -> dict:
    """Count the rows in the raw upload whose columns are shifted.

    The signature is unambiguous: City holding an issue-code list (``G1-...``)
    or a date, or Region holding an eval-set label (``S1``..``S5``). Reported
    rather than repaired — rebuilding a shifted row means guessing which
    column each value came from, and the reference already did that work
    against the source corpus.
    """
    from openpyxl import load_workbook

    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    header = [normalise(c) for c in next(rows)]
    city, region = header.index("City"), header.index("Region")
    total = shifted = 0
    examples: list[str] = []
    for raw in rows:
        if raw is None or not any(c is not None for c in raw):
            continue
        total += 1
        city_value, region_value = normalise(raw[city]), normalise(raw[region])
        if (
            re.search(r"G\d-[A-Z]+-\d|\d{4}-\d\d-\d\d", city_value)
            or re.fullmatch(r"S[1-9]", region_value)
        ):
            shifted += 1
            if len(examples) < 3:
                examples.append(f"City={city_value!r} Region={region_value!r}")
    return {
        "rows": total, "shifted": shifted, "clean": total - shifted,
        "examples": examples,
    }


def write_input_workbook(reference, path: Path, limit: int | None) -> int:
    """Write the reference's INPUT rows as a workbook the pipeline can parse.

    The reference carries them in the pipeline's own output schema, whose
    headers `_rows_to_records` already normalises onto the model's input
    fields, so the sheet is written as-is rather than remapped here — one
    header vocabulary, not two.
    """
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = "Sheet1"
    columns = list(reference.columns)
    sheet.append(columns)
    customers = list(reference.inputs)
    if limit:
        customers = customers[:limit]
    for customer in customers:
        row = reference.inputs[customer]
        sheet.append([row.get(column) for column in columns])
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return len(customers)


def produced_rows(results, headers: list[str]) -> dict[str, dict]:
    """Map customer -> the run's output row, keyed by output column name."""
    from api.output_columns import RESPONSE_COLUMNS

    field_by_header = {header: field for field, header in RESPONSE_COLUMNS.items()}
    out: dict[str, dict] = {}
    for result in results:
        row = {}
        for header in headers:
            field = field_by_header.get(header)
            if field is not None:
                row[header] = getattr(result, field, None)
        customer = normalise(getattr(result, "record_id", None))
        out[customer] = row
    return out


def render_report(summary: dict, results, reference, meta: dict) -> str:
    """The human-readable report: the score, then every failing cell."""
    lines: list[str] = []
    add = lines.append

    add("# Golden-set evaluation")
    add("")
    add(f"- reference: `{meta['reference']}`")
    add(f"- run: {meta['records']} records in {meta['elapsed']:.0f}s"
        f"{' (CACHE_FROZEN)' if meta['frozen'] else ''}")
    add(f"- graded columns: {len(reference.graded_columns)} of "
        f"{len(reference.columns)} — the rest the reference declares `skip`")
    add("")
    add("## Score")
    add("")
    add("| | n | |")
    add("|---|---:|---|")
    add(f"| records passed | **{summary['records_passed']}** / "
        f"{summary['records_produced']} | "
        f"{summary['record_accuracy']:.1%} |")
    add(f"| cells matched | **{summary['cells_matched']}** / "
        f"{summary['cells_graded']} | {summary['cell_accuracy']:.1%} |")
    if summary["records_missing"]:
        add(f"| records the run did not produce | "
            f"{len(summary['records_missing'])} | "
            f"{', '.join(summary['records_missing'][:5])} |")
    add("")

    columns = [c for c in by_column(results) if c["failed"]]
    if columns:
        add("## Where it fails, by column")
        add("")
        add("| column | failed | graded | accuracy |")
        add("|---|---:|---:|---:|")
        for entry in columns:
            add(f"| {entry['column']} | **{entry['failed']}** | "
                f"{entry['graded']} | {entry['accuracy']:.0%} |")
        add("")

    failures = [r for r in results if r.verdict == MISMATCH]
    if failures:
        add("## Every failing cell")
        add("")
        by_record: dict[str, list] = {}
        for failure in failures:
            by_record.setdefault(failure.customer, []).append(failure)
        for customer in sorted(by_record):
            add(f"### {customer}")
            add("")
            add("| column | expected | produced |")
            add("|---|---|---|")
            for failure in by_record[customer]:
                add(f"| {failure.column} | `{failure.expected or '(empty)'}` "
                    f"| `{failure.actual or '(empty)'}` |")
            add("")

    if summary["orphan_notes"]:
        add("## Cell Notes that grade nothing")
        add("")
        for customer, column, why in summary["orphan_notes"]:
            add(f"- `{customer}` / `{column}` — {why}")
        add("")
    return "\n".join(lines)


async def _main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reference", default=DEFAULT_REFERENCE)
    parser.add_argument("--out-dir", default="logs/golden")
    parser.add_argument("--input-workbook", default=None,
                        help="Feed this workbook instead of the reference's "
                             "own INPUT rows.")
    parser.add_argument("--concurrency", type=int, default=5)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--frozen", action="store_true",
                        help="CACHE_FROZEN=true — answer from the evidence "
                             "cache or record evidence-unavailable-frozen, "
                             "never the network.")
    parser.add_argument("--check-original", action="store_true",
                        help="Report the column shift in the raw upload and "
                             "exit.")
    args = parser.parse_args()

    if args.check_original:
        report = check_original(_ROOT / DEFAULT_ORIGINAL)
        print(json.dumps(report, indent=2))
        return

    if args.frozen:
        os.environ["CACHE_FROZEN"] = "true"

    out_dir = _ROOT / args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    reference = load_reference(str(_ROOT / args.reference))
    print(f"reference: {len(reference.expected)} expected rows, "
          f"{len(reference.graded_columns)} graded columns, "
          f"{len(reference.notes)} cell notes")
    if reference.orphan_notes:
        print(f"  WARNING: {len(reference.orphan_notes)} cell notes grade "
              f"nothing (unknown customer or column)")

    input_path = Path(args.input_workbook) if args.input_workbook else (
        out_dir / "golden_input.xlsx"
    )
    if not args.input_workbook:
        count = write_input_workbook(reference, input_path, args.limit)
        print(f"input:     {count} rows -> {input_path}")

    from api.models import EnrichmentOptions  # noqa: E402
    from api.routes import (  # noqa: E402
        _build_output_xlsx,
        _parse_xlsx,
        _rows_to_records,
    )
    from config import Settings  # noqa: E402
    from enrichment.orchestrator import Orchestrator  # noqa: E402

    contents = input_path.read_bytes()
    headers, row_dicts = _parse_xlsx(contents)
    records = _rows_to_records(row_dicts)

    started = time.time()
    response = await Orchestrator(Settings()).enrich_batch(
        records, EnrichmentOptions(max_concurrency=args.concurrency),
    )
    elapsed = time.time() - started
    print(f"run:       {len(response.results)} records in {elapsed:.0f}s")

    enriched_path = out_dir / "golden_enriched.xlsx"
    enriched_path.write_bytes(_build_output_xlsx(
        response.results, headers, row_dicts, source_contents=contents,
    ))
    print(f"enriched:  {enriched_path}")

    produced = produced_rows(response.results, reference.columns)
    results, summary = compare(reference, produced)

    (out_dir / "golden_eval.json").write_text(json.dumps({
        "summary": summary,
        "by_column": by_column(results),
        "cells": [
            {
                "customer": r.customer, "column": r.column,
                "verdict": r.verdict, "expected": r.expected,
                "actual": r.actual, "rule": r.rule.kind,
                "rule_source": r.rule.source,
            }
            for r in results if r.verdict != "skipped"
        ],
    }, indent=1, default=str), encoding="utf-8")

    report = render_report(summary, results, reference, {
        "reference": args.reference, "records": len(response.results),
        "elapsed": elapsed, "frozen": args.frozen,
    })
    (out_dir / "golden_eval.md").write_text(report, encoding="utf-8")

    print()
    print(f"records passed: {summary['records_passed']}/"
          f"{summary['records_produced']}  ({summary['record_accuracy']:.1%})")
    print(f"cells matched:  {summary['cells_matched']}/"
          f"{summary['cells_graded']}  ({summary['cell_accuracy']:.1%})")
    print(f"report -> {out_dir / 'golden_eval.md'}")


if __name__ == "__main__":
    asyncio.run(_main())
