"""Chapter 2 (Problem Description) frequency measurements.

Produces the measured figures for `docs/thesis/ch02_SOURCE.md` §3, §5 and §6.
Read-only: no file in the repository is modified and no external call is made.

Three measurements, each over a named workbook:

1. **Issue frequency** (§3) — runs the deterministic detector
   ``enrichment.issue_detection.detect_issues`` over every data row through the
   same record-construction path the ``POST /issues`` endpoint uses
   (``_parse_xlsx`` -> ``_rows_to_records`` -> ``_present_fields`` ->
   ``detect_issues``, ``api/routes.py:603-606``). Reports totals, the
   per-record issue-count distribution, the ranked per-code frequency, and the
   mean number of distinct SAP columns implicated per record.

   The field attribution uses per-code *locators* that mirror each rule body in
   ``enrichment/issue_detection.py`` and return the specific column(s) that made
   the rule fire. Fidelity is self-checked: the script asserts that the set of
   codes the locators fire on equals the set ``detect_issues`` returns, for
   every row. Any divergence is reported and the run exits non-zero.

2. **Duplicate prevalence** (§5) — STEP A only: the deterministic
   block-and-signature collapse of ``dedup/signatures.py``. No LLM adjudication
   is performed, so what is reported is exact-signature collapse, never an
   adjudicated cluster. The source workbook carries no ``Block ID`` column, so
   every block id is the service's *fallback* derivation
   (``derive_block_id``, ``dedup/signatures.py:45-56``) over the normalised
   (country, postal code, street, house number) tuple — not the DATAshaper
   address gate that blocks the production input.

3. **Registry-identifier coverage** (§6) — counts non-empty ``ROR ID`` and
   ``LEI ID`` cells in the enriched workbook. The pre-enrichment workbook
   carries neither column, so every populated cell is an identifier the
   pipeline supplied.

Usage:
    .venv\\Scripts\\python.exe -m scripts.ch02_measure
"""

from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from api.models import EnrichmentRecord  # noqa: E402
from api.routes import _parse_xlsx, _present_fields, _rows_to_records  # noqa: E402
from dedup.models import DedupRow  # noqa: E402
from dedup.signatures import build_signatures, group_rows_by_block  # noqa: E402
from enrichment.address_processing import (  # noqa: E402
    _BARE_MARKER_RE,
    _PO_BOX_RE,
    _STREET_TYPE_WORD_RE,
    _SUITE_PATTERNS,
    _UNIVERSITY_CENTRE_RE,
    _extract_mail_code,
    _extract_sublocations,
    _looks_like_department,
    _looks_like_street,
)
from enrichment.issue_detection import (  # noqa: E402
    _ABBREV_TOKEN_RE,
    _NAME_CONTINUATION_RE,
    _ORG_IN_STREET_RE,
    _POSTAL_FORMATS,
    _REQUIRED_FIELD_CODES,
    _SAP_NAME_LIMIT,
    _norm,
    _street_signature,
    ISSUE_CATALOGUE,
    _SUBLOCATION_SLOTS,
    issue_name,
    detect_issues,
)
from enrichment.preprocess import (  # noqa: E402
    _CO_ATTN_PREFIX_RE,
    _EMAIL_RE,
    _PHONE_RE,
    _URL_RE,
    _extract_addresses,
    _has_legal_suffix,
    _is_opaque_code,
    _normalise_dba,
    _street_person_name,
    has_multiple_contacts,
)
from utils.name_slots import (  # noqa: E402
    ADJACENT_RECORD_NAME_PAIRS,
    RECORD_NAME_FIELDS,
)
from utils.text_utils import (  # noqa: E402
    country_to_iso_code,
    is_blank,
    is_granular_unit,
    is_specific_unit_construction,
    is_unit_construction,
    looks_like_university_or_research_institute,
)

ROOT = Path(__file__).resolve().parents[1]
RAW_WORKBOOK = ROOT / "PresentationTestData.xlsx"
ENRICHED_WORKBOOK = ROOT / "PresentationTestData_enriched_checked_v1.xlsx"

# SAP column name for each EnrichmentRecord field the locators report on.
COLUMN = {
    "name_1": "Name 1", "name_2": "Name 2", "name_3": "Name 3", "name_4": "Name 4",
    "name_5": "Name 5",
    "street_1": "Street 1", "street_2": "Street 2", "street_3": "Street 3",
    "street_4": "Street 4", "street_5": "Street 5",
    "house_number": "House Number", "po_box": "PO Box",
    "postal_code": "Postal Code", "region": "Region", "city": "City",
    "country_region_key": "Country/Region Key", "language_key": "Language Key",
    "tax_jurisdiction": "Tax Jurisdiction",
    "search_term_1": "Search Term 1", "search_term_2": "Search Term 2",
    "contact": "Contact", "care_of": "Care Of", "email": "Email",
}
NAME_FIELDS = list(RECORD_NAME_FIELDS)
DEPT_FIELDS = NAME_FIELDS[1:]
NAME_PAIRS = list(ADJACENT_RECORD_NAME_PAIRS)
STREET_FIELDS = ["street_1", "street_2", "street_3", "street_4", "street_5"]


def _val(record: EnrichmentRecord, field: str) -> str | None:
    return getattr(record, field)


def locate(record: EnrichmentRecord, present: set[str] | None) -> dict[str, set[str]]:
    """Return {issue code -> set of SAP columns that made it fire}.

    Each block mirrors the correspondingly-numbered rule in
    ``enrichment/issue_detection.py``; the line reference is given inline. Only
    codes that fire appear as keys.
    """
    hits: dict[str, set[str]] = defaultdict(set)

    def add(code: str, *fields: str) -> None:
        hits[code].update(COLUMN[f] for f in fields)

    # --- G1 ---------------------------------------------------------------
    # G1-CROSS-001 (issue_detection.py:226-229)
    for f in NAME_FIELDS:
        v = _val(record, f)
        if v and _extract_addresses(v)[0]:
            add("G1-CROSS-001", f)
            break
    # G1-CROSS-002 (:234-243)
    for f in STREET_FIELDS:
        v = _val(record, f)
        if not v:
            continue
        if _ORG_IN_STREET_RE.search(_UNIVERSITY_CENTRE_RE.sub(" ", v)) and not _STREET_TYPE_WORD_RE.search(v):
            add("G1-CROSS-002", f)
            break
    # G1-CROSS-003 (:247-262) — name/street sweep first, person-name fallback second
    matched = False
    for f in NAME_FIELDS + STREET_FIELDS:
        v = _val(record, f)
        if not v:
            continue
        if _EMAIL_RE.search(v) or _PHONE_RE.search(v) or _URL_RE.search(v) or _CO_ATTN_PREFIX_RE.search(v):
            add("G1-CROSS-003", f)
            matched = True
            break
    if not matched:
        for f in STREET_FIELDS:
            v = _val(record, f)
            if v and _street_person_name(v):
                add("G1-CROSS-003", f)
                break
    # G1-ADDR-001 (:266-270)
    if is_blank(record.house_number):
        for f in STREET_FIELDS:
            if _looks_like_street(_val(record, f)):
                add("G1-ADDR-001", f, "house_number")
                break
    # G1-ADDR-003 (:273-276)
    for f in STREET_FIELDS:
        v = _val(record, f)
        if v and any(pat.search(v) for pat, _ in _SUITE_PATTERNS):
            add("G1-ADDR-003", f)
            break
    # G1-ADDR-004 (:279-282)
    for f in STREET_FIELDS:
        v = _val(record, f)
        if v and _PO_BOX_RE.search(v):
            add("G1-ADDR-004", f)
            break
    # G1-ADDR-006 (:285-288)
    for f in STREET_FIELDS:
        v = _val(record, f)
        if v and _extract_mail_code(v, allow_bare=True)[1]:
            add("G1-ADDR-006", f)
            break
    # G1-ADDR-011 (:291-294)
    for f in STREET_FIELDS:
        if _looks_like_department(_val(record, f)):
            add("G1-ADDR-011", f)
            break
    # G1-NAME-001 — every adjacent pair, not only Name 1 / Name 2.
    for upper, lower in NAME_PAIRS:
        if (
            not is_blank(_val(record, upper))
            and not is_blank(_val(record, lower))
            and not _has_legal_suffix(_val(record, upper) or "")
            and _NAME_CONTINUATION_RE.search(_val(record, lower) or "")
        ):
            add("G1-NAME-001", upper, lower)
            break
    # G1-NAME-004 — a blank slot BETWEEN two populated ones. Catalogue v2
    # renamed this to "Empty field in between populated name fields": a blank
    # Name 1 is a missing organisation name (G2-VAL-001), not a gap.
    populated = [not is_blank(_val(record, f)) for f in NAME_FIELDS]
    for idx in range(1, len(NAME_FIELDS) - 1):
        if not populated[idx] and any(populated[:idx]) and any(populated[idx + 1:]):
            add("G1-NAME-004", NAME_FIELDS[idx])
            break
    # G1-NAME-013 (:312-315)
    for f in NAME_FIELDS:
        v = _val(record, f)
        if v and _is_opaque_code(v):
            add("G1-NAME-013", f)
            break

    # --- G2 ---------------------------------------------------------------
    # G2-VAL-* (:330-334)
    for field_name, code, condition in _REQUIRED_FIELD_CODES:
        if present is not None and field_name not in present:
            continue
        # No code currently carries a condition — G2-VAL-004's US-only gate was
        # removed as unsourced. The branch stays so this mirrors the detector.
        if condition is not None and not condition(record):
            continue
        if is_blank(getattr(record, field_name)):
            add(code, field_name)
    # "No department" is the whole block below Name 1 being empty.
    dept_values = [_val(record, f) for f in DEPT_FIELDS]
    name2_blank = all(is_blank(v) for v in dept_values)
    # G2-NAME-012 (:342-343)
    if looks_like_university_or_research_institute(record.name_1) and name2_blank:
        add("G2-NAME-012", "name_1", *DEPT_FIELDS)
    # G2-NAME-009 — a granular unit in any slot with no parent elsewhere.
    for i, value in enumerate(dept_values):
        if not is_granular_unit(value):
            continue
        others = [v for j, v in enumerate(dept_values) if j != i]
        if not any(
            is_specific_unit_construction(x) or is_unit_construction(x)
            for x in others
        ):
            add("G2-NAME-009", *DEPT_FIELDS)
            break
    # G2-CONTACT-008 / -009 are withdrawn in Catalogue v2 and no longer
    # emitted. Withdrawing them removed the contact-based (Tier 2A) department
    # recovery path, which is why G2-NAME-012 now sits in G6.

    # --- G3 ---------------------------------------------------------------
    # G3-NAME-003 (:381-384)
    for f in NAME_FIELDS:
        v = _val(record, f)
        if v and _normalise_dba(v)[1]:
            add("G3-NAME-003", f)
            break
    # G3-NAME-005 — same value in two adjacent slots, at any boundary.
    for upper, lower in NAME_PAIRS:
        upper_norm = _norm(_val(record, upper))
        if upper_norm and upper_norm == _norm(_val(record, lower)):
            add("G3-NAME-005", upper, lower)
            break
    # PO-box count (:393-397)
    po_fields = [f for f in STREET_FIELDS if (_val(record, f) or "") and _PO_BOX_RE.search(_val(record, f) or "")]
    po_box_count = len(po_fields)
    if not is_blank(record.po_box):
        po_box_count += 1
        po_fields = po_fields + ["po_box"]
    # G3-ADDR-005 (:400-401)
    if po_box_count >= 2:
        add("G3-ADDR-005", *po_fields)
    # G3-ADDR-012 (:409-417)
    sig_fields: dict[tuple, list[str]] = defaultdict(list)
    for idx, f in enumerate(STREET_FIELDS):
        sig = _street_signature(_val(record, f), record.house_number if idx == 0 else None)
        if sig is not None:
            sig_fields[sig].append(f)
    dup = [fs for fs in sig_fields.values() if len(fs) > 1]
    if dup:
        for fs in dup:
            add("G3-ADDR-012", *fs)
        if "street_1" in {f for fs in dup for f in fs} and not is_blank(record.house_number):
            add("G3-ADDR-012", "house_number")
    # G3-ADDR-013 (:420-424)
    real = {}
    for f in STREET_FIELDS:
        if _looks_like_street(_val(record, f)):
            real.setdefault(_norm(_val(record, f)), []).append(f)
    if len(real) >= 2:
        add("G3-ADDR-013", *[f for fs in real.values() for f in fs])
    # G3-ADDR-014 (:427-428)
    street_like = [f for f in STREET_FIELDS if _looks_like_street(_val(record, f))]
    if po_box_count >= 1 and street_like:
        add("G3-ADDR-014", *(po_fields + street_like))
    # G3-CONTACT-007 (:431-432)
    if has_multiple_contacts(record.contact):
        add("G3-CONTACT-007", "contact")

    # --- G4 ---------------------------------------------------------------
    # G4-NAME-015 (:441-443)
    populated_names = [f for f in NAME_FIELDS if _val(record, f)]
    if sum(len(_val(record, f) or "") for f in NAME_FIELDS) > _SAP_NAME_LIMIT:
        add("G4-NAME-015", *populated_names)
    # G4-ADDR-008 (:446-449)
    for f in STREET_FIELDS:
        v = _val(record, f)
        if v and _BARE_MARKER_RE.search(v):
            add("G4-ADDR-008", f)
            break
    # G4-ADDR-025 — more distinct sub-locations than Street 2..5 can hold.
    sublocations = set()
    for f in STREET_FIELDS:
        v = _val(record, f)
        if not v:
            continue
        _rest, extracted, _bare = _extract_sublocations(v)
        for kind, value in extracted.items():
            sublocations.add((kind, value.strip().lower()))
    if len(sublocations) > _SUBLOCATION_SLOTS:
        add("G4-ADDR-025", *STREET_FIELDS)

    # G4-ADDR-026 (:452-456)
    if not is_blank(record.postal_code):
        iso = country_to_iso_code(record.country_region_key)
        fmt = _POSTAL_FORMATS.get(iso) if iso else None
        if fmt and not fmt.match((record.postal_code or "").strip()):
            add("G4-ADDR-026", "postal_code", "country_region_key")
    # G4-ADDR-027 (:459-463)
    if not is_blank(record.country_region_key):
        raw = (record.country_region_key or "").strip()
        iso = country_to_iso_code(raw)
        if iso is None or raw.upper() != iso:
            add("G4-ADDR-027", "country_region_key")

    # --- G5 ---------------------------------------------------------------
    # G5-NAME-001 (:474-475)
    if record.name_1 and _ABBREV_TOKEN_RE.search(record.name_1):
        add("G5-NAME-001", "name_1")
    # G5-NAME-002 (:478-481)
    for f in ("name_2", "name_3", "name_4"):
        v = _val(record, f)
        if v and _ABBREV_TOKEN_RE.search(v):
            add("G5-NAME-002", f)
            break

    return dict(hits)


def load(path: Path):
    headers, row_dicts = _parse_xlsx(path.read_bytes())
    records = _rows_to_records(row_dicts)
    return headers, row_dicts, records, _present_fields(headers)


def measure_issues(path: Path) -> None:
    headers, row_dicts, records, present = load(path)
    print(f"dataset          : {path.name}")
    print(f"path             : {path.relative_to(ROOT).as_posix()}")
    print(f"sheet            : first (active) sheet, header row 1")
    print(f"columns in file  : {len(headers)}")
    print(f"model fields seen: {len(present)}  (drives G2-VAL-* column gating)")
    print()

    per_row = [detect_issues(r, present) for r in records]
    total = len(records)
    with_issues = sum(1 for codes in per_row if codes)
    instances = sum(len(codes) for codes in per_row)

    print("--- 3.1 Totals ---")
    print(f"total records read from the data          : {total}")
    print(f"records with >= 1 issue                   : {with_issues}  ({with_issues / total:.1%})")
    print(f"records with 0 issues                     : {total - with_issues}  ({(total - with_issues) / total:.1%})")
    print(f"total issue instances (code x record)     : {instances}")
    print(f"mean issue codes per record (all records) : {instances / total:.2f}")
    print(f"mean issue codes per affected record      : {instances / with_issues:.2f}")
    print()
    # G2-VAL-007 (Search Term 1 Missing) fires on every row whose Search Term 1
    # column is blank. Search Term 1 is a *derived* output of enrichment
    # (api/output_columns.py), so a raw extract is expected to carry it blank
    # throughout. Reported separately so the headline figure is not read as
    # "every record has a data defect".
    UNIVERSAL = "G2-VAL-007"
    ex = [[c for c in codes if c != UNIVERSAL] for codes in per_row]
    ex_with = sum(1 for codes in ex if codes)
    ex_inst = sum(len(codes) for codes in ex)
    print(f"--- 3.1b Same totals excluding {UNIVERSAL} ({issue_name(UNIVERSAL)}) ---")
    print(f"records with >= 1 other issue             : {ex_with}  ({ex_with / total:.1%})")
    print(f"records with no other issue               : {total - ex_with}  ({(total - ex_with) / total:.1%})")
    print(f"issue instances excluding {UNIVERSAL}     : {ex_inst}")
    print(f"mean per record (all)                     : {ex_inst / total:.2f}")
    print(f"mean per affected record                  : {ex_inst / ex_with:.2f}")
    ex_dist = Counter(len(codes) for codes in ex)
    ex_four = sum(c for n, c in ex_dist.items() if n >= 4)
    print(f"{'codes on record':>16} | {'records':>8} | {'% of 500':>9}")
    print(f"{'-' * 16}-+-{'-' * 8}-+-{'-' * 9}")
    for label, cnt in (
        ("0", ex_dist.get(0, 0)), ("1", ex_dist.get(1, 0)), ("2", ex_dist.get(2, 0)),
        ("3", ex_dist.get(3, 0)), ("4+", ex_four),
    ):
        print(f"{label:>16} | {cnt:>8} | {cnt / total:>8.1%}")
    print()

    print("--- 3.2 Issue-count distribution per record ---")
    dist = Counter(len(codes) for codes in per_row)
    print(f"{'codes on record':>16} | {'records':>8} | {'% of 500':>9}")
    print(f"{'-' * 16}-+-{'-' * 8}-+-{'-' * 9}")
    for n in sorted(dist):
        print(f"{n:>16} | {dist[n]:>8} | {dist[n] / total:>8.1%}")
    four_plus = sum(c for n, c in dist.items() if n >= 4)
    print(f"{'-' * 16}-+-{'-' * 8}-+-{'-' * 9}")
    for label, cnt in (
        ("1", dist.get(1, 0)), ("2", dist.get(2, 0)), ("3", dist.get(3, 0)), ("4+", four_plus),
    ):
        print(f"{label:>16} | {cnt:>8} | {cnt / total:>8.1%}")
    print()

    print("--- 3.3 Per-code frequency, ranked by records affected ---")
    freq = Counter()
    for codes in per_row:
        freq.update(codes)
    print(f"{'rank':>4} | {'code':<15} | {'grp':<3} | {'records':>7} | {'% of 500':>9} | name")
    print(f"{'-' * 4}-+-{'-' * 15}-+-{'-' * 3}-+-{'-' * 7}-+-{'-' * 9}-+-{'-' * 44}")
    for rank, (code, n) in enumerate(sorted(freq.items(), key=lambda kv: (-kv[1], kv[0])), start=1):
        print(f"{rank:>4} | {code:<15} | {code.split('-')[0]:<3} | {n:>7} | {n / total:>8.1%} | {issue_name(code)}")
    silent = [c for c in ISSUE_CATALOGUE if c not in freq]
    print()
    print(f"codes observed in this dataset : {len(freq)} of {len(ISSUE_CATALOGUE)} declared")
    print(f"codes never observed here      : {len(silent)}")
    for code in silent:
        print(f"    {code:<15} {issue_name(code)}")
    print()

    print("--- 3.4 Distinct SAP columns implicated per record ---")
    located = [locate(r, present) for r in records]
    mismatch = 0
    for codes, loc, rec in zip(per_row, located, records):
        if set(codes) != set(loc):
            mismatch += 1
            if mismatch <= 5:
                print(f"  !! locator/detector divergence on record {rec.record_id!r}: "
                      f"detector={sorted(set(codes) - set(loc))} locator={sorted(set(loc) - set(codes))}")
    print(f"locator fidelity self-check: {total - mismatch}/{total} rows agree with detect_issues")
    if mismatch:
        print("  ABORT: locator does not mirror the detector; field counts are not reportable.")
        sys.exit(1)

    fields_per = [len({c for cols in loc.values() for c in cols}) for loc in located]
    affected = [n for n in fields_per if n]
    print(f"mean distinct columns implicated, all records      : {sum(fields_per) / total:.2f}")
    print(f"mean distinct columns implicated, affected records : {sum(affected) / len(affected):.2f}")
    print(f"max distinct columns implicated on one record      : {max(fields_per)}")
    col_dist = Counter(fields_per)
    print(f"{'columns':>8} | {'records':>8}")
    print(f"{'-' * 8}-+-{'-' * 8}")
    for n in sorted(col_dist):
        print(f"{n:>8} | {col_dist[n]:>8}")
    print()
    print("most-implicated columns (records in which the column is named by >=1 code):")
    colc = Counter()
    for loc in located:
        colc.update({c for cols in loc.values() for c in cols})
    for col, n in colc.most_common():
        print(f"    {col:<22} {n:>4}  ({n / total:>5.1%})")
    print()


def measure_field_population(path: Path) -> None:
    """§1 support: how often each mapped SAP column is actually populated."""
    headers, row_dicts, records, present = load(path)
    n = len(records)
    alias = {v: k for k, v in COLUMN.items()}
    print(f"dataset : {path.name}  ({n} rows, {len(headers)} columns)")
    print()
    print("--- 1.1 Populated rate of every column the model maps ---")
    print(f"{'SAP column':<26} | {'model field':<24} | {'populated':>9} | {'%':>6}")
    print(f"{'-' * 26}-+-{'-' * 24}-+-{'-' * 9}-+-{'-' * 6}")
    a2f = {}
    from api.routes import _input_alias_to_field, _norm_header
    a2f = _input_alias_to_field()
    seen: set[str] = set()
    for header in headers:
        field = a2f.get(_norm_header(header))
        if field is None or field in seen:
            continue
        seen.add(field)
        filled = sum(1 for r in records if not is_blank(getattr(r, field, None)))
        print(f"{header:<26} | {field:<24} | {filled:>9} | {filled / n:>5.1%}")
    print()
    print("--- 1.2 Columns present in the file that the model does not map ---")
    print("    (accepted and silently discarded: no extra='forbid' is declared,")
    print("     api/models.py:40)")
    for header in headers:
        if a2f.get(_norm_header(header)) is None:
            print(f"    {header}")
    print()


def measure_oracle_delta(path: Path) -> None:
    """§3 support: the workbook's own answer key against the measured detector."""
    from openpyxl import load_workbook

    _, _, records, present = load(path)
    measured = Counter()
    for r in records:
        measured.update(detect_issues(r, present))

    wb = load_workbook(path, read_only=True, data_only=True)
    oracle: dict[str, int] = {}
    for row in wb["Issue_Counts"].iter_rows(min_row=4, values_only=True):
        if row and row[2]:
            oracle[str(row[2]).strip()] = int(row[4])
    summary = {}
    for row in wb["Oracle_Summary"].iter_rows(min_row=4, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip()] = row[1]
    wb.close()

    print("Sheets read: 'Issue_Counts', 'Oracle_Summary' of the same workbook.")
    print("The workbook describes these as a ground-truth answer key")
    print("(Oracle_Summary row 2). They are compared here, never used as a source.")
    print()
    print("--- 3.5a Headline claims vs measured ---")
    print(f"{'metric':<32} | {'oracle':>8} | {'measured':>9}")
    print(f"{'-' * 32}-+-{'-' * 8}-+-{'-' * 9}")
    print(f"{'Total records':<32} | {str(summary.get('Total records')):>8} | {len(records):>9}")
    with_issues = sum(1 for r in records if detect_issues(r, present))
    print(f"{'Records with >=1 issue':<32} | {str(summary.get('Records with >=1 issue')):>8} | {with_issues:>9}")
    print(f"{'Clean records':<32} | {str(summary.get('Clean records')):>8} | {len(records) - with_issues:>9}")
    print(f"{'Total issue instances':<32} | {str(summary.get('Total issue instances')):>8} | {sum(measured.values()):>9}")
    print(f"{'Distinct issue codes covered':<32} | {str(summary.get('Distinct issue codes covered')):>8} | {f'{len(measured)}/{len(ISSUE_CATALOGUE)}':>9}")
    print()
    print("--- 3.5b Per-code: oracle count vs measured count ---")
    print(f"{'code':<15} | {'oracle':>7} | {'measured':>9} | {'delta':>7} | note")
    print(f"{'-' * 15}-+-{'-' * 7}-+-{'-' * 9}-+-{'-' * 7}-+-{'-' * 34}")
    for code in ISSUE_CATALOGUE:
        o = oracle.get(code)
        m = measured.get(code, 0)
        note = ""
        if o is None:
            note = "absent from the oracle sheet"
        elif o != m:
            note = "disagree"
        o_s = "-" if o is None else str(o)
        d_s = "-" if o is None else f"{m - o:+d}"
        print(f"{code:<15} | {o_s:>7} | {m:>9} | {d_s:>7} | {note}")
    only_oracle = [c for c in oracle if c not in ISSUE_CATALOGUE]
    agree = sum(1 for c in ISSUE_CATALOGUE if oracle.get(c) == measured.get(c, 0))
    print()
    print(f"codes in the declared catalogue      : {len(ISSUE_CATALOGUE)}")
    print(f"codes listed in the oracle sheet     : {len(oracle)}")
    print(f"declared codes absent from the oracle: {[c for c in ISSUE_CATALOGUE if c not in oracle]}")
    print(f"oracle codes not in the catalogue    : {only_oracle}")
    print(f"codes where oracle == measured       : {agree} of {len(ISSUE_CATALOGUE)}")
    print()


def measure_duplicates(path: Path) -> None:
    _, _, records, _ = load(path)
    rows = [
        DedupRow(
            row_id=r.record_id or f"row-{i}",
            block_id=None,
            name1=r.name_1, name2=r.name_2,
            name3=r.name_3, name4=r.name_4, name5=r.name_5,
            street=r.street_1, house_no=r.house_number,
            postal_code=r.postal_code, city=r.city, country=r.country_region_key,
        )
        for i, r in enumerate(records, start=1)
    ]
    print(f"dataset : {path.name}  ({len(rows)} rows)")
    print("method  : STEP A only - deterministic block + exact-signature collapse")
    print("          (dedup/signatures.py). No LLM adjudication was run.")
    print("blocking: fallback derive_block_id over normalised")
    print("          (country, postal_code, street, house_no) - the file carries no Block ID column.")
    print()

    blocks = group_rows_by_block(rows)
    block_sizes = Counter(len(v) for v in blocks.values())
    print("--- 5.1 Address blocks ---")
    print(f"distinct derived blocks          : {len(blocks)}")
    print(f"rows in a block of size 1        : {sum(n for s, n in ((s, s * c) for s, c in block_sizes.items()) if s == 1)}")
    multi_rows = sum(s * c for s, c in block_sizes.items() if s > 1)
    print(f"rows in a block of size > 1      : {multi_rows}  ({multi_rows / len(rows):.1%})")
    print(f"largest block                    : {max(block_sizes)} rows")
    print(f"{'block size':>10} | {'blocks':>7} | {'rows':>6}")
    print(f"{'-' * 10}-+-{'-' * 7}-+-{'-' * 6}")
    for s in sorted(block_sizes):
        print(f"{s:>10} | {block_sizes[s]:>7} | {s * block_sizes[s]:>6}")
    print()

    print("--- 5.2 Exact-signature collapse within blocks ---")
    collapsed_groups = 0
    rows_in_collapse = 0
    sig_sizes = Counter()
    multi_sig_blocks = 0
    for block_rows in blocks.values():
        sigs = build_signatures(block_rows)
        if len(sigs) > 1:
            multi_sig_blocks += 1
        for sig in sigs:
            sig_sizes[len(sig.row_ids)] += 1
            if len(sig.row_ids) > 1:
                collapsed_groups += 1
                rows_in_collapse += len(sig.row_ids)
    total_sigs = sum(sig_sizes.values())
    print(f"distinct signatures across all blocks       : {total_sigs}")
    print(f"signatures covering > 1 row (exact dupes)   : {collapsed_groups}")
    print(f"rows inside such a signature                : {rows_in_collapse}  ({rows_in_collapse / len(rows):.1%})")
    print(f"blocks holding >1 distinct signature        : {multi_sig_blocks}")
    print("  (these are the blocks an LLM adjudication would be asked to decide;")
    print("   whether their signatures merge is NOT determined by this run.)")
    print(f"{'rows/signature':>14} | {'signatures':>10}")
    print(f"{'-' * 14}-+-{'-' * 10}")
    for s in sorted(sig_sizes):
        print(f"{s:>14} | {sig_sizes[s]:>10}")
    print()


def measure_registry_ids(raw: Path, enriched: Path) -> None:
    raw_headers, _, _, _ = load(raw)
    enr_headers, enr_rows, _, _ = load(enriched)
    print(f"pre-enrichment  : {raw.name}")
    print(f"post-enrichment : {enriched.name}")
    print()
    for col in ("ROR ID", "LEI ID", "Domain", "Department Domain", "Record Type"):
        print(f"column {col!r:<20} in pre-enrichment file : {col in raw_headers}")
    print()
    n = len(enr_rows)
    ror = sum(1 for r in enr_rows if (r.get("ROR ID") or "").strip())
    lei = sum(1 for r in enr_rows if (r.get("LEI ID") or "").strip())
    either = sum(
        1 for r in enr_rows
        if (r.get("ROR ID") or "").strip() or (r.get("LEI ID") or "").strip()
    )
    both = sum(
        1 for r in enr_rows
        if (r.get("ROR ID") or "").strip() and (r.get("LEI ID") or "").strip()
    )
    print(f"rows in the enriched workbook            : {n}")
    print(f"rows with a non-empty ROR ID             : {ror}  ({ror / n:.1%})")
    print(f"rows with a non-empty LEI ID             : {lei}  ({lei / n:.1%})")
    print(f"rows with either identifier              : {either}  ({either / n:.1%})")
    print(f"rows with both identifiers               : {both}  ({both / n:.1%})")
    print(f"distinct ROR ID values                   : {len({(r.get('ROR ID') or '').strip() for r in enr_rows if (r.get('ROR ID') or '').strip()})}")
    print(f"distinct LEI ID values                   : {len({(r.get('LEI ID') or '').strip() for r in enr_rows if (r.get('LEI ID') or '').strip()})}")
    print()
    print("--- 6.1 Registry ids shared by more than one row (the dedup hint) ---")
    shared = Counter()
    for r in enr_rows:
        for key in ("ROR ID", "LEI ID"):
            v = (r.get(key) or "").strip()
            if v:
                shared[(key, v)] += 1
    repeated = {k: c for k, c in shared.items() if c > 1}
    rows_sharing = sum(repeated.values())
    print(f"identifier values carried by >1 row      : {len(repeated)}")
    print(f"rows carrying such a shared identifier   : {rows_sharing}  ({rows_sharing / n:.1%})")
    print(f"{'kind':<7} | {'rows':>5} | value")
    print(f"{'-' * 7}-+-{'-' * 5}-+-{'-' * 40}")
    for (kind, value), c in sorted(repeated.items(), key=lambda kv: (-kv[1], kv[0][1])):
        print(f"{kind:<7} | {c:>5} | {value}")
    print()


def main() -> None:
    print("=" * 78)
    print("Chapter 2 measurement run")
    print("=" * 78)
    print()
    print("#" * 78)
    print("# 1 - FIELD POPULATION (structure evidence)")
    print("#" * 78)
    print()
    measure_field_population(RAW_WORKBOOK)
    print("#" * 78)
    print("# 3 - ISSUE FREQUENCY")
    print("#" * 78)
    print()
    measure_issues(RAW_WORKBOOK)
    print("#" * 78)
    print("# 3.5 - THE WORKBOOK'S OWN ORACLE, COMPARED")
    print("#" * 78)
    print()
    measure_oracle_delta(RAW_WORKBOOK)
    print("#" * 78)
    print("# 5 - DUPLICATE PREVALENCE")
    print("#" * 78)
    print()
    measure_duplicates(RAW_WORKBOOK)
    print("#" * 78)
    print("# 6 - ENRICHMENT -> DEDUP COUPLING (registry identifiers)")
    print("#" * 78)
    print()
    measure_registry_ids(RAW_WORKBOOK, ENRICHED_WORKBOOK)
    print("=" * 78)
    print("end of run")
    print("=" * 78)


if __name__ == "__main__":
    main()
