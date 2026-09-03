#!/usr/bin/env python3
"""Derive every Issue-Catalogue figure the thesis quotes, from the source.

The counts in the module docstring, ``docs/thesis/`` and the
``PresentationTestData.xlsx`` oracle sheets must be *derived*, not asserted.
This script is the single derivation: run it and the numbers it prints are the
numbers those locations are required to carry.

    python3 scripts/issue_catalogue_census.py                  # catalogue only
    python3 scripts/issue_catalogue_census.py FILE.xlsx ...    # + per-file census
    python3 scripts/issue_catalogue_census.py --write-oracle   # rewrite the
                                                               # PresentationTestData
                                                               # oracle sheets

Definitions, stated once so every quoting location means the same thing:

  declared      entries in ``ISSUE_CATALOGUE`` — every code the module knows of
  live          status="live": in Catalogue v2 and emitted here
  unlisted      status="unlisted": emitted here, absent from Catalogue v2
  emitted       live + unlisted — the codes the detector can actually raise
                (``EMITTED_CODES``); "deterministically emitted"
  withdrawn     status="withdrawn": struck through in v2, declared for the
                audit trail, never emitted
  ndd           status="ndd": live in v2 but not deterministically detectable
  observed      codes that actually fired on a given file — a property of that
                data, not of the rule set
  fixture       codes with a positive+negative case in
                tests/fixtures/issue_catalogue_coverage.json
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from enrichment.issue_detection import (  # noqa: E402
    EMITTED_CODES,
    ISSUE_CATALOGUE,
    QUALITY_GROUPS,
    VERIFICATION_GROUPS,
    detect_issues,
)

_FIXTURE_PATH = (
    Path(__file__).resolve().parent.parent
    / "tests" / "fixtures" / "issue_catalogue_coverage.json"
)


def catalogue_census() -> dict[str, object]:
    status = Counter(e.status for e in ISSUE_CATALOGUE.values())
    live_quality = [
        e for e in ISSUE_CATALOGUE.values()
        if e.status == "live" and e.group in QUALITY_GROUPS
    ]
    fixtures = {
        case["code"] for case in json.loads(_FIXTURE_PATH.read_text())["cases"]
    }
    return {
        "declared": len(ISSUE_CATALOGUE),
        "live": status["live"],
        "unlisted": status["unlisted"],
        "withdrawn": status["withdrawn"],
        "ndd": status["ndd"],
        "emitted": len(EMITTED_CODES),
        "live_quality": len(live_quality),
        "origin": dict(Counter(e.origin for e in live_quality)),
        "group": dict(Counter(e.group for e in ISSUE_CATALOGUE.values())),
        "fixture_covered": len(fixtures & set(EMITTED_CODES)),
        "fixture_missing": sorted(set(EMITTED_CODES) - fixtures),
    }


def file_census(path: Path) -> tuple[Counter, int, int]:
    """(code -> record count, records, records with >=1 issue) for an XLSX."""
    from api.routes import (
        _flag_for_review, _parse_xlsx, _present_fields, _rows_to_records,
    )

    headers, rows = _parse_xlsx(path.read_bytes())
    records = _rows_to_records(rows)
    present = _present_fields(headers)
    counts: Counter = Counter()
    with_issue = 0
    for record, row in zip(records, rows):
        codes = detect_issues(
            record, present, flag_for_review=_flag_for_review(row, headers),
        )
        counts.update(codes)
        if codes:
            with_issue += 1
    return counts, len(records), with_issue


def _print_catalogue(c: dict) -> None:
    print("=" * 68)
    print("Issue Catalogue census — derived from enrichment/issue_detection.py")
    print("=" * 68)
    print(f"  declared                     {c['declared']}")
    print(f"  live                         {c['live']}")
    print(f"  unlisted (not in v2)         {c['unlisted']}")
    print(f"  deterministically emitted    {c['emitted']}   (live + unlisted)")
    print(f"  withdrawn                    {c['withdrawn']}")
    print(f"  not det. detectable          {c['ndd']}")
    print(f"  fixture-covered              {c['fixture_covered']} of {c['emitted']}")
    if c["fixture_missing"]:
        print(f"  !! no fixture: {c['fixture_missing']}")
    print(f"  live quality codes (G1-G6)   {c['live_quality']}")
    print(f"  origin of those              {c['origin']}")
    print(f"  entries per group            {c['group']}")
    print()
    print("  per-code status")
    for code, e in ISSUE_CATALOGUE.items():
        mark = {"live": " ", "unlisted": "?", "withdrawn": "x", "ndd": "-"}[e.status]
        print(f"    [{mark}] {code:<16} {e.group:<3} {e.origin:<4} "
              f"{'Error' if e.mandatory else 'Warning':<8} {e.name}")


_REPO = Path(__file__).resolve().parent.parent
_ORACLE_PATH = _REPO / "PresentationTestData.xlsx"
# The enriched twin carries the four Oracle_* sheets over verbatim. Its copies
# are refreshed from the same derivation so the pair does not drift apart.
_ORACLE_TWIN = _REPO / "PresentationTestData_enriched_checked_v1.xlsx"

_GROUP_NAMES = {
    "G1": "Data in Wrong Field",
    "G2": "Missing Required Data",
    "G3": "Duplicate or Conflicting Data",
    "G4": "Invalid Format or Length",
    "G5": "Non-Standard Naming",
    "G6": "Not Resolvable by Enrichment",
    "G7": "Verification Required",
    "G8": "Enrichment Unresolved",
}


def write_oracle_sheets(path: Path | None = None) -> None:
    """Rewrite ``Issue_Counts`` and the issue rows of ``Oracle_Summary``.

    Both sheets previously carried hand-authored figures ("all 36 codes",
    "Distinct issue codes covered: 36/36") that no run of the detector could
    produce. They are regenerated here from an actual audit of the workbook's
    own ``TestData_500`` sheet, so the answer key and the detector agree by
    construction rather than by assertion.

    Only the issue-related rows of ``Oracle_Summary`` are touched; the dedup,
    routing and country metrics are left exactly as they are.
    """
    from openpyxl import load_workbook

    target = path or _ORACLE_PATH
    counts, total, with_issue = file_census(target)
    observed = [c for c in EMITTED_CODES if counts.get(c)]

    wb = load_workbook(target)

    sheet = wb["Issue_Counts"]
    wb.remove(sheet)
    sheet = wb.create_sheet("Issue_Counts", wb.sheetnames.index("Oracle_Summary") + 1)
    sheet.append([
        f"Issue Counts by Code ({len(observed)} of {len(EMITTED_CODES)} "
        f"emittable codes observed on these {total} records)"
    ])
    sheet.append([
        "Derived by scripts/issue_catalogue_census.py --write-oracle from the "
        "TestData_500 sheet. Count = records exhibiting each issue."
    ])
    sheet.append(["Group", "Group Name", "Code", "Rule", "Count"])
    for code in EMITTED_CODES:
        entry = ISSUE_CATALOGUE[code]
        sheet.append([
            entry.group, _GROUP_NAMES[entry.group], code, entry.name,
            counts.get(code, 0),
        ])

    # Group_Totals — regenerated too. It previously listed G1-G5 only, which
    # silently dropped every G6 instance from the group view and had no place
    # at all for G7.
    index = wb.sheetnames.index("Group_Totals")
    wb.remove(wb["Group_Totals"])
    totals = wb.create_sheet("Group_Totals", index)
    totals.append(["Issue Instances by Group"])
    totals.append([
        "Derived by scripts/issue_catalogue_census.py --write-oracle. G6 is "
        "expected to persist through enrichment; G7 and G8 are not quality "
        "issues and are excluded from the reduction metric."
    ])
    totals.append(["Group", "Group Name", "Issue Instances"])
    for group in (*QUALITY_GROUPS, *VERIFICATION_GROUPS):
        totals.append([
            group,
            _GROUP_NAMES[group],
            sum(n for code, n in counts.items()
                if ISSUE_CATALOGUE[code].group == group),
        ])

    summary = wb["Oracle_Summary"]
    derived = {
        "Total records": total,
        "Clean records": total - with_issue,
        "Records with >=1 issue": with_issue,
        "Total issue instances": sum(counts.values()),
        "Distinct issue codes covered": f"{len(observed)}/{len(EMITTED_CODES)}",
    }
    for row in summary.iter_rows(min_row=1):
        label = row[0].value
        if label in derived:
            row[1].value = derived[label]

    wb.save(target)
    print(f"\nrewrote Issue_Counts and Oracle_Summary in {target.name}")
    for key, value in derived.items():
        print(f"    {key:<32} {value}")

    if path is None and _ORACLE_TWIN.exists():
        _mirror_oracle_sheets(wb, _ORACLE_TWIN, derived)


def _mirror_oracle_sheets(source_wb, twin_path: Path, derived: dict) -> None:
    """Copy the regenerated oracle sheets onto the enriched twin workbook.

    `07_EVALUATION.md` records that the twin carries the four ``Oracle_*``
    sheets over unchanged from the original. Refreshing only one of the pair
    would make that statement false, so the same values are written to both.
    """
    from openpyxl import load_workbook

    twin = load_workbook(twin_path)
    for name in ("Issue_Counts", "Group_Totals"):
        src = source_wb[name]
        index = twin.sheetnames.index(name)
        twin.remove(twin[name])
        dst = twin.create_sheet(name, index)
        for row in src.iter_rows(values_only=True):
            dst.append(list(row))

    summary = twin["Oracle_Summary"]
    for row in summary.iter_rows(min_row=1):
        if row[0].value in derived:
            row[1].value = derived[row[0].value]

    twin.save(twin_path)
    print(f"    mirrored onto {twin_path.name}")


def main(argv: list[str]) -> int:
    write_oracle = "--write-oracle" in argv
    paths = [Path(a) for a in argv if not a.startswith("--")]

    census = catalogue_census()
    _print_catalogue(census)

    for path in paths:
        counts, total, with_issue = file_census(path)
        observed = [c for c in EMITTED_CODES if counts.get(c)]
        print()
        print("-" * 68)
        print(f"{path.name} — {total} records")
        print("-" * 68)
        for code in EMITTED_CODES:
            if counts.get(code):
                print(f"    {code:<16} {ISSUE_CATALOGUE[code].group:<3} {counts[code]:>5}")
        print(f"  records with >=1 issue       {with_issue}")
        print(f"  clean records                {total - with_issue}")
        print(f"  total issue instances        {sum(counts.values())}")
        print(f"  distinct codes observed      {len(observed)} of "
              f"{len(EMITTED_CODES)} emittable")
        print(f"  never observed here          "
              f"{[c for c in EMITTED_CODES if not counts.get(c)]}")

    if write_oracle:
        write_oracle_sheets()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
