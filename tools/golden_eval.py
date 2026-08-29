"""The golden-set comparator: expected output vs what the pipeline produced.

`docs/SAMPLE_DATA/` holds a solved reference for 99 records — each raw record
followed by the enriched row a reviewer certified for it, in the pipeline's own
output schema. This module is the grader. It is pure: no network, no
orchestrator, no I/O beyond reading the reference workbook, so the rules it
applies can be tested without paying for a run.

**The reference grades the deterministic layer only.** Slot consolidation,
PO Box / house number / sub-location extraction, c/o and ATTN handling, e-mail
extraction, duplicate clearing, abbreviation expansion, and knowledge-based
canonical names. Everything registry- or run-dependent — ROR, LEI, Domain,
Record Type, Search Terms, and every Flag and Provenance column — is declared
``skip`` by the reference itself, because a value that depends on what a
registry answered today is not something a static file can certify.

Three levels of rule, most specific winning:

``Cell Notes``   one customer + one column. ``skip`` (with an optional reason
                after an em dash), or ``any_of: a | b | c`` widening the check
                to a set of acceptable values. ``(empty)`` inside an ``any_of``
                list means a blank cell is one of them. 87 such notes exist,
                and the reference's own Method sheet is explicit about why:
                where one reviewer's knowledge could not certify a single
                correct value, the cell carries the most defensible value and a
                note widening the check. **A disagreement on a noted cell is
                not evidence against the pipeline.**

``Match Rules`` one column, all customers. ``exact``, ``exact_ci`` (street and
                city casing is not specified by the thesis), or ``skip``.

default         a column the reference does not mention is not graded. Silence
                is not an assertion — the same principle `enrichment.locality`
                applies to evidence.

An empty expected cell in a graded column IS an assertion: the reference says
that column should be blank. An empty expected cell in a ``skip`` column is no
claim at all.
"""

from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

REFERENCE_SHEET = "Reference"
RULES_SHEET = "Match Rules"
NOTES_SHEET = "Cell Notes"

#: The two non-schema columns the Reference sheet carries to pair its rows.
PAIR_COLUMNS = ("pair_id", "row_kind")

EXACT = "exact"
EXACT_CI = "exact_ci"
EXACT_ABBREV = "exact_abbrev"
SKIP = "skip"

#: Abbreviation pairs the pipeline's own maps rewrite, in both directions.
#: `enrichment.address_processing.STREET_TYPE_ABBREVIATIONS` and
#: `DIRECTIONAL_ABBREVIATIONS` for the street half,
#: `utils.text_utils.expand_abbreviations` for the name half.
#:
#: Used ONLY by :data:`EXACT_ABBREV`, to recognise that two strings differ by an
#: abbreviation the pipeline is documented as rewriting. It never rewrites a
#: value and it is not a normaliser — it decides whether a disagreement is about
#: the organisation or about the convention for writing it down.
ABBREVIATION_PAIRS: tuple[tuple[str, str], ...] = (
    ("street", "st"), ("street", "str"), ("avenue", "ave"),
    ("boulevard", "blvd"),
    ("drive", "dr"), ("road", "rd"), ("lane", "ln"), ("court", "ct"),
    ("highway", "hwy"), ("parkway", "pkwy"), ("route", "rte"),
    ("north", "n"), ("south", "s"), ("east", "e"), ("west", "w"),
    ("northwest", "nw"), ("northeast", "ne"),
    ("southwest", "sw"), ("southeast", "se"),
    ("university", "univ"), ("department", "dept"), ("institute", "inst"),
    ("laboratory", "lab"), ("laboratories", "labs"), ("center", "ctr"),
    ("centre", "ctr"), ("medical", "med"), ("services", "svcs"),
    ("group", "grp"), ("national", "natl"), ("international", "intl"),
    ("administration", "admin"), ("management", "mgmt"),
    ("building", "bldg"), ("division", "div"), ("reference", "ref"),
)

_TO_LONG_FORM: dict[str, str] = {}
for _long, _short in ABBREVIATION_PAIRS:
    _TO_LONG_FORM[_short] = _long
    _TO_LONG_FORM[_long] = _long


def fold_abbreviations(text: str) -> str:
    """Rewrite every abbreviation in *text* to its long form, lower case.

    Deliberately whole-word and singular/plural aware: `laboratories` folds to
    `laboratories`, not to `laboratory`, so a pipeline that turned
    `Bio-Rad Laboratories` into `Bio-Rad Laboratory` is still caught.
    """
    return " ".join(
        _TO_LONG_FORM.get(word, word)
        for word in re.findall(r"[a-z0-9&]+", text.casefold())
    )

#: What a blank looks like inside an `any_of` list.
EMPTY_TOKEN = "(empty)"

MATCH = "match"
MISMATCH = "mismatch"
SKIPPED = "skipped"


def normalise(value: Any) -> str:
    """Render a cell as the string both sides are compared as.

    Excel hands back an int for a numeric postal code in one file and a string
    in another, and a ``datetime`` for a date the other file stores as text.
    Comparing those raw reports a difference that is not one — the reference
    and the pipeline agree about the value and disagree about the type — so
    every cell is rendered the same way before any rule is applied.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, (_dt.datetime, _dt.date)):
        # The pipeline round-trips dates as text; take the date part only, so
        # a midnight timestamp and a plain date are the same day.
        value = value.date() if isinstance(value, _dt.datetime) else value
        return value.isoformat()
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(v) for v in value if v is not None)
    text = str(value).strip()
    # Excel's non-breaking spaces and the reference's typographic dashes are
    # invisible differences that are never what a reviewer meant to assert.
    text = text.replace(" ", " ").replace("‑", "-")
    return re.sub(r"\s+", " ", text)


@dataclass(frozen=True)
class Rule:
    """How one cell is graded."""

    kind: str                       # EXACT | EXACT_CI | SKIP
    alternatives: tuple[str, ...] = ()   # non-empty only for a widened cell
    reason: str = ""                # why a cell note skipped or widened it
    source: str = "column"          # "column" | "cell-note" | "default"

    @property
    def graded(self) -> bool:
        return self.kind != SKIP


SKIP_RULE = Rule(kind=SKIP, source="default", reason="column not in Match Rules")


def parse_note(raw: Any) -> Rule | None:
    """Parse one Cell Notes rule. Returns None when it says nothing usable."""
    text = normalise(raw)
    if not text:
        return None
    lowered = text.casefold()
    if lowered.startswith("skip"):
        # "skip" or "skip — <reason>"; the dash form is the common one.
        reason = re.sub(r"^skip\s*[—–-]?\s*", "", text, flags=re.I)
        return Rule(kind=SKIP, reason=reason, source="cell-note")
    if lowered.startswith("any_of:"):
        body = text.split(":", 1)[1]
        alts = tuple(
            part.strip() for part in body.split("|") if part.strip()
        )
        if not alts:
            return None
        return Rule(
            kind=EXACT,
            alternatives=alts,
            reason=f"widened to {len(alts)} acceptable values",
            source="cell-note",
        )
    return None


def compare_cell(expected: Any, actual: Any, rule: Rule) -> str:
    """Grade one cell. Returns MATCH, MISMATCH or SKIPPED."""
    if not rule.graded:
        return SKIPPED

    got = normalise(actual)
    if rule.alternatives:
        accepted = {
            "" if alt.casefold() == EMPTY_TOKEN else alt
            for alt in rule.alternatives
        }
    else:
        accepted = {normalise(expected)}

    if rule.kind == EXACT_CI:
        return MATCH if got.casefold() in {
            a.casefold() for a in accepted
        } else MISMATCH
    if rule.kind == EXACT_ABBREV:
        return MATCH if fold_abbreviations(got) in {
            fold_abbreviations(a) for a in accepted
        } else MISMATCH
    return MATCH if got in accepted else MISMATCH


@dataclass
class CellResult:
    customer: str
    column: str
    verdict: str
    expected: str
    actual: str
    rule: Rule


@dataclass
class Reference:
    """The reference workbook, parsed."""

    columns: list[str]
    inputs: dict[str, dict[str, Any]]
    expected: dict[str, dict[str, Any]]
    rules: dict[str, Rule]
    notes: dict[tuple[str, str], Rule]
    #: Notes naming a customer or column that does not exist. Surfaced rather
    #: than dropped: a note that grades nothing is a defect in the reference,
    #: and silently ignoring it would hide a column-name drift.
    orphan_notes: list[tuple[str, str, str]] = field(default_factory=list)
    #: Column rules replaced from the overrides file, as "col: was -> now".
    overrides_applied: list[str] = field(default_factory=list)

    def rule_for(self, customer: str, column: str) -> Rule:
        note = self.notes.get((customer, column))
        if note is not None:
            return note
        return self.rules.get(column, SKIP_RULE)

    @property
    def graded_columns(self) -> list[str]:
        return [c for c in self.columns if self.rules.get(c, SKIP_RULE).graded]


def apply_overrides(rules: dict[str, Rule], overrides_path: str) -> list[str]:
    """Replace column rules from an overrides file. Returns what changed.

    The reference workbook is left exactly as its author wrote it; the
    corrections live in `docs/SAMPLE_DATA/reference_overrides.json` so every
    deviation from the authored rules is reviewable in one place and revertible
    with `--no-overrides`. The bar for an entry is stated in that file: the
    reference asserts a convention the pipeline is documented and tested as
    deliberately doing otherwise.
    """
    import json

    data = json.loads(Path(overrides_path).read_text(encoding="utf-8"))
    applied: list[str] = []
    for column, spec in (data.get("column_rules") or {}).items():
        kind = str(spec.get("rule", "")).strip()
        if kind not in (EXACT, EXACT_CI, EXACT_ABBREV, SKIP):
            raise ValueError(f"{column}: unknown override rule {kind!r}")
        if column not in rules:
            raise ValueError(f"{column}: not a column of the reference")
        was = rules[column].kind
        rules[column] = Rule(kind=kind, reason="override", source="override")
        applied.append(f"{column}: {was} -> {kind}")
    return applied


def load_reference(path: str, overrides: str | None = None) -> Reference:
    """Read the solved-reference workbook into a :class:`Reference`."""
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)

    sheet = book[REFERENCE_SHEET]
    rows = sheet.iter_rows(values_only=True)
    header = [normalise(c) for c in next(rows)]
    columns = [c for c in header if c and c not in PAIR_COLUMNS]

    inputs: dict[str, dict[str, Any]] = {}
    expected: dict[str, dict[str, Any]] = {}
    for raw in rows:
        if raw is None or not any(c is not None for c in raw):
            continue
        row = dict(zip(header, raw))
        customer = normalise(row.get("Customer"))
        if not customer:
            continue
        kind = normalise(row.get("row_kind")).upper()
        target = inputs if kind == "INPUT" else expected if kind == "EXPECTED" else None
        if target is not None:
            target[customer] = row

    rules: dict[str, Rule] = {}
    for raw in book[RULES_SHEET].iter_rows(values_only=True):
        column, kind = normalise(raw[0]), normalise(raw[1]).casefold()
        if not column or column.casefold() == "column":
            continue
        why = normalise(raw[2]) if len(raw) > 2 else ""
        if kind in (EXACT, EXACT_CI, SKIP):
            rules[column] = Rule(kind=kind, reason=why, source="column")

    overrides_applied: list[str] = []
    if overrides:
        overrides_applied = apply_overrides(rules, overrides)

    notes: dict[tuple[str, str], Rule] = {}
    orphans: list[tuple[str, str, str]] = []
    for raw in book[NOTES_SHEET].iter_rows(values_only=True):
        customer, column = normalise(raw[0]), normalise(raw[1])
        if not customer or customer.casefold() == "customer":
            continue
        rule = parse_note(raw[2] if len(raw) > 2 else None)
        if rule is None:
            continue
        if customer not in expected or column not in columns:
            orphans.append((customer, column, "unknown customer or column"))
            continue
        notes[(customer, column)] = rule

    book.close()
    return Reference(
        columns=columns, inputs=inputs, expected=expected,
        rules=rules, notes=notes, orphan_notes=orphans,
        overrides_applied=overrides_applied,
    )


def compare(
    reference: Reference,
    produced: dict[str, dict[str, Any]],
) -> tuple[list[CellResult], dict[str, Any]]:
    """Grade *produced* (customer -> output row) against the reference.

    Returns every graded cell result and a summary. A customer the run did not
    produce is reported as such rather than counted as a pass — a missing row
    is the most complete failure there is, and scoring it as "no mismatches"
    is the one arithmetic that could hide it.
    """
    results: list[CellResult] = []
    missing: list[str] = []

    for customer, expected_row in reference.expected.items():
        actual_row = produced.get(customer)
        if actual_row is None:
            missing.append(customer)
            continue
        for column in reference.columns:
            rule = reference.rule_for(customer, column)
            verdict = compare_cell(
                expected_row.get(column), actual_row.get(column), rule,
            )
            results.append(CellResult(
                customer=customer,
                column=column,
                verdict=verdict,
                expected=normalise(expected_row.get(column)),
                actual=normalise(actual_row.get(column)),
                rule=rule,
            ))

    graded = [r for r in results if r.verdict != SKIPPED]
    failed = [r for r in graded if r.verdict == MISMATCH]
    failing_customers = {r.customer for r in failed}
    scored = set(reference.expected) - set(missing)

    summary = {
        "records_expected": len(reference.expected),
        "records_produced": len(scored),
        "records_missing": sorted(missing),
        "records_passed": len(scored - failing_customers),
        "records_failed": len(failing_customers),
        "cells_graded": len(graded),
        "cells_matched": len(graded) - len(failed),
        "cells_failed": len(failed),
        "cell_accuracy": (
            (len(graded) - len(failed)) / len(graded) if graded else 0.0
        ),
        "record_accuracy": (
            len(scored - failing_customers) / len(scored) if scored else 0.0
        ),
        "orphan_notes": reference.orphan_notes,
    }
    return results, summary


def by_column(results: list[CellResult]) -> list[dict[str, Any]]:
    """Per-column tallies, worst first — where the failures actually are."""
    tally: dict[str, dict[str, int]] = {}
    for r in results:
        entry = tally.setdefault(r.column, {"graded": 0, "failed": 0})
        if r.verdict == SKIPPED:
            continue
        entry["graded"] += 1
        if r.verdict == MISMATCH:
            entry["failed"] += 1
    rows = [
        {
            "column": column,
            "graded": e["graded"],
            "failed": e["failed"],
            "accuracy": (e["graded"] - e["failed"]) / e["graded"]
            if e["graded"] else 0.0,
        }
        for column, e in tally.items()
    ]
    rows.sort(key=lambda r: (-r["failed"], r["column"]))
    return rows
