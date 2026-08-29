"""Merge the two enriched sample sets into one thesis workbook.

s2_now.xlsx + s3_now.xlsx -> enriched_samples_200.xlsx

Both carry the identical 81-column schema (73 pipeline output columns plus the
eval metadata the source workbooks brought with them), so the merge is a
concatenation with the header taken once.

Three scoring columns are APPENDED at the end, after every original column, so
nothing downstream that reads by position or by name is disturbed:

  type_match    correct / wrong / undecided   -- Record Type vs record_type_hint
  has_identity  y / n                         -- a ROR id or an LEI id
  domain_conf   verified / provisional / low / -

`undecided` is kept distinct from `wrong` on purpose: `unknown` asserts nothing,
and collapsing the two would score an honest abstention as an error.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SOURCES = [
    ("S2", "logs/compare/s2_now.xlsx"),
    ("S3", "logs/compare/s3_now.xlsx"),
]
OUT = "logs/compare/enriched_samples_200.xlsx"

EXTRA = ["type_match", "has_identity", "domain_conf"]

book = openpyxl.Workbook()
sheet = book.active
sheet.title = "enriched_200"

header = None
rows_written = 0
tally = {}

for tag, path in SOURCES:
    ws = openpyxl.load_workbook(path, read_only=True).active
    it = ws.iter_rows(values_only=True)
    head = list(next(it))
    if header is None:
        header = head + EXTRA
        sheet.append(header)
    else:
        assert head + EXTRA == header, f"{path} has a different schema"
    ix = {h: i for i, h in enumerate(head)}

    for row in it:
        if row[ix["Name 1"]] is None:
            continue
        rtype = (row[ix["Record Type"]] or "").strip()
        hint = (row[ix["record_type_hint"]] or "").strip()
        if rtype in ("", "unknown"):
            verdict = "undecided"
        elif rtype == hint:
            verdict = "correct"
        else:
            verdict = "wrong"

        ror = (row[ix["ROR ID"]] or "").strip()
        lei = (row[ix["LEI ID"]] or "").strip()
        prov = (row[ix["Domain Provenance"]] or "").strip()
        conf = "-"
        for level in ("verified", "provisional", "low"):
            if f":{level}" in prov:
                conf = level
                break

        sheet.append(list(row) + [verdict, "y" if (ror or lei) else "n", conf])
        rows_written += 1
        tally[(tag, verdict)] = tally.get((tag, verdict), 0) + 1

# ── presentation: freeze the header, bold it, widen the columns people read ──
head_font = Font(bold=True, color="FFFFFF")
head_fill = PatternFill("solid", fgColor="44546A")
for cell in sheet[1]:
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(vertical="center", wrap_text=False)
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions

WIDTHS = {
    "Name 1": 38, "Name 2": 28, "City": 16, "Record Type": 18,
    "record_type_hint": 17, "type_match": 12, "has_identity": 12,
    "domain_conf": 12, "Domain": 26, "Domain Provenance": 30,
    "ROR ID": 26, "LEI ID": 22, "Flag Codes": 24, "eval_set": 9,
}
for i, name in enumerate(header, start=1):
    if name in WIDTHS:
        sheet.column_dimensions[get_column_letter(i)].width = WIDTHS[name]

book.save(OUT)

print(f"{OUT}  —  {rows_written} rows x {len(header)} columns\n")
print(f"{'set':6s} {'correct':>8s} {'wrong':>7s} {'undecided':>10s}")
for tag, _ in SOURCES:
    c = tally.get((tag, "correct"), 0)
    w = tally.get((tag, "wrong"), 0)
    u = tally.get((tag, "undecided"), 0)
    print(f"{tag:6s} {c:8d} {w:7d} {u:10d}")
tot = lambda v: sum(tally.get((t, v), 0) for t, _ in SOURCES)
print(f"{'ALL':6s} {tot('correct'):8d} {tot('wrong'):7d} {tot('undecided'):10d}")
