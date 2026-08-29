"""Ticket 19 - build the 'currently lost' population and every query variant.

Lost = a record that left the pipeline with an empty ``ror_id`` AND an empty
``lei_id``.  Two corpora:

* **A** ``docs/thesis/chemspeed_us_100.xlsx`` joined to ticket 11's live run
  artefact ``tmp/run100b.json`` (raw input names available, so raw-vs-control
  is a clean comparison here).
* **D** the two 100-record labelled files in ``docs/results/``.  Their ``Name 1``
  column is the *enriched* value, which is also exactly what ticket 11's
  harness D fed back in, so the numbers stay comparable - but "raw" on corpus D
  means "the file's Name 1", not the untouched SAP input.

Writes ``tmp/q19/population.json``.  No network.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_ROOT))

from api.routes import _parse_xlsx, _rows_to_records  # noqa: E402
from enrichment.preprocess import preprocess_record  # noqa: E402
from enrichment.tier1_ror import _expand_state_abbrevs  # noqa: E402
from enrichment.tier1_lei import _LEGAL_FORM_TOKENS  # noqa: E402
from utils.text_utils import (  # noqa: E402
    country_to_iso_code,
    expand_abbreviations,
    strip_address_fragments,
)

OUT = _ROOT / ".scratch/agentic-enrichment/tmp/q19"
OUT.mkdir(parents=True, exist_ok=True)

CHEM = _ROOT / "docs/thesis/chemspeed_us_100.xlsx"
RUN100 = _ROOT / ".scratch/agentic-enrichment/tmp/q19/runA.json"
S2 = _ROOT / "docs/results/demo_S2_large_corporate_100_v1 (1)_enriched.xlsx"
S3 = _ROOT / "docs/results/demo_S3_government_labs_100_v1 (1)_enriched.xlsx"

_MULTI = {"co", "com", "org", "net", "gov", "ac", "edu"}


def registrable(domain):
    if not domain:
        return None
    d = str(domain).strip().lower().rstrip(".")
    if d.startswith("www."):
        d = d[4:]
    parts = d.split(".")
    if len(parts) < 2:
        return None
    if len(parts) >= 3 and parts[-2] in _MULTI and len(parts[-1]) == 2:
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


_SFX = re.compile(
    r"[,\s]+(?:" + "|".join(sorted(re.escape(t) for t in _LEGAL_FORM_TOKENS)) + r")\.?\s*$",
    re.I,
)


def strip_legal_suffix(name):
    out = name
    for _ in range(3):
        nxt = _SFX.sub("", out).strip(" ,.")
        if nxt == out or not nxt:
            break
        out = nxt
    return out or name


def strip_location_tokens(name, city, state, country):
    toks = set()
    for src in (city, state, country):
        for t in re.findall(r"[A-Za-z]{3,}", str(src or "")):
            toks.add(t.lower())
    if not toks:
        return name
    kept = [w for w in name.split()
            if re.sub(r"[^A-Za-z]", "", w).lower() not in toks]
    out = " ".join(kept).strip(" ,-")
    return out if len(out) >= 3 else name


def variants(rec, control):
    """The query strings, keyed by strategy id.  ``control`` is what the
    pipeline sends today."""
    raw = (rec["name1_raw"] or "").strip()
    v = {"control": control, "raw": raw}
    # `call_ror` already applies `_expand_state_abbrevs` to the query it sends,
    # so the *isolated* variable is `expand_abbreviations` - which today reaches
    # only the rescore list, never the query.
    v["expand_query"] = (expand_abbreviations(control) or control).strip()
    v["nosuffix"] = strip_legal_suffix(control)
    v["noloc"] = strip_location_tokens(control, rec["city"], rec["state"], rec["country"])
    n2 = (rec["name2_raw"] or "").strip()
    if n2:
        v["name1_name2"] = (control + " " + n2).strip()
    if "/" in control or "/" in raw:
        src = control if "/" in control else raw
        v["slashfix"] = src.replace("/", " ").strip()
    return {k: s for k, s in v.items() if s}


def control_for(rec_obj):
    """Reproduce drive_tier1.py's entry conditions exactly."""
    pre = preprocess_record(
        name1=rec_obj.name1, name2=rec_obj.name2, name3=rec_obj.name3,
        name4=rec_obj.name4, name5=rec_obj.name5, contact=rec_obj.contact,
        email=rec_obj.email, street1=rec_obj.street, street2=rec_obj.street2,
        street3=rec_obj.street3, street4=rec_obj.street4, street5=rec_obj.street5,
        house_number=rec_obj.house_number, llm_person_verdicts={},
    )
    pp = (pre.name1 or "").strip()
    if not pp:
        return "", ""
    cleaned = strip_address_fragments(
        pp, street=(pre.street1 or rec_obj.street), city=rec_obj.city,
        state=rec_obj.state, zip_code=rec_obj.zip,
    ) or pp
    return pp, cleaned


def load_xlsx_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    rows = [dict(zip(hdr, r)) for r in it if any(c is not None for c in r)]
    wb.close()
    return rows


def build():
    out = []

    # ---- corpus A ---------------------------------------------------------
    _h, rows = _parse_xlsx(CHEM.read_bytes())
    recs = _rows_to_records(rows)
    run = json.loads(RUN100.read_text(encoding="utf-8"))
    assert len(recs) == len(run["results"]) == len(run["inputs"])
    for i, (r, res, inp) in enumerate(zip(recs, run["results"], run["inputs"])):
        assert (r.name1 or "").strip().lower() == (inp["name1"] or "").strip().lower()
        pp, cleaned = control_for(r)
        out.append({
            "corpus": "A", "idx": i,
            "name1_raw": r.name1, "name2_raw": r.name2,
            "city": r.city, "state": r.state, "country": r.country,
            "cc": country_to_iso_code(r.country),
            "control": cleaned, "preprocessed": pp,
            "domain": registrable(res.get("domain")),
            "domain_raw": res.get("domain"),
            "domain_provenance": res.get("domain_provenance"),
            "ror_id": res.get("ror_id"), "lei_id": res.get("lei_id"),
            "name1_enriched": res.get("name1_enriched"),
            "hint": None,
        })

    # ---- corpus D ---------------------------------------------------------
    for tag, path in (("S2", S2), ("S3", S3)):
        _h, rows = _parse_xlsx(path.read_bytes())
        recs = _rows_to_records(rows)
        raws = load_xlsx_rows(path)
        assert len(recs) == len(raws)
        for i, (r, raw) in enumerate(zip(recs, raws)):
            pp, cleaned = control_for(r)
            out.append({
                "corpus": tag, "idx": i,
                "name1_raw": r.name1, "name2_raw": r.name2,
                "city": r.city, "state": r.state, "country": r.country,
                "cc": country_to_iso_code(r.country),
                "control": cleaned, "preprocessed": pp,
                "domain": registrable(raw.get("Domain")),
                "domain_raw": raw.get("Domain"),
                "domain_provenance": raw.get("Domain Provenance"),
                "ror_id": raw.get("ROR ID") or None,
                "lei_id": raw.get("LEI ID") or None,
                "name1_enriched": raw.get("Name 1"),
                "hint": raw.get("record_type_hint"),
            })
    return out


def main():
    pop = build()
    for d in pop:
        d["lost"] = not d["ror_id"] and not d["lei_id"]
        d["variants"] = variants(d, d["control"]) if d["control"] else {}
    (OUT / "population.json").write_text(json.dumps(pop, indent=1), encoding="utf-8")

    from collections import Counter
    print("total records: %d" % len(pop))
    for c in ("A", "S2", "S3"):
        sub = [d for d in pop if d["corpus"] == c]
        lost = [d for d in sub if d["lost"]]
        withdom = [d for d in lost if d["domain"]]
        print("  %s: n=%d lost=%d lost-with-domain=%d lost-without-domain=%d"
              % (c, len(sub), len(lost), len(withdom), len(lost) - len(withdom)))
    lost = [d for d in pop if d["lost"]]
    print("TOTAL lost = %d  with domain = %d"
          % (len(lost), sum(1 for d in lost if d["domain"])))
    print("\nvariant strings that DIFFER from control, among lost records:")
    diff = Counter()
    for d in lost:
        for k, s in d["variants"].items():
            if k != "control" and s.strip().lower() != d["control"].strip().lower():
                diff[k] += 1
    for k, n in diff.most_common():
        print("   %-14s %d" % (k, n))
    print("\nsample lost records:")
    for d in lost[:12]:
        print("   [%s] raw=%r ctl=%r dom=%s"
              % (d["corpus"], d["name1_raw"], d["control"], d["domain"]))


if __name__ == "__main__":
    main()
