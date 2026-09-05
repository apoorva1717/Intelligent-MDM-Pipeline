"""Tests for the deterministic scoring + golden-record election.

Offline pytest — no network, no LLM. Drives ``score_row`` /
``elect_golden_records`` directly for the arithmetic, the routes via httpx
for the HTTP contract, and ``score_workbook`` for the XLSX round-trip.

UNCONFIRMED values flagged in the model (verify with Bernd before go-live):
- combined_presence_bonus = 10 (Bernd gave no number)
- sales_order_partner_count tiers (assumed to mirror sales_order_count)
- the tie-break ordering (score, last_order_year, equipment_count,
  company_code_count, lowest row_id)
"""

from __future__ import annotations

import datetime
import io
import json
import os
import random
import sys
from pathlib import Path

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ.setdefault("ENV", "local")
os.environ.setdefault("MOCK_EXTERNAL_CALLS", "true")
os.environ.setdefault("OPENAI_API_KEY", "test-key-for-mocks")

from dedup.scoring import (
    DuplicateRowIdError,
    ScoringRow,
    coerce_weights,
    derived_counts,
    detect_issues,
    elect_golden_records,
    load_weights,
    score_row,
)
from dedup.scoring_xlsx import (
    BREAKDOWN_COLUMNS,
    score_workbook,
)


WEIGHTS = load_weights()

# The two *_last_used ladders are banded on the OFFSET from the election's
# reference year, not on absolute years, so every year literal in this file is
# expressed relative to that anchor. elect_golden_records / score_workbook
# resolve the anchor from the clock, so the tests must too: a hard-coded 2026
# would start failing on 1 January 2027 — the very drift the relative ladder
# exists to prevent.
THIS_YEAR = datetime.date.today().year
Y0 = THIS_YEAR      # this year          -> offset 0 -> 20 pts
Y1 = THIS_YEAR - 1  # last year          -> offset 1 -> 15 pts
Y2 = THIS_YEAR - 2  # two years back     -> offset 2 -> 10 pts
Y3 = THIS_YEAR - 3  # three years back   -> offset 3 -> 5 pts
Y4 = THIS_YEAR - 4  # off the ladder     -> 0 pts
Y5 = THIS_YEAR - 5
Y6 = THIS_YEAR - 6


def _score(field_values: dict, *, current_year: int = None) -> tuple[dict, list]:
    row = ScoringRow(row_id="1", **field_values)
    return score_row(
        row, WEIGHTS,
        current_year=THIS_YEAR if current_year is None else current_year,
    )


def _points(field_values: dict, criterion: str, *, current_year: int = None) -> int:
    breakdown, _ = _score(field_values, current_year=current_year)
    return breakdown[criterion]


def _by_row(results):
    return {r.row_id: r for r in results}


# ---------------------------------------------------------------------------
# Band boundaries
# ---------------------------------------------------------------------------

class TestBands:
    @pytest.mark.parametrize("year,expected", [
        (Y0, 20), (Y1, 15), (Y2, 10), (Y3, 5),
        (Y4, 0), (Y5, 0), (None, 0),
    ])
    def test_sales_order_last_used(self, year, expected):
        assert _points({"last_order_year": year}, "sales_order_last_used") == expected

    # §2: the band starts at 1, not 0 — the click report encodes "none" as
    # NULL and never as a literal 0 (min value is 1 in all 22,224 rows), so a
    # literal 0 is not "has activity" and must not score the first tier.
    @pytest.mark.parametrize("count,expected", [
        (0, 0), (1, 5), (5, 5), (6, 15), (10, 15), (11, 25), (100, 25), (None, 0),
    ])
    def test_sales_order_count(self, count, expected):
        # G1: count points are only AWARDED when the row owns its (here trivial,
        # context-free) most-recent year, so a year must be present for the band
        # mapping to apply. The band VALUES themselves are unchanged.
        assert _points(
            {"last_order_year": Y0, "order_count": count}, "sales_order_count"
        ) == expected

    @pytest.mark.parametrize("year,expected", [
        (Y0, 20), (Y1, 15), (Y2, 10), (Y3, 5), (Y6, 0), (None, 0),
    ])
    def test_partner_last_used(self, year, expected):
        assert _points(
            {"partner_last_order_year": year}, "sales_order_partner_last_used"
        ) == expected

    # UNCONFIRMED: partner count tiers mirror sales_order_count. CONFIRM w/ Bernd.
    @pytest.mark.parametrize("count,expected", [
        (0, 0), (1, 5), (5, 5), (6, 15), (10, 15), (11, 25), (None, 0),
    ])
    def test_partner_count(self, count, expected):
        # G1: partner count mirrors the recency gate — a partner year must be
        # present for the (unchanged) band values to apply.
        assert _points(
            {"partner_last_order_year": Y0, "partner_order_count": count},
            "sales_order_partner_count",
        ) == expected

    # §2: same NULL-encoding evidence as the two count ladders above.
    @pytest.mark.parametrize("count,expected", [
        (0, 0), (1, 5), (3, 5), (4, 12), (8, 12), (9, 20), (15, 20), (16, 30),
        (None, 0),
    ])
    def test_equipment_count(self, count, expected):
        assert _points({"equipment_count": count}, "equipment_count") == expected

    # §3: two explicit values. "Yes" scores a KNOWN zero (no warning), the
    # same way blocked: 0 does — see test_known_zero_bands_do_not_warn.
    @pytest.mark.parametrize("band,expected", [
        ("No", 15), ("Yes", 0), (None, 0),
    ])
    def test_sleeping_bands(self, band, expected):
        assert _points({"sleeping_band": band}, "sleeping_customer") == expected

    @pytest.mark.parametrize("status,expected", [
        ("active", 10), ("blocked", 0), (None, 0),
    ])
    def test_customer_status(self, status, expected):
        assert _points({"customer_status": status}, "customer_status") == expected

    @pytest.mark.parametrize("group,expected", [
        ("DRIT", 20), ("0002", 15), ("SHIP2", 15), ("0003", 10), ("0004", 10),
        ("0005", 5), ("LIEF", 5), ("MLIEF", 5),
        ("DBRU", 0),  # parked -> 0 (explicit anything-else band)
        ("XXXX", 0), (None, 0),
    ])
    def test_account_group(self, group, expected):
        assert _points({"account_group": group}, "account_group") == expected

    @pytest.mark.parametrize("codes,expected", [
        (None, 0),                       # 0 codes -> 0
        ("1001", 5),                     # 1 -> 5
        ("1001;1002", 15),               # 2 -> 15
        ("1001;1002;1003;1004", 15),     # company-code edge: 4 sits in the 15 band
        ("1001;1002;1003;1004;1005", 25),  # 5 -> 25
    ])
    def test_company_code_count(self, codes, expected):
        assert _points(
            {"company_code_consolidated": codes}, "company_code_count"
        ) == expected

    # UNCONFIRMED bonus value (10); sales org has no standalone tier.
    def test_combined_presence_bonus(self):
        assert _points(
            {"company_code_consolidated": "1001", "sales_org_consolidated": "2001"},
            "combined_presence_bonus",
        ) == 10
        assert _points(
            {"company_code_consolidated": "1001"}, "combined_presence_bonus"
        ) == 0
        assert _points(
            {"sales_org_consolidated": "2001"}, "combined_presence_bonus"
        ) == 0

    def test_salesforce_instances_x10_non_empty_only(self):
        breakdown, _ = _score({
            "salesforce_ids": ["a", None, "", "  ", "b", None, "c", None],
        })
        assert breakdown["salesforce_instance_count"] == 30


# ---------------------------------------------------------------------------
# Coercion — permissive, never raises
# ---------------------------------------------------------------------------

class TestCoercion:
    def test_all_none_scores_zero_no_exception(self):
        breakdown, warnings = _score({})
        assert sum(breakdown.values()) == 0
        assert warnings == []
        # Every criterion key is present for a column-stable audit trail.
        assert set(breakdown) == set(BREAKDOWN_COLUMNS)

    def test_unrecognized_enums_warn_not_422(self):
        # "Yes" is now a KNOWN sleeping band (§3), so the unrecognised case
        # needs a genuinely unknown value.
        breakdown, warnings = _score({
            "sleeping_band": "dormant", "customer_status": "n/a",
        })
        assert breakdown["sleeping_customer"] == 0
        assert breakdown["customer_status"] == 0
        assert any("sleeping_band 'dormant' unrecognized" in w for w in warnings)
        assert any("customer_status 'n/a' unrecognized" in w for w in warnings)

    def test_known_zero_bands_do_not_warn(self):
        """§3: a known value scoring a known zero is as clean as an absent one.

        sleeping_band "Yes" and customer_status "blocked" both carry an
        explicit 0 band, so neither may produce an "unrecognized" warning —
        that channel is reserved for values the table has never heard of.
        """
        breakdown, warnings = _score({
            "sleeping_band": "Yes", "customer_status": "blocked",
        })
        assert breakdown["sleeping_customer"] == 0
        assert breakdown["customer_status"] == 0
        assert warnings == []

    @pytest.mark.parametrize("status", [" Active ", "ACTIVE", "active"])
    def test_status_whitespace_case_variants(self, status):
        assert _points({"customer_status": status}, "customer_status") == 10

    @pytest.mark.parametrize("band", ["no", "NO", " No "])
    def test_sleeping_case_variants(self, band):
        assert _points({"sleeping_band": band}, "sleeping_customer") == 15

    def test_excel_float_hits_year_band(self):
        assert _points({"last_order_year": float(Y0)}, "sales_order_last_used") == 20
        assert _points({"last_order_year": f"{Y0}.0"}, "sales_order_last_used") == 20

    def test_non_numeric_scores_zero_with_warning(self):
        # G1: field renamed order_count -> orders_in_last_used_year (old name
        # still accepted on input via alias; coercion warning uses the new name).
        breakdown, warnings = _score({"order_count": "lots"})
        assert breakdown["sales_order_count"] == 0
        assert any("orders_in_last_used_year" in w for w in warnings)

    def test_absence_is_not_activity(self):
        """Absence never defaults into a scoring band — and neither does a
        literal 0.

        The count/equipment half of this used to be the opposite decision: the
        bands started at 0, so a literal ``equipment_count=0`` scored the first
        tier (5) while a blank scored 0. The click report settles it — it
        encodes "none" as NULL, never as 0: across all 22,224 rows of
        ``US_Qlic report data_2026-07-30.xlsx`` the minimum value is 1 in
        Sales Order Total Count, Sales Order Partner Total Count AND Equipment
        Total Count, with zero occurrences of a literal 0. So the two
        encodings of the same fact ("no equipment") were scoring differently.
        The bands now start at 1 and mean "has activity", which is what the
        data can actually support. This is a deliberate deviation from Bernd's
        literal "0 to 3 is 5%" (BerndScoring1 19:24) — see 12_RATIONALE.md.
        """
        breakdown, _ = _score({})
        assert breakdown["customer_status"] == 0
        assert breakdown["sleeping_customer"] == 0
        assert breakdown["equipment_count"] == 0
        # A literal 0 now scores the same as absence; 1 opens the first band.
        assert _points({"equipment_count": 0}, "equipment_count") == 0
        assert _points({"equipment_count": 1}, "equipment_count") == 5


# ---------------------------------------------------------------------------
# Derived counts
# ---------------------------------------------------------------------------

class TestDerivedCounts:
    @pytest.mark.parametrize("value,expected", [
        (None, 0), ("", 0), ("1003;1017;1042", 3),
        ("1003;1017;", 2),   # trailing ";" ignored
        (" 1003 ; ;1017", 2),
    ])
    def test_company_code_split(self, value, expected):
        row = ScoringRow(row_id="1", company_code_consolidated=value)
        assert derived_counts(row)[0] == expected

    def test_sales_org_and_sf_counts(self):
        row = ScoringRow(
            row_id="1",
            sales_org_consolidated="2001;2002",
            salesforce_ids=[None, "x", "", "y", None, None, None, None],
        )
        assert derived_counts(row) == (0, 2, 2)


# ---------------------------------------------------------------------------
# Election
# ---------------------------------------------------------------------------

def _cluster_row(row_id, cluster_id="C1", **kw):
    return ScoringRow(row_id=row_id, cluster_id=cluster_id, **kw)


class TestElection:
    def test_highest_score_wins(self):
        results = elect_golden_records([
            _cluster_row("1", last_order_year=Y0, order_count=20),   # 20+25
            _cluster_row("2", last_order_year=Y3),                   # 5
        ], WEIGHTS)
        by = _by_row(results)
        assert by["1"].is_golden_record is True
        assert by["1"].golden_record_id == "1"
        assert by["2"].is_golden_record is False
        assert by["2"].golden_record_id == "1"
        assert by["1"].election_status == "proposed"
        assert by["2"].election_status == "proposed"

    def test_unique_row_self_references(self):
        results = elect_golden_records(
            [ScoringRow(row_id="9", cluster_id=None)], WEIGHTS
        )
        (r,) = results
        assert r.is_golden_record is True
        assert r.golden_record_id == "9"
        assert r.election_status == "unique"

    def test_single_member_cluster_degrades_to_unique(self):
        results = elect_golden_records([_cluster_row("9", cluster_id="C1")], WEIGHTS)
        (r,) = results
        assert r.is_golden_record is True
        assert r.golden_record_id == "9"
        assert r.election_status == "unique"

    def test_blocked_scores_zero_but_can_win(self):
        results = elect_golden_records([
            _cluster_row("1", customer_status="blocked",
                         last_order_year=Y0, order_count=20),
            _cluster_row("2", customer_status="active"),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["1"].score_breakdown["customer_status"] == 0
        assert by["1"].is_golden_record is True     # eligibility unaffected
        assert by["1"].election_status == "proposed"  # not all blocked

    def test_all_blocked_cluster_manual_review(self):
        results = elect_golden_records([
            _cluster_row("1", customer_status="blocked", last_order_year=Y0),
            _cluster_row("2", customer_status="BLOCKED "),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["1"].is_golden_record is True     # still elects a winner
        assert by["1"].election_status == "manual_review"
        assert by["2"].election_status == "manual_review"

    def test_low_confidence_merge_demoted_to_manual_review(self):
        """Q2: a merge below the confidence threshold keeps its membership but
        enters election as manual_review — a human confirms before a block."""
        results = elect_golden_records([
            _cluster_row("1", confidence=0.90, last_order_year=Y0, order_count=20),
            _cluster_row("2", confidence=0.90),
        ], WEIGHTS, confidence_threshold=0.95)
        by = _by_row(results)
        # Membership + winner still computed…
        assert by["1"].is_golden_record is True
        assert by["1"].golden_record_id == "1"
        assert by["2"].golden_record_id == "1"
        # …but the whole cluster is demoted.
        assert by["1"].election_status == "manual_review"
        assert by["2"].election_status == "manual_review"

    def test_confident_merge_stays_proposed(self):
        """A merge at/above threshold is a normal proposal."""
        results = elect_golden_records([
            _cluster_row("1", confidence=0.97, last_order_year=Y0, order_count=20),
            _cluster_row("2", confidence=0.96),
        ], WEIGHTS, confidence_threshold=0.95)
        by = _by_row(results)
        assert by["1"].election_status == "proposed"
        assert by["2"].election_status == "proposed"

    def test_lowest_member_confidence_gates_the_cluster(self):
        """Per-cluster confidence is the LOWEST member's — one low-confidence
        join demotes the merge even if others are confident."""
        results = elect_golden_records([
            _cluster_row("1", confidence=0.99, last_order_year=Y0),
            _cluster_row("2", confidence=0.80),  # dragged the merge down
        ], WEIGHTS, confidence_threshold=0.95)
        assert all(r.election_status == "manual_review" for r in results)

    def test_none_confidence_never_gates(self):
        """A deterministic identical-collapse (no LLM merge → confidence None)
        is fully trusted: it elects as proposed, never gated."""
        results = elect_golden_records([
            _cluster_row("1", confidence=None, last_order_year=Y0),
            _cluster_row("2", confidence=None),
        ], WEIGHTS, confidence_threshold=0.95)
        assert all(r.election_status == "proposed" for r in results)

    def test_confidence_threshold_from_env(self, monkeypatch):
        """The threshold is env-overridable without re-running the LLM."""
        monkeypatch.setenv("CONFIDENCE_MERGE_THRESHOLD", "0.85")
        results = elect_golden_records([
            _cluster_row("1", confidence=0.90, last_order_year=Y0),
            _cluster_row("2", confidence=0.90),
        ], WEIGHTS)  # 0.90 >= 0.85 → proposed
        assert all(r.election_status == "proposed" for r in results)

    def test_inherited_manual_review_survives_confident_neighbours(self):
        """Q3: a row clustering flagged manual_review can NEVER leave election
        as proposed/unique — even surrounded by confident rows. Election only
        ever propagates uncertainty, never upgrades it."""
        results = elect_golden_records([
            # A confident cluster…
            _cluster_row("1", cluster_id="C1", confidence=0.99, last_order_year=Y0),
            _cluster_row("2", cluster_id="C1", confidence=0.99),
            # …and a lone row clustering could not resolve.
            ScoringRow(row_id="9", cluster_id=None, routing="manual_review",
                       last_order_year=Y0, order_count=20),
        ], WEIGHTS, confidence_threshold=0.95)
        by = _by_row(results)
        assert by["1"].election_status == "proposed"   # neighbours unaffected
        assert by["2"].election_status == "proposed"
        assert by["9"].election_status == "manual_review"  # never upgraded to unique

    def test_inherited_manual_review_demotes_whole_cluster(self):
        """An uncertain member propagates manual_review to its cluster, even at
        high confidence and with no blocked members."""
        results = elect_golden_records([
            _cluster_row("1", cluster_id="C1", confidence=0.99,
                         routing="manual_review", last_order_year=Y0),
            _cluster_row("2", cluster_id="C1", confidence=0.99, routing="cluster"),
        ], WEIGHTS, confidence_threshold=0.95)
        by = _by_row(results)
        assert by["1"].election_status == "manual_review"
        assert by["2"].election_status == "manual_review"
        # Winner is still elected — membership is preserved, only routing demoted.
        assert by["1"].is_golden_record is True
        assert by["2"].golden_record_id == "1"

    def test_manual_review_singleton_not_upgraded_in_summary(self):
        """A lone manual_review row is counted as manual_review, not unique, and
        does not mint a phantom cluster."""
        from dedup.scoring import build_summary
        results = elect_golden_records([
            ScoringRow(row_id="9", cluster_id=None, routing="manual_review"),
            ScoringRow(row_id="8", cluster_id=None, routing="unique"),
        ], WEIGHTS)
        summary = build_summary(results)
        assert summary.rows_manual_review == 1
        assert summary.rows_unique == 1
        assert summary.clusters == 0
        assert summary.all_blocked_clusters == 0

    def test_approval_and_proposed_golden_fields(self):
        """Q5: cluster rows get approval_status='proposed' and a
        proposed_golden_id; unique rows get neither (nothing to approve)."""
        results = elect_golden_records([
            _cluster_row("1", cluster_id="C1", last_order_year=Y0, order_count=20),
            _cluster_row("2", cluster_id="C1"),
            ScoringRow(row_id="9", cluster_id=None),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["1"].approval_status == "proposed"
        assert by["2"].approval_status == "proposed"
        assert by["1"].proposed_golden_id == "1" == by["2"].proposed_golden_id
        # Unique: nothing to approve.
        assert by["9"].approval_status is None
        assert by["9"].proposed_golden_id is None
        assert by["9"].election_status == "unique"

    def test_apply_approval_promotes_golden_and_rejects(self):
        """Q5: approving a cluster sets approval_status and promotes the proposed
        winner into the golden fields; rejecting only sets the status. Unknown
        cluster raises."""
        from dedup.scoring import apply_approval, ClusterNotFoundError
        results = elect_golden_records([
            _cluster_row("1", cluster_id="C1", customer_status="blocked",
                         last_order_year=Y0),
            _cluster_row("2", cluster_id="C1", customer_status="blocked"),
        ], WEIGHTS)  # all-blocked → manual_review
        approved, updated = apply_approval(results, "C1", "approved")
        by = {r.row_id: r for r in approved}
        assert set(updated) == {"1", "2"}
        assert all(r.approval_status == "approved" for r in approved)
        assert by["1"].is_golden_record is True and by["1"].golden_record_id == "1"
        assert by["2"].is_golden_record is False and by["2"].golden_record_id == "1"

        rejected, _ = apply_approval(results, "C1", "rejected")
        assert all(r.approval_status == "rejected" for r in rejected)
        with pytest.raises(ClusterNotFoundError):
            apply_approval(results, "NOPE", "approved")

    def test_duplicate_row_id_raises(self):
        with pytest.raises(DuplicateRowIdError) as exc:
            elect_golden_records([
                ScoringRow(row_id="7"), ScoringRow(row_id="7"),
            ], WEIGHTS)
        assert exc.value.row_ids == ["7"]

    def test_empty_rows(self):
        assert elect_golden_records([], WEIGHTS) == []

    # UNCONFIRMED tie-break ordering (confirm with Bernd): total score, most
    # recent last_order_year, equipment_count, company_code_count, lowest
    # row_id (numeric when all ids in the cluster parse as ints).

    def test_tiebreak_recent_year_wins_on_equal_score(self):
        # Y5 and Y4 both score 0 -> equal total, more recent year wins.
        rows = [
            _cluster_row("1", last_order_year=Y5),
            _cluster_row("2", last_order_year=Y4),
        ]
        by = _by_row(elect_golden_records(rows, WEIGHTS))
        assert by["2"].is_golden_record is True

    def test_tiebreak_equipment_on_equal_score_and_year(self):
        # 4 and 8 sit in the same 12-point band -> equal score and year,
        # higher raw equipment_count wins.
        rows = [
            _cluster_row("1", last_order_year=Y4, equipment_count=4),
            _cluster_row("2", last_order_year=Y4, equipment_count=8),
        ]
        by = _by_row(elect_golden_records(rows, WEIGHTS))
        assert by["2"].is_golden_record is True

    def test_tiebreak_company_codes_on_equal_score(self):
        # Row 1: 1 code (5) + active status (10) = 15.
        # Row 2: 2 codes (15). Equal totals, equal year/equipment (absent) ->
        # higher company_code_count wins.
        rows = [
            _cluster_row("1", company_code_consolidated="1001",
                         customer_status="active"),
            _cluster_row("2", company_code_consolidated="1001;1002"),
        ]
        by = _by_row(elect_golden_records(rows, WEIGHTS))
        assert by["2"].is_golden_record is True

    def test_tiebreak_lowest_numeric_row_id_last(self):
        # Everything equal -> lowest row_id compared NUMERICALLY
        # ("3" < "20" as ints even though "20" < "3" lexically).
        rows = [
            _cluster_row("20", last_order_year=Y0, equipment_count=2),
            _cluster_row("3", last_order_year=Y0, equipment_count=2),
        ]
        by = _by_row(elect_golden_records(rows, WEIGHTS))
        assert by["3"].is_golden_record is True

    def test_tiebreak_lexical_when_non_numeric_ids(self):
        rows = [
            _cluster_row("BP-10", last_order_year=Y0),
            _cluster_row("BP-2", last_order_year=Y0),
        ]
        by = _by_row(elect_golden_records(rows, WEIGHTS))
        # "BP-10" < "BP-2" lexically.
        assert by["BP-10"].is_golden_record is True

    def test_winner_invariant_under_shuffle(self):
        rows = [
            _cluster_row(str(i), cluster_id="C1",
                         last_order_year=Y3 + (i % 4),
                         order_count=i, equipment_count=i % 7)
            for i in range(1, 12)
        ] + [
            ScoringRow(row_id=str(i)) for i in range(100, 105)
        ] + [
            _cluster_row(str(i), cluster_id="C2", last_order_year=Y0)
            for i in range(200, 204)
        ]
        baseline = {
            r.row_id: (r.is_golden_record, r.golden_record_id, r.election_status)
            for r in elect_golden_records(rows, WEIGHTS)
        }
        rng = random.Random(42)
        for _ in range(10):
            shuffled = rows[:]
            rng.shuffle(shuffled)
            outcome = {
                r.row_id: (r.is_golden_record, r.golden_record_id, r.election_status)
                for r in elect_golden_records(shuffled, WEIGHTS)
            }
            assert outcome == baseline

    def test_table_invariant(self):
        rows = [
            _cluster_row("1", last_order_year=Y0),
            _cluster_row("2"),
            _cluster_row("3", cluster_id="C2", equipment_count=20),
            _cluster_row("4", cluster_id="C2"),
            ScoringRow(row_id="5"),
            _cluster_row("6", cluster_id="C3"),  # degrades to unique
        ]
        results = elect_golden_records(rows, WEIGHTS)
        by = _by_row(results)
        for r in results:
            if r.is_golden_record:
                assert r.golden_record_id == r.row_id  # survivors self-reference
            else:
                target = by[r.golden_record_id]
                assert target.is_golden_record is True
                assert target.cluster_id == r.cluster_id  # resolves within cluster

    def test_score_equals_breakdown_sum(self):
        results = elect_golden_records([
            _cluster_row("1", last_order_year=Y0, order_count=12,
                         sleeping_band="No", customer_status="active",
                         account_group="DRIT",
                         company_code_consolidated="1;2;3",
                         sales_org_consolidated="9",
                         salesforce_ids=["a", "b"]),
        ], WEIGHTS)
        (r,) = results
        assert r.score == sum(r.score_breakdown.values())
        # year 20 + orders 25 + partner 0/0 + equipment 0 + sleeping 15
        # + status 10 + DRIT 20 + 3 codes 15 + bonus 10 + 2 SF ids 20
        assert r.score == 135


# ---------------------------------------------------------------------------
# JSON endpoint
# ---------------------------------------------------------------------------

from api.app import app  # noqa: E402


@pytest.fixture
def transport():
    return ASGITransport(app=app)


class TestIssues:
    def test_detect_issues_covers_each_type(self):
        rows = [
            # cA: low-confidence merge (0.90 < 0.95); r1 has a score, r2 doesn't.
            _cluster_row("1", cluster_id="cA", confidence=0.90, last_order_year=Y0),
            _cluster_row("2", cluster_id="cA", confidence=0.90),
            # cB: all blocked + zero scores → all_blocked + empty_payload + tie.
            _cluster_row("3", cluster_id="cB", customer_status="blocked"),
            _cluster_row("4", cluster_id="cB", customer_status="blocked"),
            # cC: a deterministic identity-split reasoning → verdict_contradiction.
            _cluster_row("5", cluster_id="cC", last_order_year=Y0,
                         reasoning="Split: different non-empty ROR ids (a, b) "
                                   "indicate different entities."),
            _cluster_row("6", cluster_id="cC", last_order_year=Y3),
        ]
        results = elect_golden_records(rows, WEIGHTS, confidence_threshold=0.95)
        issues = detect_issues(rows, results, confidence_threshold=0.95)
        by_type = {}
        for i in issues:
            by_type.setdefault(i.issue_type, []).append(i)

        assert "low_confidence_merge" in by_type
        assert by_type["low_confidence_merge"][0].cluster_id == "cA"
        assert "all_blocked_cluster" in by_type
        assert "empty_scoring_payload" in by_type
        assert "tiebreak_decided" in by_type
        vc = by_type["verdict_contradiction"]
        assert vc[0].row_id == "5" and "Split:" in vc[0].detail

    def test_candidate_cap_exceeded_issue_from_reasoning_marker(self):
        """The adjudicator's candidate_cap_exceeded marker (persisted in the
        Reasoning column) surfaces as a scoring issue, one per capped block."""
        rows = [
            ScoringRow(row_id="1", cluster_id="C1", routing="manual_review",
                       reasoning="candidate_cap_exceeded: 80 candidate pairs exceed the per-block cap of 50"),
            ScoringRow(row_id="2", cluster_id="C1", routing="manual_review",
                       reasoning="candidate_cap_exceeded: 80 candidate pairs exceed the per-block cap of 50"),
        ]
        results = elect_golden_records(rows, WEIGHTS)
        caps = [i for i in detect_issues(rows, results)
                if i.issue_type == "candidate_cap_exceeded"]
        assert len(caps) == 1 and caps[0].cluster_id == "C1"

    def test_no_issues_on_clean_confident_cluster(self):
        rows = [
            _cluster_row("1", cluster_id="cA", confidence=0.99, last_order_year=Y0),
            _cluster_row("2", cluster_id="cA", confidence=0.99, last_order_year=Y3),
        ]
        results = elect_golden_records(rows, WEIGHTS, confidence_threshold=0.95)
        assert detect_issues(rows, results, confidence_threshold=0.95) == []


@pytest_asyncio.fixture
async def client(transport):
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c


class TestScoreEndpoint:
    @pytest.mark.asyncio
    async def test_duplicate_row_id_400_lists_ids(self, client):
        resp = await client.post("/api/dedup/score", json={"rows": [
            {"row_id": "7"}, {"row_id": "7"}, {"row_id": "8"}, {"row_id": "8"},
        ]})
        assert resp.status_code == 400
        assert "7" in resp.json()["detail"]
        assert "8" in resp.json()["detail"]

    @pytest.mark.asyncio
    async def test_empty_rows_200_zeroed_summary(self, client):
        resp = await client.post("/api/dedup/score", json={"rows": []})
        assert resp.status_code == 200
        data = resp.json()
        assert data["rows"] == []
        assert data["summary"]["rows_in"] == 0
        assert data["summary"]["clusters"] == 0

    @pytest.mark.asyncio
    async def test_approve_endpoint_promotes_and_echoes(self, client):
        """Q5: score a cluster, then approve it — approval_status flips to
        approved and the proposed winner is promoted into the golden fields."""
        scored = await client.post("/api/dedup/score", json={"rows": [
            {"row_id": "1", "cluster_id": "C1", "last_order_year": Y0,
             "customer_status": "blocked"},
            {"row_id": "2", "cluster_id": "C1", "customer_status": "blocked"},
        ]})
        rows = scored.json()["rows"]
        assert all(r["election_status"] == "manual_review" for r in rows)

        resp = await client.post("/api/dedup/approve", json={
            "cluster_id": "C1", "decision": "approved",
            "approver": "bernd", "rows": rows,
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["approver"] == "bernd"
        assert set(data["updated_row_ids"]) == {"1", "2"}
        # Output columns use the exact file headers ("Customer", not "row_id").
        by = {r["Customer"]: r for r in data["rows"]}
        assert all(r["approval_status"] == "approved" for r in data["rows"])
        assert by["1"]["is_golden_record"] is True
        assert by["2"]["golden_record_id"] == "1"

    @pytest.mark.asyncio
    async def test_score_endpoint_returns_issues(self, client):
        """Q7: the JSON score endpoint returns an issues list."""
        resp = await client.post("/api/dedup/score", json={"rows": [
            {"row_id": "1", "cluster_id": "C1", "confidence": 0.80,
             "last_order_year": Y0},
            {"row_id": "2", "cluster_id": "C1", "confidence": 0.80},
        ]})
        assert resp.status_code == 200
        issues = resp.json()["issues"]
        assert any(i["issue_type"] == "low_confidence_merge" for i in issues)
        assert all(set(i) == {"row_id", "cluster_id", "issue_type", "detail"}
                   for i in issues)

    @pytest.mark.asyncio
    async def test_approve_unknown_cluster_404(self, client):
        resp = await client.post("/api/dedup/approve", json={
            "cluster_id": "NOPE", "decision": "approved", "approver": "x",
            "rows": [{"row_id": "1", "cluster_id": "C1", "score": 0,
                      "is_golden_record": True, "election_status": "proposed",
                      "score_breakdown": {}}],
        })
        assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_dirty_values_do_not_422(self, client):
        resp = await client.post("/api/dedup/score", json={"rows": [
            {"row_id": "1", "cluster_id": "C1", "sleeping_band": "Yes",
             "customer_status": "parked", "last_order_year": "n/a"},
            {"row_id": "2", "cluster_id": "C1", "last_order_year": float(Y0)},
        ]})
        assert resp.status_code == 200
        data = resp.json()
        by = {r["Customer"]: r for r in data["rows"]}
        # Per-row warnings are not a file column (not serialized); the dirty
        # value is surfaced via the aggregate summary instead.
        assert by["2"]["is_golden_record"] is True
        assert data["summary"]["rows_with_warnings"] == 1

    @pytest.mark.asyncio
    async def test_summary_counts(self, client):
        resp = await client.post("/api/dedup/score", json={"rows": [
            {"row_id": "1", "cluster_id": "C1", "last_order_year": Y0},
            {"row_id": "2", "cluster_id": "C1"},
            {"row_id": "3"},
            {"row_id": "4", "cluster_id": "C2", "customer_status": "blocked"},
            {"row_id": "5", "cluster_id": "C2", "customer_status": "blocked"},
        ]})
        assert resp.status_code == 200
        s = resp.json()["summary"]
        assert s["rows_in"] == 5
        assert s["clusters"] == 2
        assert s["rows_elected"] == 2
        assert s["rows_duplicates"] == 2
        assert s["rows_unique"] == 1
        assert s["rows_manual_review"] == 2
        assert s["all_blocked_clusters"] == 1
        assert s["errors"] == 0


# ---------------------------------------------------------------------------
# File endpoint / workbook round-trip
# ---------------------------------------------------------------------------

# The 45 original columns of the scoring workbook (SAP export + Phase 1/2
# expectations) — must survive the round-trip untouched.
ORIGINAL_COLUMNS = [
    "Customer", "ECC Customer Number", "Central Deletion Flag", "Comments",
    "Account group", "Company Code", "Sales Organization",
    "Distribution Channel", "Division", "Name 1", "Name 2", "Name 3", "Name 4",
    "Contact", "Street 1", "House Number", "Street 2", "Street 3", "Street 4",
    "Street 5", "PO Box", "Country/Region Key", "Postal Code", "City", "Region",
    "Language Key", "Reconciliation acct", "Tax Jurisdiction",
    "Central delivery block", "Delivery Priority", "Shipping Conditions",
    "Delivering Plant", "Created On", "Created By", "VAT Registration No.",
    "Search Term 1", "Search Term 2", "Terms of Payment", "ror_id", "lei_id",
    "classification", "expected_cluster", "expected_routing",
    "expected_llm_flag", "test_category",
]

CRM_COLUMNS = [
    "Sales_Order_Last_Used", "Sales_Order_Total_Count",
    "Sales_Order_Partner_Last_Used", "Sales_Order_Partner_Total_Count",
    "Equipment_Total_Count", "SleepingCustomer", "CustomerStatus", "Is_ZFIS",
    "Company_Code_Consolidated", "Company_Code_Count", "Sales_Org_Consolidated",
    "Sales_Org_Count", "SF_ID_Biosystems", "SF_ID_AXS", "SF_ID_3", "SF_ID_4",
    "SF_ID_5", "SF_ID_6", "SF_ID_7", "SF_ID_8", "Salesforce_Instance_Count",
]

SCORE_COLUMNS = [
    "score_SalesOrderLastUsed", "score_SalesOrderCount",
    "score_SalesOrderPartnerLastUsed", "score_SalesOrderPartnerCount",
    "score_EquipmentCount", "score_SleepingCustomer", "score_CustomerStatus",
    "score_AccountGroup", "score_CompanyCodeCount", "score_CombinedPresence",
    "score_SalesforceInstances", "score_final",
]

ALL_HEADERS = ORIGINAL_COLUMNS + CRM_COLUMNS + SCORE_COLUMNS


def _data_row(customer, *, cluster="unique", routing="unique", year=None,
              orders=None, p_year=None, p_orders=None, equipment=None,
              sleeping=None, status=None, group="DRIT", codes=None, orgs=None,
              sf=(None,) * 8, name=None):
    values = {h: None for h in ALL_HEADERS}
    values.update({
        "Customer": customer, "Account group": group,
        "Name 1": name or f"Firm {customer}",
        "Street 1": "Hauptstrasse 1", "Postal Code": "10115", "City": "Berlin",
        "Country/Region Key": "DE",
        "expected_cluster": cluster, "expected_routing": routing,
        "Sales_Order_Last_Used": year, "Sales_Order_Total_Count": orders,
        "Sales_Order_Partner_Last_Used": p_year,
        "Sales_Order_Partner_Total_Count": p_orders,
        "Equipment_Total_Count": equipment, "SleepingCustomer": sleeping,
        "CustomerStatus": status, "Is_ZFIS": 0,
        "Company_Code_Consolidated": codes, "Sales_Org_Consolidated": orgs,
        "SF_ID_Biosystems": sf[0], "SF_ID_AXS": sf[1], "SF_ID_3": sf[2],
        "SF_ID_4": sf[3], "SF_ID_5": sf[4], "SF_ID_6": sf[5], "SF_ID_7": sf[6],
        "SF_ID_8": sf[7],
    })
    return [values[h] for h in ALL_HEADERS]


def _weights_sheet_rows():
    rows = [("Criterion", "Band", "Points", "Note")]
    for criterion, bands in WEIGHTS.items():
        for band, points in bands.items():
            rows.append((criterion, band, points, None))
    return rows


def _build_workbook(data_rows, *, weights_rows=None, drop_weights_sheet=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(ALL_HEADERS)
    for row in data_rows:
        ws.append(row)
    if not drop_weights_sheet:
        weights = wb.create_sheet("Weights")
        for row in (weights_rows or _weights_sheet_rows()):
            weights.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestScoreWorkbook:
    def test_manual_review_blanks_golden_in_file(self):
        """Q5: an all-blocked (manual_review) cluster leaves is_golden_record
        and golden_record_id EMPTY in the file; the proposal survives in
        proposed_golden_id, and approval_status is 'proposed'."""
        original = _build_workbook([
            _data_row("1", cluster="A1", routing="cluster", status="blocked",
                      year=Y0),
            _data_row("2", cluster="A1", routing="cluster", status="blocked"),
            _data_row("3"),  # unique
        ])
        output, _ = score_workbook(original)
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]

        def col(h):
            return headers.index(h) + 1

        # Row 2 (winner "1", row 3 in sheet): golden blanked, proposal kept.
        assert ws.cell(row=2, column=col("election_status")).value == "manual_review"
        assert ws.cell(row=2, column=col("is_golden_record")).value is None
        assert ws.cell(row=2, column=col("golden_record_id")).value is None
        assert ws.cell(row=2, column=col("proposed_golden_id")).value == "1"
        assert ws.cell(row=2, column=col("approval_status")).value == "proposed"
        # Unique row: golden filled, nothing to approve.
        assert ws.cell(row=4, column=col("is_golden_record")).value is True
        assert ws.cell(row=4, column=col("election_status")).value == "unique"
        assert ws.cell(row=4, column=col("approval_status")).value is None

    def test_issues_sheet_written_preserving_weights(self):
        """Q7: an Issues sheet is added (all_blocked_cluster) while the Weights
        sheet and original columns survive."""
        original = _build_workbook([
            _data_row("1", cluster="A1", routing="cluster", status="blocked",
                      year=Y0),
            _data_row("2", cluster="A1", routing="cluster", status="blocked"),
            _data_row("3"),
        ])
        output, _ = score_workbook(original)
        wb = load_workbook(io.BytesIO(output))
        assert "Issues" in wb.sheetnames
        assert "Weights" in wb.sheetnames  # untouched
        issues_ws = wb["Issues"]
        header = [c.value for c in issues_ws[1]]
        assert header == ["row_id", "cluster_id", "issue_type", "detail"]
        body = list(issues_ws.iter_rows(min_row=2, values_only=True))
        types = {r[2] for r in body}
        assert "all_blocked_cluster" in types

    def test_round_trip_preserves_weights_sheet_and_45_columns(self):
        original = _build_workbook([
            _data_row("72000001", cluster="A1", routing="cluster", year=Y0,
                      orders=12, sleeping="No", status="active", codes="1001",
                      orgs="2001"),
            _data_row("72000002", cluster="A1", routing="cluster", year=Y3),
            _data_row("72000003"),
        ])
        output, summary = score_workbook(original)
        wb = load_workbook(io.BytesIO(output))
        assert "Weights" in wb.sheetnames  # sheet survived

        weights_ws = wb["Weights"]
        assert [c.value for c in weights_ws[1]] == ["Criterion", "Band", "Points", "Note"]

        ws = wb["Sheet1"]
        headers = [c.value for c in ws[1]]
        assert headers[: len(ORIGINAL_COLUMNS)] == ORIGINAL_COLUMNS  # all 45, in place
        # Original cell values untouched.
        assert ws.cell(row=2, column=headers.index("Customer") + 1).value == "72000001"
        assert ws.cell(row=2, column=headers.index("Name 1") + 1).value == "Firm 72000001"
        assert ws.cell(row=2, column=headers.index("expected_cluster") + 1).value == "A1"

        # Election columns appended and filled; native Excel booleans.
        def col(h):
            return headers.index(h) + 1

        assert ws.cell(row=2, column=col("is_golden_record")).value is True
        assert ws.cell(row=3, column=col("is_golden_record")).value is False
        assert ws.cell(row=3, column=col("golden_record_id")).value == "72000001"
        assert ws.cell(row=2, column=col("election_status")).value == "proposed"
        assert ws.cell(row=4, column=col("election_status")).value == "unique"
        assert ws.cell(row=4, column=col("golden_record_id")).value == "72000003"

        # score_final is a plain value equal to the sum of the score_* cells.
        for r in (2, 3, 4):
            parts = sum(
                ws.cell(row=r, column=col(h)).value or 0
                for h in SCORE_COLUMNS if h != "score_final"
            )
            final = ws.cell(row=r, column=col("score_final")).value
            assert isinstance(final, int)
            assert final == parts

        # Derived counts written as plain values.
        assert ws.cell(row=2, column=col("Company_Code_Count")).value == 1
        assert ws.cell(row=2, column=col("Sales_Org_Count")).value == 1
        assert ws.cell(row=2, column=col("Salesforce_Instance_Count")).value == 0

        assert summary.rows_in == 3
        assert summary.clusters == 1
        assert summary.warnings == []  # intact Weights sheet accepted silently

    def test_corrupted_weights_sheet_falls_back_wholesale(self):
        # Drop one band -> the WHOLE sheet must be ignored (never merged),
        # scoring falls back to dedup/weights.json, and the summary warns.
        broken = [r for r in _weights_sheet_rows() if r[1] != "6-10"]
        original = _build_workbook(
            [_data_row("1", year=Y0)], weights_rows=broken
        )
        output, summary = score_workbook(original)
        assert any("Weights sheet ignored" in w for w in summary.warnings)
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]
        assert ws.cell(row=2, column=headers.index("score_SalesOrderLastUsed") + 1).value == 20

    def test_weights_sheet_override_applies_wholesale(self):
        retuned = [
            (c, b, (40 if (c, b) == ("sales_order_last_used", "0") else p), n)
            for c, b, p, n in _weights_sheet_rows()
        ]
        output, summary = score_workbook(
            _build_workbook([_data_row("1", year=Y0)], weights_rows=retuned)
        )
        assert summary.warnings == []
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]
        assert ws.cell(row=2, column=headers.index("score_SalesOrderLastUsed") + 1).value == 40

    def test_blank_customer_skipped_and_counted(self):
        rows = [
            _data_row("1", year=Y0),
            _data_row(None, year=Y2),  # blank Customer -> summary.errors
            _data_row("2", year=Y3),
        ]
        output, summary = score_workbook(_build_workbook(rows))
        assert summary.errors == 1
        assert summary.rows_in == 3
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]
        # The skipped row's score cells stay empty; others are filled
        # (year band + the helper's default DRIT account group = 20).
        col = headers.index("score_final") + 1
        assert ws.cell(row=2, column=col).value == 40   # Y0 (20) + DRIT (20)
        assert ws.cell(row=3, column=col).value is None
        assert ws.cell(row=4, column=col).value == 25   # Y3 (5) + DRIT (20)

    def test_manual_review_routing_keeps_cluster_membership(self):
        # Adjudicator semantics: manual_review means membership certain,
        # only a merge was uncertain -> expected_cluster is used.
        rows = [
            _data_row("1", cluster="M1", routing="manual_review", year=Y0),
            _data_row("2", cluster="M1", routing="manual_review"),
        ]
        _, summary = score_workbook(_build_workbook(rows))
        assert summary.clusters == 1
        assert summary.rows_elected == 1
        assert summary.rows_duplicates == 1

    def test_duplicate_customer_raises(self):
        rows = [_data_row("1"), _data_row("1")]
        with pytest.raises(DuplicateRowIdError):
            score_workbook(_build_workbook(rows))

    def test_production_cluster_columns(self):
        # The dedup stage appends "Routing" + "Cluster ID" (integer ids);
        # score/file must consume its output directly.
        wb = Workbook()
        ws = wb.active
        ws.append(["Customer", "Sales_Order_Last_Used", "Routing", "Cluster ID"])
        ws.append(["1", Y0, "cluster", 3])
        ws.append(["2", Y3, "cluster", 3])
        ws.append(["3", Y1, "unique", None])
        buffer = io.BytesIO()
        wb.save(buffer)

        _, summary = score_workbook(buffer.getvalue())
        assert summary.clusters == 1
        assert summary.rows_elected == 1
        assert summary.rows_duplicates == 1
        assert summary.rows_unique == 1

    def test_production_pair_beats_expected_pair(self):
        # A dedup output run on the test fixture carries BOTH pairs; the
        # production Routing/Cluster ID pair must win (and never be mixed
        # with the expected_* pair).
        wb = Workbook()
        ws = wb.active
        ws.append(["Customer", "expected_routing", "expected_cluster",
                   "Routing", "Cluster ID"])
        ws.append(["1", "cluster", "A1", "unique", None])
        ws.append(["2", "cluster", "A1", "unique", None])
        buffer = io.BytesIO()
        wb.save(buffer)

        _, summary = score_workbook(buffer.getvalue())
        # expected_* said cluster, production said unique -> unique wins.
        assert summary.clusters == 0
        assert summary.rows_unique == 2

    @pytest.mark.asyncio
    async def test_file_endpoint_end_to_end(self, client):
        contents = _build_workbook([
            _data_row("1", cluster="A1", routing="cluster", year=Y0),
            _data_row("2", cluster="A1", routing="cluster"),
        ])
        resp = await client.post(
            "/api/dedup/score/file",
            files={"file": ("scores.xlsx", contents,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert "scores_scored.xlsx" in resp.headers["content-disposition"]
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Weights" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_file_endpoint_duplicate_customer_400(self, client):
        contents = _build_workbook([_data_row("1"), _data_row("1")])
        resp = await client.post(
            "/api/dedup/score/file",
            files={"file": ("scores.xlsx", contents,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400
        assert "1" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# Pipeline field preservation: /enrich/file -> /api/dedup/file -> score/file
# ---------------------------------------------------------------------------

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestPipelineFieldPreservation:
    """Each stage's output feeds the next, so no stage may drop the CRM
    scoring columns or the Weights sheet before they reach score/file."""

    @pytest.mark.asyncio
    async def test_enrich_dedup_score_chain(self, client):
        # Two rows share a name + address -> identical signature -> the dedup
        # stage clusters them deterministically (no LLM merge needed, so the
        # conservative mock is fine). The third row is unique.
        rows = [
            _data_row("72000001", name="Acme Labs GmbH", year=Y0, orders=12,
                      sleeping="No", status="active", codes="1001;1002",
                      orgs="2001"),
            _data_row("72000002", name="Acme Labs GmbH", year=Y3),
            _data_row("72000003", name="Zeta Institut", year=Y1),
        ]
        contents = _build_workbook(rows)

        # Stage 1 — enrichment must carry the unconsumed CRM columns and the
        # Weights sheet through to its output.
        r1 = await client.post(
            "/enrich/file", files={"file": ("pipe.xlsx", contents, XLSX_MEDIA)}
        )
        assert r1.status_code == 200
        wb1 = load_workbook(io.BytesIO(r1.content))
        assert "Weights" in wb1.sheetnames
        ws1 = wb1.worksheets[0]
        h1 = [c.value for c in ws1[1]]
        for header in ("Sales_Order_Last_Used", "Sales_Order_Total_Count",
                       "SleepingCustomer", "CustomerStatus",
                       "Company_Code_Consolidated", "Sales_Org_Consolidated",
                       "SF_ID_Biosystems", "SF_ID_8"):
            assert header in h1, f"enrichment dropped {header}"
        col1 = {h: i + 1 for i, h in enumerate(h1)}
        assert str(ws1.cell(2, col1["Sales_Order_Last_Used"]).value) == str(Y0)
        assert ws1.cell(2, col1["SleepingCustomer"]).value == "No"

        # Stage 2 — dedup echoes every column, appends Routing/Cluster ID,
        # and keeps the Weights sheet.
        r2 = await client.post(
            "/api/dedup/file",
            files={"file": ("pipe_enriched.xlsx", r1.content, XLSX_MEDIA)},
        )
        assert r2.status_code == 200
        wb2 = load_workbook(io.BytesIO(r2.content))
        assert "Weights" in wb2.sheetnames
        ws2 = wb2.worksheets[0]
        h2 = [c.value for c in ws2[1]]
        assert "Routing" in h2 and "Cluster ID" in h2
        assert "Sales_Order_Last_Used" in h2

        # Stage 3 — scoring consumes the dedup output directly, using the
        # production Routing/Cluster ID pair.
        r3 = await client.post(
            "/api/dedup/score/file",
            files={"file": ("pipe_dedup.xlsx", r2.content, XLSX_MEDIA)},
        )
        assert r3.status_code == 200
        wb3 = load_workbook(io.BytesIO(r3.content))
        assert "Weights" in wb3.sheetnames
        ws3 = wb3.worksheets[0]
        h3 = [c.value for c in ws3[1]]
        col3 = {h: i + 1 for i, h in enumerate(h3)}
        by = {}
        for r in range(2, ws3.max_row + 1):
            rid = str(ws3.cell(r, col3["Customer"]).value)
            by[rid] = {
                h: ws3.cell(r, col3[h]).value
                for h in ("is_golden_record", "golden_record_id",
                          "election_status", "score_final")
            }
        assert by["72000001"]["election_status"] == "proposed"
        assert by["72000002"]["election_status"] == "proposed"
        assert by["72000001"]["is_golden_record"] is True   # higher score
        assert by["72000002"]["is_golden_record"] is False
        assert by["72000002"]["golden_record_id"] == "72000001"
        assert by["72000003"]["election_status"] == "unique"
        # CRM values survived both hops and actually scored.
        assert by["72000001"]["score_final"] >= 20 + 25 + 15 + 10  # year+orders+sleeping+status


class TestEdgeCases:
    """Q8 — robustness edge cases (offline, no LLM)."""

    def test_zero_signal_cluster_is_manual_review(self):
        """A cluster whose every member scores 0 (all-None payload) elects a
        winner by tie-break only → manual_review + empty_scoring_payload issue."""
        rows = [_cluster_row("1", cluster_id="C1"), _cluster_row("2", cluster_id="C1")]
        results = elect_golden_records(rows, WEIGHTS)
        assert all(r.score == 0 for r in results)
        assert all(r.election_status == "manual_review" for r in results)
        issues = detect_issues(rows, results)
        assert any(i.issue_type == "empty_scoring_payload" for i in issues)

    def test_mixed_numeric_and_lexical_row_ids_no_raise(self):
        """Non-numeric row_ids exercise the lexical tie-break; a cluster mixing
        numeric and non-numeric ids must not raise."""
        rows = [
            _cluster_row("DE-0001", cluster_id="C1", last_order_year=Y0),
            _cluster_row("100", cluster_id="C1", last_order_year=Y0),  # tie on score
        ]
        results = elect_golden_records(rows, WEIGHTS)  # must not raise
        winners = [r for r in results if r.is_golden_record]
        assert len(winners) == 1
        # Lexical order: "100" < "DE-0001", so "100" wins the tie.
        assert winners[0].row_id == "100"

    def test_weights_retune_flips_winner_and_changes_version(self):
        """Same cluster, different weights → possibly different winner, and a
        different scored_with_weights_version stamped on every row."""
        rows = [
            _cluster_row("1", cluster_id="C1", last_order_year=Y0),
            _cluster_row("2", cluster_id="C1", equipment_count=50),
        ]
        w_year = json.loads(json.dumps(WEIGHTS))
        w_equip = json.loads(json.dumps(WEIGHTS))
        for band in w_year["equipment_count"]:
            w_year["equipment_count"][band] = 0        # year dominates
        for band in w_equip["sales_order_last_used"]:
            w_equip["sales_order_last_used"][band] = 0
        for band in w_equip["equipment_count"]:
            w_equip["equipment_count"][band] = 999      # equipment dominates

        by_year = _by_row(elect_golden_records(rows, w_year))
        by_equip = _by_row(elect_golden_records(rows, w_equip))
        assert by_year["1"].is_golden_record is True    # year winner
        assert by_equip["2"].is_golden_record is True    # equipment winner
        # Version fingerprints differ and are stamped on the rows.
        v1 = by_year["1"].scored_with_weights_version
        v2 = by_equip["1"].scored_with_weights_version
        assert v1 and v2 and v1 != v2

    def test_partial_cluster_warns_but_does_not_fail(self):
        """A content-hash cluster_id whose submitted members don't reproduce the
        hash (a subset submitted separately) warns 'partial_cluster'; a complete
        submission does not."""
        from dedup.cluster_key import cluster_hash
        cid = cluster_hash(["1", "2"])  # id encodes the FULL membership {1, 2}

        # Only member "1" submitted → subset → partial_cluster warning.
        subset = elect_golden_records([_cluster_row("1", cluster_id=cid)], WEIGHTS)
        assert any("partial_cluster" in w for w in subset[0].warnings)

        # Both members submitted → hash matches → no partial warning.
        full = elect_golden_records([
            _cluster_row("1", cluster_id=cid, last_order_year=Y0),
            _cluster_row("2", cluster_id=cid),
        ], WEIGHTS)
        assert not any("partial_cluster" in w for r in full for w in r.warnings)

    def test_non_hash_cluster_id_never_warns_partial(self):
        """A plain id like 'C1' is not a content hash and must never trip the
        partial-cluster check."""
        results = elect_golden_records([
            _cluster_row("1", cluster_id="C1", last_order_year=Y0),
            _cluster_row("2", cluster_id="C1"),
        ], WEIGHTS)
        assert not any("partial_cluster" in w for r in results for w in r.warnings)

    def test_rerun_is_deterministic_end_to_end(self):
        """Scoring the same workbook twice yields identical scoring/election
        column values (determinism end to end)."""
        wb = _build_workbook([
            _data_row("1", cluster="A1", routing="cluster", year=Y0, orders=12),
            _data_row("2", cluster="A1", routing="cluster", year=Y3),
            _data_row("3"),
        ])
        out1, _ = score_workbook(wb)
        out2, _ = score_workbook(wb)
        ws1 = load_workbook(io.BytesIO(out1))["Sheet1"]
        ws2 = load_workbook(io.BytesIO(out2))["Sheet1"]
        vals1 = [[c.value for c in row] for row in ws1.iter_rows()]
        vals2 = [[c.value for c in row] for row in ws2.iter_rows()]
        assert vals1 == vals2

    def test_scored_with_weights_version_written_to_file(self):
        output, _ = score_workbook(_build_workbook([_data_row("1", year=Y0)]))
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]
        assert "scored_with_weights_version" in headers
        col = headers.index("scored_with_weights_version") + 1
        assert ws.cell(row=2, column=col).value  # non-empty fingerprint


class TestG1CountRecency:
    """G1 — Bernd's year-priority rule: sales-order count points are awarded
    only to the cluster's most-recent-year record(s); an older, higher-volume
    record can never out-score a more recent one on count."""

    def _comp(self, res):
        """Combined sales-order component = recency points + awarded count."""
        b = res.score_breakdown
        return b["sales_order_last_used"] + b["sales_order_count"]

    def test_bernd_example_2026_three_orders_beats_older_record_with_25(self):
        # Bernd's verbatim example: 2019+25 vs Y0+3 → the Y0 record is golden.
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", last_order_year=2019, order_count=25),
            _cluster_row("B", cluster_id="C1", last_order_year=Y0, order_count=3),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["B"].is_golden_record is True
        assert by["A"].is_golden_record is False
        assert by["A"].score_breakdown["sales_order_count"] == 0   # suppressed
        assert by["B"].score_breakdown["sales_order_count"] == 5   # 3 -> band 0-5

    def test_discovered_failure_2023_12_beaten_by_2026_3(self):
        # The exact regression: flat-additive gave Y3+12=30 > Y0+3=25.
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", last_order_year=Y3, order_count=12),
            _cluster_row("B", cluster_id="C1", last_order_year=Y0, order_count=3),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["B"].is_golden_record is True
        assert by["A"].score_breakdown["sales_order_count"] == 0

    def test_same_year_count_differentiates(self):
        # Count only "adds something" WITHIN the same year.
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", last_order_year=Y0, order_count=3),
            _cluster_row("B", cluster_id="C1", last_order_year=Y0, order_count=12),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["A"].score_breakdown["sales_order_count"] == 5
        assert by["B"].score_breakdown["sales_order_count"] == 25
        assert by["B"].is_golden_record is True

    # -- Partner mirror -----------------------------------------------------

    def test_partner_bernd_example(self):
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", partner_last_order_year=2019,
                         partner_order_count=25),
            _cluster_row("B", cluster_id="C1", partner_last_order_year=Y0,
                         partner_order_count=3),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["A"].score_breakdown["sales_order_partner_count"] == 0
        assert by["B"].score_breakdown["sales_order_partner_count"] == 5
        assert by["B"].is_golden_record is True

    def test_partner_discovered_failure(self):
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", partner_last_order_year=Y3,
                         partner_order_count=12),
            _cluster_row("B", cluster_id="C1", partner_last_order_year=Y0,
                         partner_order_count=3),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["B"].is_golden_record is True
        assert by["A"].score_breakdown["sales_order_partner_count"] == 0

    def test_partner_same_year_differentiates(self):
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", partner_last_order_year=Y0,
                         partner_order_count=3),
            _cluster_row("B", cluster_id="C1", partner_last_order_year=Y0,
                         partner_order_count=12),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["A"].score_breakdown["sales_order_partner_count"] == 5
        assert by["B"].score_breakdown["sales_order_partner_count"] == 25

    # -- Edge cases ---------------------------------------------------------

    def test_singleton_receives_count_points(self):
        (r,) = elect_golden_records([
            ScoringRow(row_id="1", cluster_id=None, last_order_year=Y3,
                       order_count=12),
        ], WEIGHTS)
        assert r.election_status == "unique"
        assert r.score_breakdown["sales_order_count"] == 25  # 12 -> >10, awarded

    def test_all_none_year_cluster_no_count_no_exception(self):
        rows = [
            _cluster_row("A", cluster_id="C1", order_count=12),
            _cluster_row("B", cluster_id="C1", order_count=25),
        ]
        results = elect_golden_records(rows, WEIGHTS)
        by = _by_row(results)
        assert by["A"].score_breakdown["sales_order_count"] == 0
        assert by["B"].score_breakdown["sales_order_count"] == 0
        # No recency competitor (cluster max year is None) → not flagged as noise.
        assert not any(
            i.issue_type == "count_suppressed_by_recency"
            for i in detect_issues(rows, results)
        )

    def test_context_free_year_none_suppression_is_not_flagged(self):
        """A lone (unique) year-None row with a present count is still zeroed by
        rule 1, but produces NO suppression warning/issue — there is no cluster
        and no more-recent competitor, so it is not 'volume lost to recency'."""
        rows = [ScoringRow(row_id="1", cluster_id=None, order_count=12)]  # no year
        results = elect_golden_records(rows, WEIGHTS)
        (r,) = results
        assert r.score_breakdown["sales_order_count"] == 0   # rule 1 still applies
        assert not any("count suppressed" in w for w in r.warnings)
        assert not any(
            i.issue_type == "count_suppressed_by_recency"
            for i in detect_issues(rows, results)
        )

    def test_max_year_record_with_none_count_still_wins_on_recency(self):
        results = elect_golden_records([
            _cluster_row("A", cluster_id="C1", last_order_year=Y0),          # max, no count
            _cluster_row("B", cluster_id="C1", last_order_year=Y3, order_count=25),
        ], WEIGHTS)
        by = _by_row(results)
        assert by["A"].is_golden_record is True
        assert by["A"].score_breakdown["sales_order_count"] == 0   # no count field
        assert by["B"].score_breakdown["sales_order_count"] == 0   # suppressed
        assert any("count suppressed (G1)" in w for w in by["B"].warnings)

    def test_suppression_emits_issue(self):
        rows = [
            _cluster_row("A", cluster_id="C1", last_order_year=Y3, order_count=12),
            _cluster_row("B", cluster_id="C1", last_order_year=Y0, order_count=3),
        ]
        results = elect_golden_records(rows, WEIGHTS)
        issues = detect_issues(rows, results)
        supp = [i for i in issues if i.issue_type == "count_suppressed_by_recency"]
        assert supp and supp[0].row_id == "A"

    # -- Provable invariant -------------------------------------------------

    def test_sales_component_never_contradicts_recency(self):
        """Within any cluster, ordering by the sales-order component (recency +
        awarded count) never contradicts ordering by last_order_year: a strictly
        more recent year yields a strictly greater component (years drawn from
        the strictly-decreasing tier range Y3-Y0)."""
        rng = random.Random(20260723)
        YEARS = [Y3, Y2, Y1, Y0]
        for _ in range(300):
            n = rng.randint(2, 5)
            spec = [(str(i), rng.choice(YEARS), rng.randint(0, 30)) for i in range(n)]
            rows = [_cluster_row(rid, cluster_id="C1", last_order_year=y,
                                 order_count=c) for rid, y, c in spec]
            by = _by_row(elect_golden_records(rows, WEIGHTS))
            comp = {rid: self._comp(by[rid]) for rid, _, _ in spec}
            year = {rid: y for rid, y, _ in spec}
            for a in year:
                for b in year:
                    if year[a] > year[b]:
                        assert comp[a] > comp[b], (spec, comp)

    def test_partner_component_never_contradicts_recency(self):
        rng = random.Random(7770723)
        YEARS = [Y3, Y2, Y1, Y0]
        for _ in range(300):
            n = rng.randint(2, 5)
            spec = [(str(i), rng.choice(YEARS), rng.randint(0, 30)) for i in range(n)]
            rows = [_cluster_row(rid, cluster_id="C1", partner_last_order_year=y,
                                 partner_order_count=c) for rid, y, c in spec]
            by = _by_row(elect_golden_records(rows, WEIGHTS))
            comp = {rid: by[rid].score_breakdown["sales_order_partner_last_used"]
                         + by[rid].score_breakdown["sales_order_partner_count"]
                    for rid, _, _ in spec}
            year = {rid: y for rid, y, _ in spec}
            for a in year:
                for b in year:
                    if year[a] > year[b]:
                        assert comp[a] > comp[b], (spec, comp)


# ---------------------------------------------------------------------------
# Dates, the relative ladder, and the header bindings the click report needs
# ---------------------------------------------------------------------------

class TestDateCoercion:
    """The two *_last_used columns arrive from the click report as real dates.

    `US_Qlic report data_2026-07-30.xlsx` stores both as datetimes, never as
    integer years. Before the fix the Scalar union rejected the cell outright
    (15,369 of 22,224 rows raised ValidationError) and the ISO-string form
    coerced to None — zeroing both ladders directly and both COUNT rules
    through the G1 gate, i.e. 90 of the 200 available points on every record.
    """

    @pytest.mark.parametrize("value", [
        datetime.datetime(2026, 7, 28, 0, 0),
        datetime.date(2026, 7, 28),
        "2026-07-28",
        "2026-07-28 00:00:00",
        "2026/07/28",
        2026,
        "2026",
    ])
    def test_year_is_extracted_from_a_date(self, value):
        assert _points(
            {"last_order_year": value}, "sales_order_last_used", current_year=2026
        ) == 20
        assert _points(
            {"partner_last_order_year": value},
            "sales_order_partner_last_used", current_year=2026,
        ) == 20

    def test_a_date_does_not_raise_at_the_model_boundary(self):
        # Scalar must ACCEPT a date: the file route builds ScoringRow straight
        # from openpyxl cell values, so a rejection here 500s the whole batch.
        row = ScoringRow(row_id="1", last_order_year=datetime.datetime(2026, 7, 28))
        assert isinstance(row.last_order_year, datetime.datetime)

    def test_a_date_unblocks_the_g1_count_gate(self):
        # The count rules are gated on the row having a year at all, so a
        # date that coerced to None took the counts down with it.
        breakdown, warnings = _score(
            {"last_order_year": datetime.datetime(2026, 3, 1), "order_count": 7,
             "partner_last_order_year": datetime.datetime(2026, 3, 1),
             "partner_order_count": 7},
            current_year=2026,
        )
        assert breakdown["sales_order_count"] == 15
        assert breakdown["sales_order_partner_count"] == 15
        assert warnings == []

    @pytest.mark.parametrize("value", [
        datetime.datetime(2026, 7, 28), datetime.date(2026, 7, 28), "2026-07-28",
    ])
    def test_a_date_in_a_count_column_is_still_dirt(self, value):
        """A date is only meaningful in the two *_last_used columns.

        The count columns are plain integers, so a date landing in one is a
        broken upstream join, not a year — it must keep scoring 0 WITH a
        warning rather than quietly scoring as ~2026 (which would hit the
        top equipment band and silently award 30 points).
        """
        breakdown, warnings = _score({"equipment_count": value})
        assert breakdown["equipment_count"] == 0
        assert any("equipment_count" in w for w in warnings)

    @pytest.mark.parametrize("text", ["lots", "n/a", "1234-56-78", "0007-01-01"])
    def test_non_dates_keep_todays_behaviour(self, text):
        breakdown, warnings = _score({"last_order_year": text})
        assert breakdown["sales_order_last_used"] == 0
        assert any("last_order_year" in w for w in warnings)


class TestRelativeLadder:
    """Both *_last_used ladders are banded on the OFFSET from a reference year.

    Bernd described the rule relatively — "sales order last year, last two
    years, last three years" (BerndScoring1 15:00) — and only then instantiated
    it as 2026/2025/2024/2023. Hard-coded years would make the whole model read
    zero from 1 January 2027.
    """

    @pytest.mark.parametrize("offset,expected", [
        (0, 20), (1, 15), (2, 10), (3, 5), (4, 0), (10, 0),
    ])
    @pytest.mark.parametrize("reference", [2026, 2027, 2031, 2100])
    def test_ladder_follows_the_reference_year(self, offset, expected, reference):
        assert _points(
            {"last_order_year": reference - offset},
            "sales_order_last_used", current_year=reference,
        ) == expected
        assert _points(
            {"partner_last_order_year": reference - offset},
            "sales_order_partner_last_used", current_year=reference,
        ) == expected

    def test_the_same_year_ages_out_of_the_ladder(self):
        """The regression the relative ladder exists to prevent: a fixed 2026
        scored 20 forever under absolute bands, and 0 for everything from
        1 January 2027 under a table that was never retuned."""
        assert _points({"last_order_year": 2026}, "sales_order_last_used",
                       current_year=2026) == 20
        assert _points({"last_order_year": 2026}, "sales_order_last_used",
                       current_year=2029) == 5
        assert _points({"last_order_year": 2026}, "sales_order_last_used",
                       current_year=2030) == 0

    @pytest.mark.parametrize("ahead", [1, 2, 5])
    def test_a_future_dated_order_scores_zero(self, ahead):
        # A negative offset matches no band. It must not wrap onto a tier.
        assert _points({"last_order_year": 2026 + ahead}, "sales_order_last_used",
                       current_year=2026) == 0

    def test_reference_year_is_resolved_once_per_election(self):
        # Every row of one election carries the same anchor — a per-row clock
        # read would let a batch straddling midnight on 31 December band its
        # first and last rows differently.
        results = elect_golden_records(
            [_cluster_row(str(i), last_order_year=Y0) for i in range(5)], WEIGHTS
        )
        assert {r.scored_with_reference_year for r in results} == {THIS_YEAR}

    def test_reference_year_is_stamped_on_every_row(self):
        results = elect_golden_records([
            _cluster_row("1", cluster_id="C1", last_order_year=Y0),
            _cluster_row("2", cluster_id="C1", last_order_year=Y3),
            _cluster_row("3"),  # unique
        ], WEIGHTS)
        for r in results:
            assert r.scored_with_reference_year == THIS_YEAR

    def test_reference_year_written_to_file(self):
        output, _ = score_workbook(_build_workbook([_data_row("1", year=Y0)]))
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]
        assert "scored_with_reference_year" in headers
        col = headers.index("scored_with_reference_year") + 1
        assert ws.cell(row=2, column=col).value == THIS_YEAR

    def test_tiebreak_still_orders_on_the_absolute_year(self):
        """The offset must exist ONLY inside the two ladder lookups.

        The tie-break orders by most-recent year; if it saw an offset the
        comparison would invert and the OLDEST record would win a tie.
        """
        rows = [
            _cluster_row("1", cluster_id="C1", last_order_year=Y4),
            _cluster_row("2", cluster_id="C1", last_order_year=Y5),
        ]
        by = _by_row(elect_golden_records(rows, WEIGHTS))
        # Both score 0 on the ladder (off the end), so the tie-break decides:
        # the more recent year must win.
        assert by["1"].is_golden_record is True
        assert by["2"].is_golden_record is False


class TestClickReportHeaders:
    """The click report spells two scoring headers differently, and _norm's
    tolerance (lowercase + strip non-alphanumerics) is not wide enough:
    "Customer Account Group" -> "customeraccountgroup" != "accountgroup", and
    "Customer No." -> "customerno" != "customer"."""

    @pytest.mark.parametrize("key", ["Customer", "Customer No.", "row_id"])
    def test_json_route_binds_every_customer_spelling(self, key):
        assert ScoringRow.model_validate({key: "0013012902"}).row_id == "0013012902"

    @pytest.mark.parametrize(
        "key", ["Account group", "Customer Account Group", "account_group"]
    )
    def test_json_route_binds_every_account_group_spelling(self, key):
        row = ScoringRow.model_validate({"row_id": "1", key: "DRIT"})
        assert row.account_group == "DRIT"

    def test_file_route_scores_the_click_report_spellings(self):
        """All nine scoring columns bind off the click report's own headers,
        and account_group alone is worth 20 points a record."""
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Customer No.", "Customer Account Group", "Sales Order Last Used",
            "Sales Order Total Count", "Sales Order Partner Last Used",
            "Sales Order Partner Total Count", "Equipment Total Count",
            "SleepingCustomer", "CustomerStatus",
        ])
        ws.append([
            "0013012902", "DRIT", datetime.datetime(THIS_YEAR, 7, 28), 7,
            datetime.datetime(THIS_YEAR, 7, 28), 7, 5, "No", "Active",
        ])
        buffer = io.BytesIO()
        wb.save(buffer)

        output, summary = score_workbook(buffer.getvalue())
        ws_out = load_workbook(io.BytesIO(output)).worksheets[0]
        headers = [c.value for c in ws_out[1]]

        def cell(h):
            return ws_out.cell(row=2, column=headers.index(h) + 1).value

        # The click report's own headers survive; nothing is renamed.
        assert "Customer No." in headers and "Customer Account Group" in headers
        assert cell("Customer No.") == "0013012902"
        assert cell("score_AccountGroup") == 20        # was 0 — bound on name
        assert cell("score_SalesOrderLastUsed") == 20  # was 0 — date not a year
        assert cell("score_SalesOrderCount") == 15
        assert cell("score_SalesOrderPartnerLastUsed") == 20
        assert cell("score_SalesOrderPartnerCount") == 15
        assert cell("score_EquipmentCount") == 12
        assert cell("score_SleepingCustomer") == 15
        assert cell("score_CustomerStatus") == 10
        # 20 group + 20 ladder + 15 count + 20 partner + 15 partner count
        # + 12 equipment + 15 sleeping + 10 status
        assert cell("score_final") == 127
        assert summary.errors == 0


class TestStaleWeightsTable:
    """§6: six band labels changed, so every previously-valid weights table is
    now stale. coerce_weights rejects a stale table WHOLESALE — a half-applied
    retune is worse than none — and says which pair it missed."""

    @pytest.mark.parametrize("criterion,stale_band,new_band", [
        ("sales_order_last_used", "2026", "0"),
        ("sales_order_partner_last_used", "2026", "0"),
        ("sales_order_count", "0-5", "1-5"),
        ("sales_order_partner_count", "0-5", "1-5"),
        ("equipment_count", "0-3", "1-3"),
        ("sleeping_customer", ">5", "Yes"),
    ])
    def test_each_changed_label_rejects_a_stale_table(
        self, criterion, stale_band, new_band
    ):
        stale = {c: dict(b) for c, b in WEIGHTS.items()}
        stale[criterion] = {
            (stale_band if k == new_band else k): v
            for k, v in stale[criterion].items()
        }
        weights, reason = coerce_weights(stale, WEIGHTS, source="Weights sheet")
        assert weights is None
        assert reason == (
            f"Weights sheet ignored: missing (criterion, band) pair "
            f"({criterion!r}, {new_band!r}); using dedup/weights.json"
        )

    def test_a_stale_weights_sheet_is_ignored_wholesale_not_merged(self):
        # The whole point: the retuned points on the OTHER criteria must not
        # leak in. A stale sheet falls all the way back to weights.json.
        stale_rows = [("Criterion", "Band", "Points", "Note")]
        for criterion, bands in WEIGHTS.items():
            for band, points in bands.items():
                if (criterion, band) == ("equipment_count", "1-3"):
                    band = "0-3"                      # the stale label
                if criterion == "account_group":
                    points = 99                       # a retune that must NOT apply
                stale_rows.append((criterion, band, points, None))

        output, summary = score_workbook(_build_workbook(
            [_data_row("1", year=Y0, group="DRIT")], weights_rows=stale_rows
        ))
        assert any(
            w == "Weights sheet ignored: missing (criterion, band) pair "
                 "('equipment_count', '1-3'); using dedup/weights.json"
            for w in summary.warnings
        )
        ws = load_workbook(io.BytesIO(output))["Sheet1"]
        headers = [c.value for c in ws[1]]
        # weights.json's DRIT = 20, NOT the sheet's 99.
        assert ws.cell(
            row=2, column=headers.index("score_AccountGroup") + 1
        ).value == 20

    def test_the_current_table_is_accepted(self):
        weights, reason = coerce_weights(WEIGHTS, WEIGHTS)
        assert reason is None
        assert weights == WEIGHTS
