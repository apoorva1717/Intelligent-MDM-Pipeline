"""Name-2 slot classification — ``DEDUP_V2_NAME2`` (change C).

What the text below Name 1 IS, and what follows from getting it right: the
signature key, the ``has_name2`` boundary that the deterministic asymmetry rule
turns on, the six columns the file route used to drop, the prompt, and the two
extra candidate rules for large blocks.

Every classification case here is a real row from the stress batch, cited by
Customer number, because the rules were written against these shapes and a
synthetic example would not tell you whether they hold on the data.
"""

from __future__ import annotations

import asyncio
import os
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ.setdefault("ENV", "local")
os.environ.setdefault("MOCK_EXTERNAL_CALLS", "true")
os.environ.setdefault("OPENAI_API_KEY", "test-key-for-mocks")

from dedup.candidates import CandidateUnit, generate_candidate_pairs  # noqa: E402
from dedup.models import DedupRow  # noqa: E402
from dedup.name_slots import classify_slots  # noqa: E402
from dedup.prompts import PROMPT_VERSION, PROMPT_VERSION_V2  # noqa: E402
from dedup.signatures import build_signatures  # noqa: E402
from tests.dedup_v2_support import (  # noqa: E402
    V2_FLAGS,
    SpecOracleLLM,
    fixture_dedup_rows,
    load_fixture,
    run_clustering,
)


@pytest.fixture
def flags_on(monkeypatch: pytest.MonkeyPatch) -> None:
    for flag in V2_FLAGS:
        monkeypatch.setenv(flag, "true")


@pytest.fixture
def name2_only(monkeypatch: pytest.MonkeyPatch) -> None:
    """Only C. The flags must work one at a time as well as together."""
    monkeypatch.setenv("DEDUP_V2_BLOCKING", "false")
    monkeypatch.setenv("DEDUP_V2_NAME2", "true")
    monkeypatch.setenv("DEDUP_V2_ID_CONFLICT", "false")


@pytest.fixture(scope="module")
def fixture() -> dict:
    return load_fixture()


# ---------------------------------------------------------------------------
# classify_slots
# ---------------------------------------------------------------------------

#: (label, Name 1, Name 2, other Name 1s in the block, kind, institution, department)
SLOT_CASES = [
    ("13044976 a delivery desk is not a department",
     "University of Texas", "Central Receiving", [],
     "logistics", "University of Texas", ""),
    ("13161437 accounts payable is not a department",
     "University of California, San Francisco", "Accounts Payable", [],
     "logistics", "University of California, San Francisco", ""),
    ("13216611 a trading name is not a department",
     "Lee Memorial Health System", "DBA Lee Health", [],
     "alias", "Lee Memorial Health System", ""),
    ("13226604 a parent company is not a department",
     "Global Equipment Services Inc", "A Kimball Electronics Company", [],
     "alias", "Global Equipment Services Inc", ""),
    ("13135468 Name 1's own tail is not a department",
     "EMD Serono Research & Development", "Institute, Inc", [],
     "overflow", "EMD Serono Research & Development Institute, Inc", ""),
    # With the full spelling present in the block this is a SPLIT name, not a
    # truncation: institution_split is asked first and matches. The outcome —
    # one institution, no department — is the same either way.
    ("13345937 a name split across two cells, with the full form in the block",
     "Palo Alto Veterans Institute for", "Research",
     ["Palo Alto Veterans Institute for Research"],
     "institution_split", "Palo Alto Veterans Institute for Research", ""),
    ("13011226 a truncated Name 1 continues into Name 2, nothing to match",
     "American School of Classical Studies at", "Athens", [],
     "overflow", "American School of Classical Studies at Athens", ""),
    ("13130623 an opaque Name 1 means Name 2 is the institution",
     "GHW23", "Case Western Reserve University",
     ["Case Western Reserve University"],
     "institution", "Case Western Reserve University", ""),
    ("13342545 a person is not a department",
     "UCSF", "Emanuela Zacco - LCA Core", [],
     "contact", "UCSF", ""),
    ("13367825 a real department stays a department",
     "Stanford University", "Fairchild Science", [],
     "department", "Stanford University", "Fairchild Science"),
    ("13036862 the whole block below Name 1 is the department",
     "National Aeronautics and Space Administration",
     "Intelligent Systems Division", [],
     "department", "National Aeronautics and Space Administration",
     "Intelligent Systems Division"),
]


@pytest.mark.parametrize(
    ("label", "name1", "name2", "others", "kind", "institution", "department"),
    SLOT_CASES, ids=[case[0] for case in SLOT_CASES],
)
def test_classify_slots(label, name1, name2, others, kind, institution, department) -> None:
    result = classify_slots(name1, name2, block_name1s=[name1, *others])
    assert (result.kind, result.institution, result.department) == (
        kind, institution, department,
    )


def test_a_two_word_department_is_not_read_as_a_person() -> None:
    """The trap this batch is built around, stated as its own assertion.

    "Fairchild Science" has the shape of a first and last name. Reading it as a
    contact empties the department, and the Stanford record then merges into
    the two bare "Stanford University" rows at the same door — the single
    highest-cost misclassification available here, because it destroys a real
    distinction rather than merely failing to find one.
    """
    result = classify_slots("Stanford University", "Fairchild Science")
    assert result.kind == "department"
    assert result.department == "Fairchild Science"


def test_a_unit_of_the_records_own_institution_is_not_a_person() -> None:
    """"Army Contracting Command - Detroit Arsenal" has the person-plus-site
    shape and is neither. It shares "Army" with its own Name 1, which is what
    a unit does and a contact does not."""
    result = classify_slots(
        "United States Army", "Army Contracting Command - Detroit Arsenal"
    )
    assert result.kind == "department"


def test_logistics_text_is_dropped_from_a_lower_slot_too() -> None:
    """A desk is not a unit wherever it was typed (13047774: Name 3)."""
    result = classify_slots(
        "The University of Texas Southwestern", "Medical Center", "Central Receiving"
    )
    assert result.department == "Medical Center"


def test_an_alias_is_kept_and_reported_not_discarded() -> None:
    """The department is emptied; the trading name is not thrown away.

    It is the other spelling of this institution, and the cross-slot candidate
    rule is built on exactly this field.
    """
    result = classify_slots("Lee Memorial Health System", "DBA Lee Health")
    assert result.aliases == ["DBA Lee Health"]


def test_an_opaque_name1_is_kept_as_an_alias() -> None:
    result = classify_slots(
        "KMB3 LLC", "Case Western Reserve University",
        block_name1s=["KMB3 LLC", "Case Western Reserve University"],
    )
    assert result.aliases == ["KMB3 LLC"]


def test_an_opaque_name1_needs_a_matching_name1_in_the_block() -> None:
    """Otherwise every short upper-case Name 1 promotes whatever is under it.

    "UCSF" is the same shape as "GHW23"; what separates them is that no other
    record in UCSF's block spells "Emanuela Zacco - LCA Core" as an institution.
    """
    result = classify_slots("UCSF", "Emanuela Zacco - LCA Core", block_name1s=["UCSF"])
    assert result.kind != "institution"
    assert result.institution == "UCSF"


# ---------------------------------------------------------------------------
# The signature key and has_name2
# ---------------------------------------------------------------------------

def _rows(*pairs) -> list[DedupRow]:
    return [
        DedupRow(row_id=str(i), name1=n1, name2=n2, country="US",
                 postal_code="12345", street="Main St", house_no="1")
        for i, (n1, n2) in enumerate(pairs, start=1)
    ]


def test_a_desk_no_longer_splits_a_record_from_its_own_institution(name2_only) -> None:
    """UTSA, the shortest statement of what C is for.

    Two records at 7703 Floyd Curl Dr: one bare "University of Texas", one with
    "Central Receiving" below it. v1 read the second as a departmental record,
    which the asymmetry rule then forbade from ever sharing an entity with the
    first — the rule working exactly as designed, on a false premise.
    """
    signatures = build_signatures(
        _rows(("University of Texas", None), ("University of Texas", "Central Receiving"))
    )
    assert len(signatures) == 1
    assert signatures[0].has_name2 is False


def test_a_real_department_still_splits(name2_only) -> None:
    signatures = build_signatures(
        _rows(("Stanford University", None), ("Stanford University", "Fairchild Science"))
    )
    assert len(signatures) == 2
    assert sorted(s.has_name2 for s in signatures) == [False, True]


def test_institution_split_empties_a_department_that_is_the_institution(name2_only) -> None:
    """The EMD block: one company written four ways, two of them across cells.

    "EMD Serono, Inc." + "Research and Development Institute" is not a
    department of EMD Serono. Read as one, it lands on the populated side of
    the asymmetry rule while the bare "EMD Serono Research and Development
    Institute, Inc." lands on the empty side, and the rule then declares two
    spellings of one company to be different entities — correctly applying
    itself to a false premise.
    """
    signatures = build_signatures(_rows(
        ("EMD Serono, Inc.", "Research and Development Institute"),
        ("EMD Serono Research Institute, Inc.", "Research and Development Institute"),
        ("EMD Serono Research and Development Institute, Inc.", None),
    ))
    # One signature, not three: with the full name selected from the block,
    # all three rows resolve to the SAME institution and collapse before any
    # model is consulted. That is the point of selecting the name rather than
    # composing it — a composed name is a third spelling that matches nothing.
    assert len(signatures) == 1
    assert signatures[0].has_name2 is False
    assert signatures[0].institution == (
        "EMD Serono Research and Development Institute, Inc."
    )
    assert len(signatures[0].row_ids) == 3


def test_a_split_names_fragments_are_hints_and_never_aliases(name2_only) -> None:
    """A piece of a name is not another name for the whole thing.

    "EMD Serono, Inc." is half of "EMD Serono Research and Development
    Institute, Inc." — and it is also, separately, the entire name of a
    different company at the same address. Filed as an alias it read as "this
    institute is also called EMD Serono, Inc.", the cross_slot rule matched it
    against that company, and the model merged the two: a company swallowed by
    its own research arm. Both fragments are kept, as hints, which are shown
    and never matched on.
    """
    result = classify_slots(
        "EMD Serono, Inc.", "Research and Development Institute",
        block_name1s=[
            "EMD Serono, Inc.",
            "EMD Serono Research and Development Institute, Inc.",
        ],
    )
    assert result.kind == "institution_split"
    assert result.aliases == []
    assert result.hints == [
        "EMD Serono, Inc.", "Research and Development Institute",
    ]


def test_a_split_fragment_is_not_cross_slot_evidence(name2_only) -> None:
    """The regression this rule exists to prevent, asserted at the rule."""
    from dedup.candidates import pair_evidence

    class Unit:
        def __init__(self, institution, aliases=(), ror=None):
            self.institution = institution
            self.aliases = aliases
            self.operating_name = None
            self.suggested_name = None
            self.ror_id = ror
            self.lei_id = None

    bare = Unit("EMD Serono, Inc.", ror="https://ror.org/027zrs220")
    institute = Unit(
        "EMD Serono Research and Development Institute, Inc.",
        ror="https://ror.org/027zrs220",
    )
    assert pair_evidence(bare, institute) == ("id",)


def test_institution_split_fires_when_the_slots_overlap(name2_only) -> None:
    """13138597's Name 1 already repeats "Research" and "Institute".

    So the concatenation stutters and Jaro-Winkler reads it at 0.88, under the
    0.92 bar. The containment arm is what catches it: neither slot introduces a
    token the full name does not already carry.
    """
    result = classify_slots(
        "EMD Serono Research Institute, Inc.", "Research and Development Institute",
        block_name1s=[
            "EMD Serono Research Institute, Inc.",
            "EMD Serono Research and Development Institute, Inc.",
        ],
    )
    assert result.kind == "institution_split"
    assert result.department == ""


def test_institution_split_needs_the_full_name_in_the_block(name2_only) -> None:
    """Without a record spelling the whole name, a department stays a department.

    The rule reads a name that IS present somewhere in the block; it never
    invents one out of two slots that merely sit next to each other.
    """
    result = classify_slots(
        "EMD Serono, Inc.", "Research and Development Institute",
        block_name1s=["EMD Serono, Inc.", "Pfizer Inc."],
    )
    assert result.kind == "department"
    assert result.department == "Research and Development Institute"


def test_overflow_rebuilds_one_institution_from_two_cells(name2_only) -> None:
    signatures = build_signatures(
        _rows(("Palo Alto Veterans Institute for", "Research"),
              ("Palo Alto Veterans Institute for Research", None))
    )
    assert len(signatures) == 1, "the two spellings are one signature"
    assert signatures[0].institution == "Palo Alto Veterans Institute for Research"


def test_the_signature_carries_the_bound_phase_one_columns(name2_only) -> None:
    """C.4's six columns reach a signature, not just a DedupRow."""
    rows = [DedupRow(
        row_id="1", name1="The Assay Depot Inc", country="US", postal_code="92075",
        street="Lomas Santa Fe Dr", house_no="505",
        operating_name="scientist com", suggested_name="The Assay Depot, Inc.",
        record_type="unknown", ror_id_provenance="ror:verified",
        building="Suite 110",
    )]
    sig = build_signatures(rows)[0]
    assert sig.operating_name == "scientist com"
    assert sig.suggested_name == "The Assay Depot, Inc."
    assert sig.record_type == "unknown"
    assert sig.ror_provenance == "ror:verified"


def test_building_reaches_neither_the_key_nor_the_block(flags_on) -> None:
    """A building is not an entity, and two buildings are not two doors.

    The change request binds Building as a hint only. This is that promise as a
    test: the same address with different Building values is one block and one
    signature.
    """
    from dedup.signatures import build_blocks

    rows = [
        DedupRow(row_id="1", name1="Acme Labs", country="US", postal_code="12345",
                 street="Main St", house_no="1", building="A"),
        DedupRow(row_id="2", name1="Acme Labs", country="US", postal_code="12345",
                 street="Main St", house_no="1", building="B"),
    ]
    assert len(build_blocks(rows)) == 1
    assert len(build_signatures(rows)) == 1


# ---------------------------------------------------------------------------
# The prompt
# ---------------------------------------------------------------------------

def test_the_prompt_version_moves_with_the_flag(flags_on, name2_only) -> None:
    from dedup.prompts import prompt_version

    assert prompt_version() == PROMPT_VERSION_V2


def test_the_v1_prompt_version_is_untouched(monkeypatch: pytest.MonkeyPatch) -> None:
    from dedup.prompts import prompt_version

    for flag in V2_FLAGS:
        monkeypatch.setenv(flag, "false")
    assert prompt_version() == PROMPT_VERSION


def test_the_model_is_shown_the_new_fields(flags_on, fixture) -> None:
    """Including the five columns the file route used to drop before this."""
    llm = SpecOracleLLM(fixture)
    asyncio.run(run_clustering(fixture_dedup_rows(fixture), llm))
    prompts = "\n".join(llm.prompts)
    for field in (
        "institution", "department", "aliases", "operating_name",
        "suggested_name", "record_type", "ror_id", "lei_id", "street_match",
    ):
        assert f'"{field}"' in prompts, f"{field} never reached the model"


def test_the_model_is_never_told_a_blocked_pair_is_incompatible(flags_on, fixture) -> None:
    """Two records only reach one prompt after their delivery point matched.

    "incompatible" would be describing a question that was already answered,
    in a word that invites the model to re-answer it the other way.
    """
    llm = SpecOracleLLM(fixture)
    asyncio.run(run_clustering(fixture_dedup_rows(fixture), llm))
    assert "incompatible" not in "\n".join(llm.prompts)


def test_the_department_rule_is_stated_in_both_directions(flags_on) -> None:
    """v3 said when records are the SAME and left the asymmetric case silent.

    One record with a department and one without, at one door, is the single
    most common shape in this batch; the prompt has to answer it outright.
    """
    from dedup.prompts import system_prompt

    text = system_prompt()
    assert "SAME ENTITY when" in text and "DIFFERENT ENTITY when" in text
    assert "one has a department and the other has none" in text


# ---------------------------------------------------------------------------
# C.5 — the extra candidate rules
# ---------------------------------------------------------------------------

def _unit(index: int, name: str, **kw) -> CandidateUnit:
    # ``adjudicated=False``: a lone signature the bucketed pass never put to the
    # model. That is what residue nomination is FOR, and a pair where both
    # sides were already adjudicated in the same bucket is skipped before any
    # rule is consulted.
    return CandidateUnit(
        index=index, name=name, ror_id=None, lei_id=None,
        has_name2=False, adjudicated=False, **kw,
    )


def _rules(units, extra_rules: bool) -> dict[tuple[int, int], str]:
    pairs = generate_candidate_pairs(
        units, name_threshold=0.85, token_threshold=0.6, extra_rules=extra_rules,
    )
    return {(c.a, c.b): c.rule for c in pairs}


def test_acronym_nominates_a_short_name_against_its_initials() -> None:
    units = [_unit(0, "Global Equipment Services Inc"), _unit(1, "GES Inc")]
    assert _rules(units, extra_rules=True) == {(0, 1): "acronym"}
    assert _rules(units, extra_rules=False) == {}


def test_acronym_drops_the_connector_words() -> None:
    """"University of Texas Health" initialises to UTH, not UOTH."""
    units = [_unit(0, "University of Texas Health"), _unit(1, "UTH")]
    assert _rules(units, extra_rules=True) == {(0, 1): "acronym"}


def test_acronym_does_not_fire_on_two_letters() -> None:
    """"HP" matches the initials of any two-word H-P name by coincidence.

    Hewlett Packard Enterprise is a different company at the same street
    address; an acronym line asserting otherwise would be handing the model
    evidence for the merge this batch exists to forbid.
    """
    units = [_unit(0, "Hewlett Packard Enterprise Company"), _unit(1, "HP Inc")]
    assert _rules(units, extra_rules=True) == {}


def test_acronym_does_not_fire_on_a_long_name() -> None:
    """Only the short side may BE the acronym, or initials match noise."""
    units = [
        _unit(0, "Global Equipment Services Inc"),
        _unit(1, "General Electric Systems Corporation"),
    ]
    assert _rules(units, extra_rules=True) == {}


def test_cross_slot_nominates_through_an_alias() -> None:
    units = [
        _unit(0, "Assay Depot Inc", aliases=("Scientist.com",)),
        _unit(1, "Scientist com"),
    ]
    assert _rules(units, extra_rules=True) == {(0, 1): "cross_slot"}


def test_cross_slot_nominates_through_a_suggested_name() -> None:
    units = [
        _unit(0, "Takeda Pharmaceuticals",
              suggested_name="TAKEDA PHARMACEUTICALS U.S.A., INC."),
        _unit(1, "Takeda Pharmaceutical U.S.A., Inc."),
    ]
    assert (0, 1) in _rules(units, extra_rules=True)


def test_id_convergence_still_outranks_the_new_rules() -> None:
    units = [
        CandidateUnit(index=0, name="Global Equipment Services Inc", ror_id="r1",
                      lei_id=None, has_name2=False, adjudicated=False),
        CandidateUnit(index=1, name="GES Inc", ror_id="r1", lei_id=None,
                      has_name2=False, adjudicated=False),
    ]
    assert _rules(units, extra_rules=True) == {(0, 1): "id"}


def test_the_extra_rules_stay_off_for_a_small_block(flags_on, fixture) -> None:
    """They are for blocks Mode A cannot cover in one call.

    Every block in this fixture is small, so nothing here should nominate on an
    acronym or an alias — the pairs those rules would find were all put to the
    model already, in one partition call, and asking again costs a call per
    pair for a second opinion on a settled question.
    """
    llm = SpecOracleLLM(fixture)
    results, _summary = asyncio.run(
        run_clustering(fixture_dedup_rows(fixture), llm)
    )
    reasons = " ".join(r.reasoning or "" for r in results.values())
    assert "[acronym]" not in reasons and "[cross_slot]" not in reasons


def test_the_reasoning_names_the_rule_that_nominated_the_pair(flags_on, fixture) -> None:
    """An id convergence and a guess at an acronym deserve different trust."""
    llm = SpecOracleLLM(fixture)
    results, _summary = asyncio.run(
        run_clustering(fixture_dedup_rows(fixture), llm)
    )
    reasons = [r.reasoning or "" for r in results.values()]
    assert any("[id]" in reason for reason in reasons)


# ---------------------------------------------------------------------------
# The third outcome — link for review
# ---------------------------------------------------------------------------

def test_an_uncertain_pair_with_evidence_shares_a_link_id(flags_on, fixture) -> None:
    """One organisation, two records, and no claim that they are one record.

    Before this the file had two outcomes: a Cluster ID, or nothing. A pair the
    model recognises as one organisation but declines to merge on legal-entity
    grounds had to be reported as one or the other, so the finding was either
    overstated as a duplicate or thrown away as "unique".
    """
    results, _summary = asyncio.run(
        run_clustering(fixture_dedup_rows(fixture), SpecOracleLLM(fixture))
    )
    left, right = results["13185655"], results["13350355"]
    assert left.link_id is not None and left.link_id == right.link_id
    assert left.cluster_id is None and right.cluster_id is None, (
        "a link is not a merge"
    )
    assert left.routing == right.routing == "manual_review"
    assert left.reasoning, "the model's reason must survive to the row"


def _sig(sid, name, relation=None, ror=None):
    from dedup.signatures import Signature

    return Signature(
        signature_id=sid, norm_name1=name.lower(), norm_name2="",
        name1=name, name2="", ror_id=ror, row_ids=[sid],
        institution=name, institution_relation=relation,
    )


def _entities(*signatures):
    from dedup.adjudicator import Entity

    return [
        Entity(entity_id=f"e{i}", signatures=[sig])
        for i, sig in enumerate(signatures, start=1)
    ]


def test_two_unrelated_institutions_are_not_linked() -> None:
    """A link needs a reason to exist, and sharing a block is not one."""
    from dedup.adjudicator import _institution_links

    links, conflicts = _institution_links(
        _entities(_sig("r1", "Acme Biotech", "different"),
                  _sig("r2", "Zenith Shipping", "different"))
    )
    assert links == {} and conflicts == []


def test_a_shared_registry_id_links_regardless_of_the_model() -> None:
    """The registry has standing of its own; a link is not the model's to veto."""
    from dedup.adjudicator import _institution_links

    links, conflicts = _institution_links(
        _entities(_sig("r1", "EMD Serono, Inc.", "different", ror="r"),
                  _sig("r2", "EMD Serono R&D Institute, Inc.", "different", ror="r"))
    )
    assert len(set(links.values())) == 1 and set(links) == {"r1", "r2"}
    assert [sig.signature_id for sig, _ in conflicts] == ["r1", "r2"], (
        "the registry and the model disagree — that is the steward's question"
    )


def test_evidence_plus_a_same_verdict_links_without_review() -> None:
    from dedup.adjudicator import _flag_institution_conflicts, _institution_links

    entities = _entities(
        _sig("r1", "Stanford University", "same"),
        _sig("r2", "Stanford University", "same"),
    )
    links, conflicts = _institution_links(entities)
    _flag_institution_conflicts(conflicts)
    assert len(set(links.values())) == 1
    assert not any(s.uncertain for e in entities for s in e.signatures), (
        "agreement is not a question"
    )


def test_an_uncertain_verdict_with_evidence_links_and_asks() -> None:
    from dedup.adjudicator import _institution_links

    links, _conflicts = _institution_links(
        _entities(_sig("r1", "United States Gypsum Company", "uncertain"),
                  _sig("r2", "USG Corporation, Inc.", "uncertain"))
    )
    assert len(set(links.values())) == 1


def test_the_link_id_column_is_absent_with_the_flags_off(
    monkeypatch: pytest.MonkeyPatch, fixture,
) -> None:
    """The v1 workbook keeps exactly v1's columns."""
    from api.routes import _DEDUP_RESULT_COLUMNS, _build_dedup_xlsx, _rows_to_dedup_rows
    from config import Settings
    from dedup.adjudicator import cluster_blocks
    from tests.dedup_v2_support import V1ReplayLLM, fixture_row_dicts

    import io
    import openpyxl

    for flag in V2_FLAGS:
        monkeypatch.setenv(flag, "false")
    row_dicts = fixture_row_dicts(fixture)
    headers = fixture["input_columns"]
    rows = _rows_to_dedup_rows(row_dicts)
    response = asyncio.run(
        cluster_blocks(rows, V1ReplayLLM(fixture), settings=Settings())
    )
    workbook = openpyxl.load_workbook(
        io.BytesIO(_build_dedup_xlsx(headers, row_dicts, rows, response))
    )
    written = [cell.value for cell in workbook.active[1]]
    assert written[-len(_DEDUP_RESULT_COLUMNS):] == _DEDUP_RESULT_COLUMNS
    assert "Link ID" not in written


def test_the_link_id_column_is_written_with_the_flags_on(flags_on, fixture) -> None:
    from api.routes import _build_dedup_xlsx, _rows_to_dedup_rows
    from config import Settings
    from dedup.adjudicator import cluster_blocks
    from tests.dedup_v2_support import fixture_row_dicts

    import io
    import openpyxl

    row_dicts = fixture_row_dicts(fixture)
    headers = fixture["input_columns"]
    rows = _rows_to_dedup_rows(row_dicts)
    response = asyncio.run(
        cluster_blocks(rows, SpecOracleLLM(fixture), settings=Settings())
    )
    workbook = openpyxl.load_workbook(
        io.BytesIO(_build_dedup_xlsx(headers, row_dicts, rows, response))
    )
    header_row = [cell.value for cell in workbook.active[1]]
    assert header_row.index("Link ID") == header_row.index("Cluster ID") + 1
    assert "Link ID" in [cell.value for cell in workbook["Dedup Debug"][1]]


# ---------------------------------------------------------------------------
# Making a run identifiable
# ---------------------------------------------------------------------------

def _run_sheet(fixture, llm_cls) -> dict[str, str]:
    import io

    import openpyxl

    from api.routes import _build_dedup_xlsx, _rows_to_dedup_rows
    from config import Settings
    from dedup.adjudicator import cluster_blocks
    from tests.dedup_v2_support import fixture_row_dicts

    row_dicts = fixture_row_dicts(fixture)
    rows = _rows_to_dedup_rows(row_dicts)
    response = asyncio.run(
        cluster_blocks(rows, llm_cls(fixture), settings=Settings())
    )
    workbook = openpyxl.load_workbook(io.BytesIO(
        _build_dedup_xlsx(fixture["input_columns"], row_dicts, rows, response)
    ))
    sheet = workbook["Run"]
    return {row[0].value: row[1].value for row in sheet.iter_rows(min_row=2)}


def test_the_run_sheet_says_which_configuration_produced_the_workbook(
    flags_on, fixture,
) -> None:
    """A workbook with no provenance is one nobody can place.

    Two runs can differ in every cluster and look identical on the front sheet,
    and the only tell that the flags were unset was a missing column — which is
    precisely how a v1 run gets mistaken for a v2 one.
    """
    run = _run_sheet(fixture, SpecOracleLLM)
    assert run["DEDUP_V2_BLOCKING"] == "true"
    assert run["DEDUP_V2_NAME2"] == "true"
    assert run["DEDUP_V2_ID_CONFLICT"] == "true"
    assert run["dedup_v2_active"] == "true"
    assert run["prompt_version"] == PROMPT_VERSION_V2
    assert run["fixture_cache"] == "off"
    assert run["rows_in"] == "200"


def test_the_run_sheet_reports_a_v1_run_as_a_v1_run(
    monkeypatch: pytest.MonkeyPatch, fixture,
) -> None:
    from tests.dedup_v2_support import V1ReplayLLM

    for flag in V2_FLAGS:
        monkeypatch.setenv(flag, "false")
    run = _run_sheet(fixture, V1ReplayLLM)
    assert run["dedup_v2_active"] == "false"
    assert run["prompt_version"] == PROMPT_VERSION


def test_the_run_sheet_does_not_disturb_the_data_sheets(
    monkeypatch: pytest.MonkeyPatch, fixture,
) -> None:
    """Additive means additive: v1's two sheets keep their exact shape."""
    import io

    import openpyxl

    from api.routes import (
        _DEDUP_DEBUG_COLUMNS, _DEDUP_RESULT_COLUMNS, _build_dedup_xlsx,
        _rows_to_dedup_rows,
    )
    from config import Settings
    from dedup.adjudicator import cluster_blocks
    from tests.dedup_v2_support import V1ReplayLLM, fixture_row_dicts

    for flag in V2_FLAGS:
        monkeypatch.setenv(flag, "false")
    row_dicts = fixture_row_dicts(fixture)
    rows = _rows_to_dedup_rows(row_dicts)
    response = asyncio.run(
        cluster_blocks(rows, V1ReplayLLM(fixture), settings=Settings())
    )
    workbook = openpyxl.load_workbook(io.BytesIO(
        _build_dedup_xlsx(fixture["input_columns"], row_dicts, rows, response)
    ))
    assert workbook.sheetnames == ["Sheet", "Dedup Debug", "Run"]
    main = [cell.value for cell in workbook["Sheet"][1]]
    assert main[-len(_DEDUP_RESULT_COLUMNS):] == _DEDUP_RESULT_COLUMNS
    assert "Link ID" not in main
    assert [c.value for c in workbook["Dedup Debug"][1]] == _DEDUP_DEBUG_COLUMNS
    assert workbook["Sheet"].max_row == 201
