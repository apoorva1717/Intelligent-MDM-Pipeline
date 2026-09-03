"""Tests for API routes using httpx AsyncClient."""

from __future__ import annotations

import io
import logging
import os
import sys
from pathlib import Path

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ["ENV"] = "local"
os.environ["MOCK_EXTERNAL_CALLS"] = "true"
# FIX(Bug 6): use direct OpenAI env vars, not Azure
os.environ.setdefault("OPENAI_API_KEY", "test-key-for-mocks")

from api.app import app

_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@pytest.fixture
def transport():
    return ASGITransport(app=app)


@pytest_asyncio.fixture
async def client(transport):
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c


class TestRoutes:
    """Test API endpoints."""

    @pytest.mark.asyncio
    async def test_health(self, client):
        resp = await client.get("/health")
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "healthy"
        assert data["version"] == "1.0.0"
        assert data["mock_mode"] is True

    @pytest.mark.asyncio
    async def test_tiers(self, client):
        resp = await client.get("/tiers")
        assert resp.status_code == 200
        data = resp.json()
        # FIX(Bug 1): single threshold now
        assert "ror_confidence_threshold" in data
        assert data["mock_mode"] is True

    @pytest.mark.asyncio
    async def test_enrich_single_record(self, client):
        payload = {
            "records": [
                {
                    "record_id": "ROUTE_001",
                    "name1": "Massachusetts Institute of Technology",
                    "name2": "Department of Chemistry",
                }
            ],
            "options": {"max_concurrency": 1},
        }
        resp = await client.post("/enrich", json=payload)
        assert resp.status_code == 200
        data = resp.json()
        assert len(data["results"]) == 1
        assert data["results"][0]["Customer"] == "ROUTE_001"
        assert data["summary"]["total"] == 1

    @pytest.mark.asyncio
    async def test_enrich_validation_error(self, client):
        """Empty records list should return 422."""
        payload = {"records": [], "options": {}}
        resp = await client.post("/enrich", json=payload)
        assert resp.status_code == 422

    @pytest.mark.asyncio
    async def test_enrich_record_without_identifier_accepted(self, client):
        """No field is mandatory: a record with no customer identifier is
        accepted (record_id falls back to empty string)."""
        payload = {
            "records": [{"name1": "MIT"}],
            "options": {},
        }
        resp = await client.post("/enrich", json=payload)
        assert resp.status_code == 200
        data = resp.json()
        assert len(data["results"]) == 1
        assert data["results"][0]["Customer"] == ""

    @pytest.mark.asyncio
    async def test_enrich_batch(self, client):
        """Batch of 3 records returns 3 results."""
        payload = {
            "records": [
                {"record_id": "R1", "name1": "MIT", "name2": "Department of Physics"},
                {"record_id": "R2", "name1": "Pfizer Inc", "name2": "R&D"},
                {"record_id": "R3", "name1": "UCLA"},
            ],
            "options": {"max_concurrency": 2},
        }
        resp = await client.post("/enrich", json=payload)
        assert resp.status_code == 200
        data = resp.json()
        assert len(data["results"]) == 3
        assert data["summary"]["total"] == 3

    @pytest.mark.asyncio
    async def test_enrich_always_200(self, client):
        """Valid request always returns 200, errors in result.error."""
        payload = {
            "records": [
                {"record_id": "ERR_001", "name1": None},
            ],
        }
        resp = await client.post("/enrich", json=payload)
        assert resp.status_code == 200

    @staticmethod
    def _xlsx_bytes(header: list, *rows: list) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _xlsx_upload(self, data: bytes, filename: str = "input.xlsx"):
        return {
            "file": (
                filename,
                data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }

    @staticmethod
    def _col(header: list, field: str) -> int:
        """Index of a result field's column, resolved via its output header."""
        from api.output_columns import RESPONSE_COLUMNS
        return header.index(RESPONSE_COLUMNS[field])

    @pytest.mark.asyncio
    async def test_enrich_file_returns_xlsx(self, client):
        """XLSX upload is enriched and returned as an XLSX file whose columns
        match the /enrich response body."""
        from api.output_columns import RESPONSE_COLUMNS

        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2"],
            ["ROUTE_001", "Massachusetts Institute of Technology", "Department of Chemistry"],
            [12345, "Pfizer Inc", "R&D"],  # numeric customer is coerced to str
            [None, None, None],  # blank row is skipped
        )
        resp = await client.post(
            "/enrich/file?max_concurrency=2", files=self._xlsx_upload(data)
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in resp.headers["content-disposition"]

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]

        # Columns are exactly the response-body fields, in order.
        assert header == list(RESPONSE_COLUMNS.values())
        # Two data rows, blank row skipped; record_id is the first column.
        assert len(rows) == 2
        assert rows[0][0] == "ROUTE_001"
        assert rows[1][0] == "12345"

    @pytest.mark.asyncio
    async def test_enrich_file_columns_match_response_body(self, client):
        """Output columns are driven by RESPONSE_COLUMNS: every key is a real
        result field and the header row is exactly the mapping's values."""
        from api.models import EnrichmentResult
        from api.output_columns import RESPONSE_COLUMNS

        data = self._xlsx_bytes(["Customer", "Name 1"], ["R1", "MIT"])
        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        header = [c.value for c in load_workbook(io.BytesIO(resp.content)).active[1]]
        # Every mapped key is a valid serialised result field (curatable
        # subset/order — not required to match the model one-to-one).
        result_fields = set(EnrichmentResult(record_id="x").model_dump().keys())
        assert set(RESPONSE_COLUMNS.keys()) <= result_fields
        assert header == list(RESPONSE_COLUMNS.values())

    @pytest.mark.asyncio
    async def test_enrich_file_tolerates_messy_headers(self, client):
        """Headers differing only by case/spacing still map to input fields,
        so the record is enriched."""
        data = self._xlsx_bytes(
            ["Customer ", "NAME 1", "Name  2 "],
            ["R1", "MIT", "Department of Chemistry"],
        )
        from api.output_columns import RESPONSE_COLUMNS

        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[1]]
        row = [c.value for c in ws[2]]
        # Recognised on input despite messy headers → name1 was enriched.
        name1_col = header.index(RESPONSE_COLUMNS["name1_enriched"])
        assert row[name1_col] == "Massachusetts Institute of Technology"

    @pytest.mark.asyncio
    async def test_enrich_file_dedupes_repeated_streets(self, client):
        """A street repeated across Street 1/2/3 is kept once; the repeats
        are blanked rather than echoed in every cleaned-street column."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Street 1", "Street 2", "Street 3"],
            ["R1", "Acme Corp", "235 East 42nd Street", "235 East 42nd Street", "235 East 42nd Street"],
        )
        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[1]]
        row = [c.value for c in ws[2]]
        assert row[self._col(header, "street_cleaned")]  # first occurrence kept
        assert row[self._col(header, "street_2_cleaned")] is None  # duplicate dropped
        assert row[self._col(header, "street_3_cleaned")] is None  # duplicate dropped

    @pytest.mark.asyncio
    async def test_enrich_file_dedupes_address_from_name_field(self, client):
        """An address in Name 2 that duplicates Street 1 (incl. trailing
        direction) is extracted in full and deduped, not echoed as a near-
        duplicate street."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Street 1", "Street 2"],
            [
                "13344992",
                "IDEXX Reference Labs",
                "10901 Roosevelt Blvd N",       # address living in Name 2
                "10901 ROOSEVELT BLVD N",        # same address in Street 1
                "Pinellas Bus Ctr, Ste 400D",
            ],
        )
        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[1]]
        row = [c.value for c in ws[2]]
        assert row[self._col(header, "street_cleaned")]            # Street 1 kept
        assert row[self._col(header, "street_3_cleaned")] is None  # Name 2 dup dropped

    @pytest.mark.asyncio
    async def test_enrich_file_street_junk_reaches_output(self, client):
        """End-to-end: URL / email / person / connector / orphan-marker junk
        in a street column is cleaned in the OUTPUT street_2_cleaned (not
        just in the helper functions)."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Street 1", "Street 2"],
            ["A", "Acme", "100 Main St", "ALSO 250 CENTRAL Ave"],
            ["B", "Acme", "100 Main St", "Ste"],
            ["C", "Acme", "100 Main St", "Dr Sarah Johnson - Lab Director"],
            ["D", "Acme", "100 Main St", "http://www.lockheedmartin.com/us/supplie"],
            ["E", "Acme", "100 Main St", "harrisapinvoices@harris.com"],
            ["F", "Acme", "100 Main St", "440 NICKERSON Rd"],
        )
        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        h = [c.value for c in ws[1]]
        rows = {
            r[self._col(h, "record_id")]: r
            for r in ws.iter_rows(min_row=2, values_only=True)
        }
        s2 = self._col(h, "street_2_cleaned")
        contact = self._col(h, "contact_enriched")
        email = self._col(h, "email_enriched")

        # Fix 5: the street value is cased on the way out, so the assertions
        # here are the cased forms. What this test pins is the junk removal —
        # "ALSO " stripped, orphan/person/URL/email dropped — not the casing.
        assert rows["A"][s2] == "250 Central Ave"   # connector stripped
        assert rows["B"][s2] is None                # orphan marker dropped
        assert rows["C"][s2] is None                # person removed...
        assert rows["C"][contact] == "Dr Sarah Johnson"  # ...routed to contact
        assert rows["D"][s2] is None                # URL removed
        assert rows["E"][s2] is None                # email removed...
        assert rows["E"][email] == "harrisapinvoices@harris.com"  # ...to email
        assert rows["F"][s2] == "440 Nickerson Rd"  # real address kept

    @pytest.mark.asyncio
    async def test_enrich_file_rejects_non_xlsx(self, client):
        resp = await client.post(
            "/enrich/file",
            files={"file": ("data.txt", b"not a spreadsheet", "text/plain")},
        )
        assert resp.status_code == 400

    @pytest.mark.asyncio
    async def test_enrich_file_row_without_identifier_accepted(self, client):
        """No field is mandatory: a row with no customer identifier is
        processed rather than rejected."""
        data = self._xlsx_bytes(["Name 1"], ["MIT"])
        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1

    @pytest.mark.asyncio
    async def test_enrich_file_no_data_rows(self, client):
        """Header only, no data rows → 400."""
        data = self._xlsx_bytes(["Customer", "Name 1"])
        resp = await client.post("/enrich/file", files=self._xlsx_upload(data))
        assert resp.status_code == 400

    # ── /issues endpoint ────────────────────────────────────────────────

    @pytest.mark.asyncio
    async def test_issues_echoes_input_and_appends_column(self, client):
        """The uploaded sheet is returned unchanged with an appended
        ``Issues`` column listing the detected codes per row."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Postal Code", "Country/Region Key"],
            ["R1", "Acme Corp", "PO BOX 115350", "12345", "US"],   # has issues
            ["R2", "Acme Corp", "Engineering", "12345", "US"],     # clean-ish
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in resp.headers["content-disposition"]

        ws = load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[1]]
        # Original headers preserved, "Issues" appended last.
        assert header == [
            "Customer", "Name 1", "Name 2", "Postal Code",
            "Country/Region Key", "Issues",
        ]

        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        assert len(rows) == 2
        # Original cell values are echoed verbatim.
        assert rows[0][:5] == ["R1", "Acme Corp", "PO BOX 115350", "12345", "US"]
        # PO Box in a name field surfaces as address-content-in-name.
        assert "G1-CROSS-001" in (rows[0][-1] or "")

    @pytest.mark.asyncio
    async def test_issues_multiple_codes_semicolon_joined(self, client):
        data = self._xlsx_bytes(
            ["Name 1", "Postal Code", "Country/Region Key"],
            ["Univ of Florida", "", "USA"],  # G5-NAME-001 + G2-VAL-002 + G4-ADDR-027
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        issues_cell = [c.value for c in ws[2]][-1]
        codes = {c.strip() for c in (issues_cell or "").split(";")}
        assert {"G2-VAL-002", "G4-ADDR-027", "G5-NAME-001"} <= codes

    @pytest.mark.asyncio
    async def test_issues_rejects_non_xlsx(self, client):
        resp = await client.post(
            "/issues",
            files={"file": ("data.txt", b"not a spreadsheet", "text/plain")},
        )
        assert resp.status_code == 400

    @pytest.mark.asyncio
    async def test_issues_column_aware_skips_absent_columns(self, client):
        """A file that doesn't carry Postal Code is not reported as missing it."""
        data = self._xlsx_bytes(["Customer", "Name 1", "Name 2"], ["R1", "Acme Corp", "Sales"])
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        issues_cell = [c.value for c in ws[2]][-1] or ""
        assert "G2-VAL-002" not in issues_cell  # Postal Code column absent

    # ── /issues/json endpoint ───────────────────────────────────────────

    @pytest.mark.asyncio
    async def test_issues_json_returns_record_id_and_issues(self, client):
        """One {record_id, issues} entry per record, in request order."""
        resp = await client.post(
            "/issues/json",
            json={
                "records": [
                    {
                        "Customer": "R1",
                        "Name 1": "Acme Corp",
                        "Name 2": "PO BOX 115350",
                        "Postal Code": "12345",
                        "Country/Region Key": "US",
                    },
                    {
                        "Customer": "R2",
                        "Name 1": "Acme Corp",
                        "Name 2": "Engineering",
                        "Postal Code": "12345",
                        "Country/Region Key": "US",
                    },
                ]
            },
        )
        assert resp.status_code == 200
        results = resp.json()["results"]
        assert [r["record_id"] for r in results] == ["R1", "R2"]
        # PO Box in a name field surfaces as address-content-in-name.
        assert "G1-CROSS-001" in results[0]["issues"]

    @pytest.mark.asyncio
    async def test_issues_json_matches_the_file_endpoint(self, client):
        """The JSON twin returns exactly what the file endpoint writes into
        its Issues column, for the same rows."""
        headers = ["Customer", "Name 1", "Postal Code", "Country/Region Key"]
        rows = [
            ["R1", "Univ of Florida", "", "USA"],
            ["R2", "Acme Corp", "12345", "US"],
        ]
        data = self._xlsx_bytes(headers, *rows)
        file_resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert file_resp.status_code == 200
        ws = load_workbook(io.BytesIO(file_resp.content)).active
        from_file = [
            {c.strip() for c in ([cell.value for cell in row][-1] or "").split(";") if c.strip()}
            for row in ws.iter_rows(min_row=2)
        ]

        json_resp = await client.post(
            "/issues/json",
            json={"records": [dict(zip(headers, row)) for row in rows]},
        )
        assert json_resp.status_code == 200
        from_json = [set(r["issues"]) for r in json_resp.json()["results"]]
        assert from_json == from_file
        assert {"G2-VAL-002", "G4-ADDR-027", "G5-NAME-001"} <= from_json[0]

    @pytest.mark.asyncio
    async def test_issues_json_column_aware_skips_absent_columns(self, client):
        """A payload that doesn't carry Postal Code is not reported as missing it."""
        resp = await client.post(
            "/issues/json",
            json={"records": [{"Customer": "R1", "Name 1": "Acme Corp", "Name 2": "Sales"}]},
        )
        assert resp.status_code == 200
        assert "G2-VAL-002" not in resp.json()["results"][0]["issues"]

    @pytest.mark.asyncio
    async def test_issues_json_blank_value_counts_as_present_but_empty(self, client):
        """Sending the column with an empty value is "present but blank" —
        the same distinction the workbook draws between a missing column and
        an empty cell."""
        resp = await client.post(
            "/issues/json",
            json={"records": [{"Customer": "R1", "Name 1": "Acme Corp", "Postal Code": ""}]},
        )
        assert resp.status_code == 200
        assert "G2-VAL-002" in resp.json()["results"][0]["issues"]

    @pytest.mark.asyncio
    async def test_issues_json_suppresses_the_same_codes_as_the_column(self, client):
        """The suppressed G6 codes and G7-VERIFY-001 are withheld here too."""
        resp = await client.post(
            "/issues/json",
            json={
                "records": [
                    {
                        "Customer": "R1",
                        "Name 1": "Acme Corp",
                        "Postal Code": "12345",
                        "Country/Region Key": "US",
                        "Flag for Review": "X",
                    }
                ]
            },
        )
        assert resp.status_code == 200
        issues = resp.json()["results"][0]["issues"]
        assert not ({"G2-VAL-003", "G2-VAL-006", "G2-NAME-012", "G7-VERIFY-001"} & set(issues))

    @pytest.mark.asyncio
    async def test_issues_json_record_without_identifier_gets_empty_id(self, client):
        """No field is mandatory: a record with no customer identifier is
        audited rather than rejected."""
        resp = await client.post(
            "/issues/json", json={"records": [{"Name 1": "MIT"}]}
        )
        assert resp.status_code == 200
        assert resp.json()["results"][0]["record_id"] == ""

    @pytest.mark.asyncio
    async def test_issues_json_integral_float_is_not_a_malformed_postal_code(self, client):
        """openpyxl hands back a whole-numbered cell as an int, so a postal
        code sent as the JSON number 12345.0 must reach the detector as
        "12345" — "12345.0" would be read as malformed (G4-ADDR-026), an issue
        the workbook never raises for the same value."""
        resp = await client.post(
            "/issues/json",
            json={
                "records": [
                    {
                        "Customer": "R1",
                        "Name 1": "Acme Corp",
                        "Postal Code": 12345.0,
                        "Country/Region Key": "US",
                    }
                ]
            },
        )
        assert resp.status_code == 200
        assert "G4-ADDR-026" not in resp.json()["results"][0]["issues"]

    @pytest.mark.asyncio
    async def test_issues_json_keeps_an_all_empty_record(self, client):
        """The one intentional divergence from the file endpoint: a blank
        spreadsheet row is a parsing artefact and is skipped, but an empty
        JSON object was deliberately sent and the response is positional, so
        it gets its own result rather than misaligning the caller's join."""
        resp = await client.post(
            "/issues/json",
            json={"records": [{"Customer": "R1", "Name 1": "Acme Corp"}, {}]},
        )
        assert resp.status_code == 200
        results = resp.json()["results"]
        assert len(results) == 2
        assert results[1]["record_id"] == ""

    @pytest.mark.asyncio
    async def test_issues_json_null_and_absent_and_blank_agree(self, client):
        """null, "" and an omitted key are the same "blank cell" — but only
        omitting it from *every* record removes the column."""
        payload = [
            {"Customer": "R1", "Name 1": "Acme Corp", "Postal Code": None},
            {"Customer": "R2", "Name 1": "Acme Corp", "Postal Code": ""},
            {"Customer": "R3", "Name 1": "Acme Corp"},
        ]
        resp = await client.post("/issues/json", json={"records": payload})
        assert resp.status_code == 200
        issues = [set(r["issues"]) for r in resp.json()["results"]]
        assert issues[0] == issues[1] == issues[2]
        assert "G2-VAL-002" in issues[0]  # the column is present across records

    @pytest.mark.asyncio
    async def test_issues_json_rejects_empty_records(self, client):
        resp = await client.post("/issues/json", json={"records": []})
        assert resp.status_code == 422

    # ── /issues/compare endpoint ────────────────────────────────────────

    @staticmethod
    def _summary(ws) -> dict:
        """Read the label/value pairs from the Summary sheet."""
        out = {}
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None and len(row) >= 2 and row[1] is not None:
                out[row[0]] = row[1]
        return out

    @pytest.mark.asyncio
    async def test_issues_compare_reports_reduction(self, client):
        # Original: PO box sits in a name field (G1-CROSS-001) + missing postal.
        original = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Postal Code", "Country/Region Key"],
            ["R1", "Acme Corp", "10901 Roosevelt Blvd N", "", "US"],
        )
        # Enriched: address moved out of the name, postal filled. Enriched
        # schema omits Postal Code entirely (column-aware → not counted missing).
        enriched = self._xlsx_bytes(
            ["record_id", "Name 1", "Name 2", "Country/Region Key"],
            ["R1", "Acme Corporation", "Sales Department", "US"],
        )
        resp = await client.post(
            "/issues/compare",
            files={
                "original": ("original.xlsx", original,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                "enriched": ("enriched.xlsx", enriched,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            },
        )
        assert resp.status_code == 200
        assert "attachment" in resp.headers["content-disposition"]

        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Summary", "Per Record", "Remaining Issues"]

        summary = self._summary(wb["Summary"])
        assert summary["Records matched (joined by id)"] == 1
        # Headline figures are the "Reduced" block (G1-G5) — see the segment
        # tests below for why the old undifferentiated totals are gone.
        assert summary["Reduced: issues before"] > summary["Reduced: issues after"]
        assert summary["Reduced: net reduction"] >= 1
        assert summary["Reduced: issues resolved"] >= 1

        per_record = [
            [c.value for c in row] for row in wb["Per Record"].iter_rows(min_row=2)
        ]
        assert per_record[0][0] == "R1"
        assert "G1-CROSS-001" in (per_record[0][3] or "")  # resolved column

        # Remaining Issues: one row per (code, customer) for issues still
        # present after enrichment, naming the customer id every time.
        remaining_ws = wb["Remaining Issues"]
        assert [c.value for c in next(remaining_ws.iter_rows())] == [
            "Code", "Name", "Group", "Segment", "Customer",
        ]
        remaining = [
            [c.value for c in row]
            for row in remaining_ws.iter_rows(min_row=2)
        ]
        # Every data row carries a code, its segment, and the customer id
        # (R1 is the only record).
        assert all(row[0] and row[4] == "R1" for row in remaining)
        assert all(
            row[3] in ("Reduced", "Expected to persist", "Verification")
            for row in remaining
        )
        # The "Reduced" rows are exactly the after-count of the reduction block.
        reduced_rows = [row for row in remaining if row[3] == "Reduced"]
        assert len(reduced_rows) == summary["Reduced: issues after"]

    @pytest.mark.asyncio
    async def test_issues_suppresses_unresolvable_g6_codes(self, client):
        """The suppressed G6 codes are withheld from the /issues column even in
        the conditions that raise them; the sibling required-field codes still
        fire, so the suppression is targeted rather than a column-gating side
        effect. Name 1 is a research institution with an empty Name 2, which is
        what G2-NAME-012 reads."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Tax Jurisdiction", "Language Key",
             "Postal Code", "Country/Region Key"],
            ["R1", "University of Florida", "", "", "", "", "US"],
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        codes = {c.strip() for c in ([c.value for c in ws[2]][-1] or "").split(";")}
        assert "G2-VAL-003" not in codes
        assert "G2-VAL-006" not in codes
        assert "G2-NAME-012" not in codes
        assert "G2-VAL-002" in codes  # Postal Code blank — still reported

    @pytest.mark.asyncio
    async def test_issues_g7_never_raised_on_a_raw_input_audit(self, client):
        """A raw file has no Flag for Review column, so G7 cannot apply. The
        record below is deliberately dirty: it must collect quality codes and
        still no G7."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Postal Code", "Country/Region Key"],
            ["R1", "Acme Corp", "10901 Roosevelt Blvd N", "", "US"],
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        cell = [c.value for c in ws[2]][-1] or ""
        assert "G1-CROSS-001" in cell
        assert "G7-VERIFY-001" not in cell

    @pytest.mark.asyncio
    async def test_issues_suppresses_g7_on_flagged_enriched_rows(self, client):
        """G7-VERIFY-001 is withheld from the /issues column even for a row the
        pipeline flagged — the case where the detector does raise it. The
        detector-level behaviour is pinned in tests/test_issue_detection.py."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Flag for Review", "Flag Reason"],
            ["R1", "Acme Corporation", "TRUE", "LLM canonical form — verify"],
            ["R2", "Beta Industries", "FALSE", ""],
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        flagged = [c.value for c in ws[2]][-1] or ""
        unflagged = [c.value for c in ws[3]][-1] or ""
        assert "G7-VERIFY-001" not in flagged
        assert "G7-VERIFY-001" not in unflagged

    @pytest.mark.asyncio
    async def test_issues_column_carries_the_flag_derived_codes(self, client):
        """The three codes the pipeline's own review flags map onto. Unlike
        G7-VERIFY-001 they are NOT suppressed: each names which of three
        different jobs a reviewer has, which is what an Issues column is for."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Flag for Review", "Flag Codes"],
            ["R1", "Acme Corporation", "TRUE", "opaque-code; email-conflict"],
            ["R2", "Beta Industries", "TRUE", "domain-unverified"],
            ["R3", "Gamma Labs", "TRUE", "person-unresolved"],
            ["R4", "Delta Works", "FALSE", ""],
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        cells = [([c.value for c in row][-1] or "") for row in ws.iter_rows(min_row=2)]

        assert "G6-RESOLVE-001" in cells[0]
        assert "G7-CONFIRM-001" in cells[1]
        assert "G8-VERIFY-001" in cells[2]
        # Two flags mapping to one code say it once, and a row the pipeline
        # flagged nothing on carries none of them.
        assert cells[0].count("G6-RESOLVE-001") == 1
        for code in ("G6-RESOLVE-001", "G7-CONFIRM-001", "G8-VERIFY-001"):
            assert code not in cells[3]

    @pytest.mark.asyncio
    async def test_issues_column_reads_the_retired_low_from_provenance(self, client):
        """`low-confidence-unchanged` cannot appear in Flag Codes — it was
        retired — so G8 is derived from `input:low` on the name provenance,
        which is where that state now lives. R2 pins the other direction: a
        settled name raises nothing."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Flag Codes", "Name 1 Provenance"],
            ["R1", "Acme Corporation", "", "input:low"],
            ["R2", "Beta Industries", "", "ror:verified"],
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        cells = [([c.value for c in row][-1] or "") for row in ws.iter_rows(min_row=2)]
        assert "G8-VERIFY-001" in cells[0]
        assert "G8-VERIFY-001" not in cells[1]

    @pytest.mark.asyncio
    async def test_raw_input_never_gets_a_flag_derived_code(self, client):
        """A file with neither a Flag Codes nor a provenance column is a raw
        audit, and the question was never asked of it."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2"],
            ["R1", "E004120188", "10901 Roosevelt Blvd N"],
        )
        resp = await client.post("/issues", files=self._xlsx_upload(data))
        assert resp.status_code == 200
        ws = load_workbook(io.BytesIO(resp.content)).active
        issues = ([c.value for c in ws[2]][-1] or "")
        for code in ("G6-RESOLVE-001", "G7-CONFIRM-001", "G8-VERIFY-001"):
            assert code not in issues

    @pytest.mark.asyncio
    async def test_issues_compare_segments_g6_and_g7_out_of_the_metric(self, client):
        """G6 codes that survive enrichment are expected persistence, not
        unreduced defects, and a G7 raised by the enrichment must not inflate
        the post-pipeline total. Neither may touch the reduction figures."""
        # Before: a G1 defect (address in a name) plus two G6 codes that no
        # automated path can fix (Tax Jurisdiction, Language).
        original = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Tax Jurisdiction", "Language Key",
             "Country/Region Key"],
            ["R1", "Acme Corp", "10901 Roosevelt Blvd N", "", "", "US"],
        )
        # After: the G1 defect is fixed, both G6 codes still there (correctly),
        # and the record is flagged for steward verification.
        enriched = self._xlsx_bytes(
            ["record_id", "Name 1", "Name 2", "Tax Jurisdiction", "Language Key",
             "Country/Region Key", "Flag for Review", "Flag Reason"],
            ["R1", "Acme Corporation", "Sales Department", "", "", "US",
             "TRUE", "domain-unverified"],
        )
        resp = await client.post(
            "/issues/compare",
            files={
                "original": ("original.xlsx", original, _XLSX_MIME),
                "enriched": ("enriched.xlsx", enriched, _XLSX_MIME),
            },
        )
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        summary = self._summary(wb["Summary"])

        # The reduction block sees only the G1 defect: 1 before, 0 after, 100%.
        assert summary["Reduced: issues before"] == 1
        assert summary["Reduced: issues after"] == 0
        assert summary["Reduction %"] == 100.0
        # G7 fired, and did not enter the reduction block.
        assert summary["Verification: records requiring verification"] == 1
        # Both G6 codes persisted, and are reported as such rather than as
        # unreduced defects.
        assert summary["Expected to persist: issues before"] == 2
        assert summary["Expected to persist: issues after"] == 2
        assert summary["Expected to persist: persisted as expected"] == 2

        rows = {
            row[0]: row
            for row in (
                [c.value for c in r] for r in wb["Summary"].iter_rows(min_row=2)
            )
            if row and isinstance(row[0], str) and row[0].startswith("G")
        }
        assert rows["G2-VAL-003"][2:4] == ["G6", "Expected to persist"]
        assert rows["G7-VERIFY-001"][2:4] == ["G7", "Verification"]
        assert rows["G1-CROSS-001"][2:4] == ["G1", "Reduced"]

    @pytest.mark.asyncio
    async def test_issues_compare_rejects_non_xlsx(self, client):
        good = self._xlsx_bytes(["Customer", "Name 1"], ["R1", "Acme"])
        resp = await client.post(
            "/issues/compare",
            files={
                "original": ("a.txt", b"nope", "text/plain"),
                "enriched": ("enriched.xlsx", good,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            },
        )
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
# Shared dedup fixture: one block, every DedupRow field, spelled the way
# /enrich/file writes it (api.output_columns.RESPONSE_COLUMNS). Used by both
# the file-route tests and the JSON/file equivalence test below.
# ---------------------------------------------------------------------------

# Header spelling -> DedupRow field:
#   Customer->row_id, Block ID->block_id, Name 1->name1, Name 2->name2,
#   Street 1->street, House Number->house_no, Postal Code->postal_code,
#   City->city, Country/Region Key->country, ROR ID->ror_id, LEI ID->lei_id,
#   Enriched Name->enriched_name
_FULL_DEDUP_HEADERS = [
    "Customer", "Block ID", "Name 1", "Name 2", "Street 1", "House Number",
    "Postal Code", "City", "Country/Region Key", "ROR ID", "LEI ID",
    "Enriched Name",
]

_FULL_DEDUP_ROWS = [
    ["SAP-1", "b1", "University of Stuttgart", "Department of Chemistry",
     "Pfaffenwaldring", "55", "70569", "Stuttgart", "DE",
     "https://ror.org/04vnq7t77", None, "University of Stuttgart"],
    ["SAP-2", "b1", "Universitaet Stuttgart", "Institut fuer Chemie",
     "Pfaffenwaldring", "55", "70569", "Stuttgart", "DE",
     "https://ror.org/04vnq7t77", None, "University of Stuttgart"],
    ["SAP-3", "b1", "Carl Zeiss AG", None,
     "Pfaffenwaldring", "55", "70569", "Stuttgart", "DE",
     None, "529900W18LQJJN6SJ336", "Carl Zeiss AG"],
]

# The same three rows as JSON-route payload dicts, keyed by DedupRow field.
_FULL_DEDUP_JSON = [
    dict(zip(
        ["row_id", "block_id", "name1", "name2", "street", "house_no",
         "postal_code", "city", "country", "ror_id", "lei_id", "enriched_name"],
        row,
    ))
    for row in _FULL_DEDUP_ROWS
]


def _capture_cluster_rows(monkeypatch) -> dict:
    """Intercept the DedupRow list each route hands to ``cluster_blocks``.

    Both endpoints call ``cluster_blocks`` from the ``api.routes`` namespace,
    so patching it there captures what the route actually built — the real
    wiring, not a re-run of the parsing helper.
    """
    import api.routes as routes

    captured: dict = {}
    real = routes.cluster_blocks

    async def _spy(rows, llm, **kwargs):
        captured["rows"] = list(rows)
        return await real(rows, llm, **kwargs)

    monkeypatch.setattr(routes, "cluster_blocks", _spy)
    return captured


class TestDedupFile:
    """The /api/dedup/file XLSX endpoint (file twin of cluster-block)."""

    _XTYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # Q4: exactly one cluster key on the main sheet (Cluster ID = stable hash);
    # internal keys (Block ID, Signature ID) live on the Dedup Debug sheet.
    _EXTRA = ["Cluster ID", "Routing", "LLM Flag", "Confidence", "Reasoning"]

    @staticmethod
    def _xlsx_bytes(header: list, *rows: list) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _upload(self, data: bytes, filename: str = "candidates.xlsx"):
        return {"file": (filename, data, self._XTYPE)}

    @pytest.mark.asyncio
    async def test_dedup_file_returns_annotated_xlsx(self, client):
        """Accepts enriched-export headers and appends cluster columns,
        one output row per input row, with row_ids echoed."""
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "Name 2", "Street 1", "House Number",
             "Postal Code", "City", "Country/Region Key"],
            ["SAP-1", "University of Stuttgart", "Department of Chemistry",
             "Pfaffenwaldring", "55", "70569", "Stuttgart", "DE"],
            ["SAP-2", "University of Stuttgart", "Department of Chemistry",
             "Pfaffenwaldring", "55", "70569", "Stuttgart", "DE"],
            ["SAP-3", "Carl Zeiss", None,
             "Carl-Zeiss-Straße", "22", "73447", "Oberkochen", "DE"],
        )
        resp = await client.post("/api/dedup/file", files=self._upload(data))
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(self._XTYPE)
        assert "candidates_dedup.xlsx" in resp.headers["content-disposition"]

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        for col in self._EXTRA:
            assert col in header
        # Internal keys moved off the main sheet onto the debug sheet.
        assert "Dedup Debug" in wb.sheetnames
        assert "Block ID" not in header and "Signature ID" not in header

        body = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(body) == 3  # one row per input row
        # Customer column echoed verbatim; routing is a valid bucket.
        cust_i = header.index("Customer")
        routing_i = header.index("Routing")
        cluster_i = header.index("Cluster ID")
        assert {r[cust_i] for r in body} == {"SAP-1", "SAP-2", "SAP-3"}
        assert all(r[routing_i] in {"cluster", "unique", "manual_review"}
                   for r in body)

        # The two identical Stuttgart rows share a stable-hash Cluster ID;
        # the unique Carl Zeiss row has none.
        by_cust = {r[cust_i]: r[cluster_i] for r in body}
        assert by_cust["SAP-1"] == by_cust["SAP-2"] is not None
        assert str(by_cust["SAP-1"]).startswith("c_")
        assert by_cust["SAP-3"] != by_cust["SAP-1"]

    @pytest.mark.asyncio
    async def test_dedup_file_accepts_snake_case_headers(self, client):
        data = self._xlsx_bytes(
            ["row_id", "name1", "name2", "postal_code", "country"],
            ["R1", "Acme", "Sales", "12345", "US"],
        )
        resp = await client.post("/api/dedup/file", files=self._upload(data))
        assert resp.status_code == 200

    @pytest.mark.asyncio
    async def test_dedup_file_carries_lei_and_ror_ids(self, client, monkeypatch):
        """The registry identifiers must reach the adjudicator, not just the
        output file.

        "LEI ID" bound to nothing before this was fixed: the column was echoed
        into the output workbook (so a column-presence check passed) while the
        value never reached DedupRow, and adjudication ran without the
        company legal-entity signal the JSON route carries.
        """
        captured = _capture_cluster_rows(monkeypatch)
        data = self._xlsx_bytes(_FULL_DEDUP_HEADERS, *_FULL_DEDUP_ROWS)

        resp = await client.post("/api/dedup/file", files=self._upload(data))
        assert resp.status_code == 200

        by_id = {r.row_id: r for r in captured["rows"]}
        assert by_id["SAP-1"].ror_id == "https://ror.org/04vnq7t77"
        assert by_id["SAP-1"].lei_id is None
        assert by_id["SAP-3"].lei_id == "529900W18LQJJN6SJ336"
        assert by_id["SAP-3"].ror_id is None

    @pytest.mark.asyncio
    async def test_dedup_file_warns_on_unrecognised_header(self, client, caplog):
        """A header the alias table does not know is named at WARNING level.

        One missing alias is a typo; silently discarding the column is what
        let it survive unnoticed.

        Note "LEI Code" is used rather than a whitespace variant like
        "LEI  I D": ``_norm_header`` strips all non-alphanumerics, so the
        latter normalises to "leiid" and is correctly recognised.
        """
        data = self._xlsx_bytes(
            ["Customer", "Name 1", "LEI Code", "Widget Count"],
            ["R1", "Acme GmbH", "529900W18LQJJN6SJ336", "7"],
        )
        with caplog.at_level(logging.WARNING, logger="api.routes"):
            resp = await client.post("/api/dedup/file", files=self._upload(data))
        assert resp.status_code == 200

        warnings = [
            r.getMessage() for r in caplog.records
            if r.levelno == logging.WARNING and "matched no DedupRow field" in r.getMessage()
        ]
        assert len(warnings) == 1, "expected exactly one line listing all headers"
        assert "'LEI Code'" in warnings[0]
        assert "'Widget Count'" in warnings[0]
        # Recognised columns must not be named.
        assert "'Customer'" not in warnings[0]
        assert "'Name 1'" not in warnings[0]

    @pytest.mark.asyncio
    async def test_dedup_file_rejects_non_xlsx(self, client):
        resp = await client.post(
            "/api/dedup/file",
            files={"file": ("data.txt", b"not a spreadsheet", "text/plain")},
        )
        assert resp.status_code == 400

    @pytest.mark.asyncio
    async def test_dedup_file_missing_row_id_is_validation_error(self, client):
        data = self._xlsx_bytes(["Name 1", "City"], ["Acme", "Boston"])
        resp = await client.post("/api/dedup/file", files=self._upload(data))
        assert resp.status_code == 422


class TestDedupRouteEquivalence:
    """The two intake routes must build identical DedupRow objects.

    This is the invariant the missing LEI alias violated. Asserting it
    field-for-field guards every alias at once, so the next omission fails
    here rather than silently weakening adjudication — which is how the LEI
    column went unnoticed.
    """

    _XTYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @pytest.mark.asyncio
    async def test_json_and_file_routes_build_identical_rows(
        self, client, monkeypatch,
    ):
        captured_json = _capture_cluster_rows(monkeypatch)
        resp = await client.post(
            "/api/dedup/cluster-block", json={"rows": _FULL_DEDUP_JSON},
        )
        assert resp.status_code == 200
        json_rows = captured_json["rows"]

        wb = Workbook()
        ws = wb.active
        ws.append(_FULL_DEDUP_HEADERS)
        for row in _FULL_DEDUP_ROWS:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)

        captured_file = _capture_cluster_rows(monkeypatch)
        resp = await client.post(
            "/api/dedup/file",
            files={"file": ("candidates.xlsx", buf.getvalue(), self._XTYPE)},
        )
        assert resp.status_code == 200
        file_rows = captured_file["rows"]

        assert len(file_rows) == len(json_rows) == len(_FULL_DEDUP_ROWS)
        for json_row, file_row in zip(json_rows, file_rows):
            assert file_row.model_dump() == json_row.model_dump()
