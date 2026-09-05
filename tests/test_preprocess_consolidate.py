"""Tests for the row-grain -> customer-grain consolidation stage.

Offline pytest — no network, no LLM. Drives ``consolidate_rows`` directly for
the transform, ``consolidate_workbook`` for the XLSX round-trip, the routes
via httpx for the HTTP contract, and ``split_consolidated`` for the
downstream delimiter contract the whole stage depends on.

The stage is a COLUMN-APPEND, not a collapse: almost every test below is
ultimately an assertion that no row was dropped, merged, reordered or
otherwise altered.
"""

from __future__ import annotations

import io
import json
import os
import random
import sys
from pathlib import Path

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ.setdefault("ENV", "local")
os.environ.setdefault("MOCK_EXTERNAL_CALLS", "true")
os.environ.setdefault("OPENAI_API_KEY", "test-key-for-mocks")

from dedup.consolidate import (
    CONSOLIDATED_DELIMITER,
    EXAMPLE_REQUEST,
    EXAMPLE_RESPONSE,
    EXAMPLE_ROWS_IN,
    EXAMPLE_ROWS_OUT,
    EXAMPLE_SUMMARY,
    ConsolidateRequest,
    ConsolidateResponse,
    COMPANY_CODE_CONSOLIDATED_HEADER as CC_OUT,
    SALES_ORG_CONSOLIDATED_HEADER as SO_OUT,
    consolidate_rows,
    consolidate_values,
    customer_key,
)
from dedup.consolidate_xlsx import ConsolidateFileError, consolidate_workbook
from dedup.scoring import ScoringRow, derived_counts, split_consolidated


def row(customer, company_code=None, sales_org=None, **extra) -> dict:
    """One row-grain input row, in the extract's own column spelling."""
    return {
        "Customer": customer,
        "Company Code": company_code,
        "Sales Organization": sales_org,
        **extra,
    }


def _values(rows: list[dict]) -> list[tuple[str, str]]:
    return [(r[CC_OUT], r[SO_OUT]) for r in rows]


# ---------------------------------------------------------------------------
# The transform
# ---------------------------------------------------------------------------

class TestColumnAppend:
    def test_row_count_preserved_and_originals_untouched(self):
        rows = [
            row("13119338", "1140", None, **{"Name 1": "Contra Costa County"}),
            row("13119338", "1207", None, **{"Name 1": "Contra Costa County"}),
            row("13118369", "1101", "1021", **{"Name 1": "Merck & Co Inc"}),
        ]
        before = [dict(r) for r in rows]

        out, summary = consolidate_rows(rows)

        assert len(out) == len(rows) == summary.rows_in == summary.rows_out == 3
        # The caller's dicts are not mutated, and every original key/value
        # survives verbatim in the output.
        assert rows == before
        for original, result in zip(before, out):
            for key, value in original.items():
                assert result[key] == value

    def test_column_order_is_stable_and_appends_exactly_two(self):
        rows = [row("1", "1140", "2401", **{"Name 1": "A", "City": "B"})]
        out, _ = consolidate_rows(rows)
        assert list(out[0]) == [
            "Customer", "Company Code", "Sales Organization",
            "Name 1", "City", CC_OUT, SO_OUT,
        ]
        # Explicitly NOT the counts — those are the scorer's to derive.
        assert "Company_Code_Count" not in out[0]
        assert "Sales_Org_Count" not in out[0]

    def test_rows_returned_in_input_order(self):
        rows = [row(c, "1140") for c in ("300", "100", "200", "100")]
        out, _ = consolidate_rows(rows)
        assert [r["Customer"] for r in out] == ["300", "100", "200", "100"]


class TestValueSet:
    def test_deduplication_14_rows_11_distinct_company_codes(self):
        """Customer 13118369 in dedup_STRESS_200_v1-verified_sapgrain.xlsx."""
        codes = [
            "1101", "1101", "1101", "1140", "1140", "1207", "1225",
            "1240", "1505", "1506", "1507", "1543", "1569", "1571",
        ]
        out, summary = consolidate_rows([row("13118369", c) for c in codes])

        assert len(out) == 14
        expected = "1101,1140,1207,1225,1240,1505,1506,1507,1543,1569,1571"
        assert {r[CC_OUT] for r in out} == {expected}
        assert len(expected.split(CONSOLIDATED_DELIMITER)) == 11
        assert summary.customers_with_multiple_company_codes == 1

    def test_blanks_dropped_not_emitted_as_empty_positions(self):
        """7 rows, one sales org -> "2401", never ",,2401,,,,"."""
        rows = [row("13119338", f"11{n:02d}", "2401" if n == 3 else None)
                for n in range(1, 8)]
        out, _ = consolidate_rows(rows)
        assert {r[SO_OUT] for r in out} == {"2401"}
        assert len(out) == 7

    def test_the_two_lists_are_independent_in_length(self):
        rows = [
            row("42", "1101", "1021"),
            row("42", "1140", "1021"),
            row("42", "1207", None),
            row("42", "1225", "2251"),
        ]
        out, _ = consolidate_rows(rows)
        company_codes, sales_orgs = out[0][CC_OUT], out[0][SO_OUT]
        assert company_codes == "1101,1140,1207,1225"
        assert sales_orgs == "1021,2251"
        # 4 vs 2: nothing is zipped, padded or positionally aligned.
        assert len(company_codes.split(",")) != len(sales_orgs.split(","))

    def test_numeric_before_non_numeric_ascending(self):
        assert consolidate_values(["ZZ", "1207", "AB", "1140", "9"]) == \
            "9,1140,1207,AB,ZZ"

    def test_excel_float_codes_render_as_integers(self):
        # openpyxl hands numeric cells back as floats; 1140.0 and "1140" are
        # the same company code and must not become two.
        assert consolidate_values([1140.0, "1140", 1207]) == "1140,1207"


class TestGrouping:
    def test_leading_zeros_group_together(self):
        rows = [row("0013119338", "1140"), row("13119338", "1207")]
        out, summary = consolidate_rows(rows)
        assert summary.customers == 1
        assert {r[CC_OUT] for r in out} == {"1140,1207"}
        # The Customer cell itself is never rewritten.
        assert [r["Customer"] for r in out] == ["0013119338", "13119338"]

    def test_single_row_customer_gets_a_one_element_list(self):
        out, summary = consolidate_rows([row("77", "1140", "2401")])
        assert out[0][CC_OUT] == "1140"
        assert out[0][SO_OUT] == "2401"
        assert summary.max_rows_per_customer == 1
        assert summary.customers_with_multiple_company_codes == 0

    def test_empty_group_is_an_empty_string(self):
        """Never None, never "None", never "[]"."""
        out, summary = consolidate_rows([row("77", None, None)] * 2)
        assert out[0][CC_OUT] == ""
        assert out[0][SO_OUT] == ""
        assert out[0][CC_OUT] is not None
        assert summary.customers_with_no_company_code == 1
        assert summary.customers_with_no_sales_org == 1


class TestDeterminismAndIdempotency:
    def test_idempotent_second_run_is_a_no_op(self):
        rows = [row("42", "1207", "2071"), row("42", "1140", None)]
        first, _ = consolidate_rows(rows)
        second, _ = consolidate_rows(first)
        assert json.dumps(first, sort_keys=False) == json.dumps(second, sort_keys=False)
        # Overwritten in place, not appended a second time.
        assert list(first[0]) == list(second[0])

    def test_stale_values_are_recomputed_not_preserved(self):
        rows = [{**row("42", "1207", "2071"), CC_OUT: "9999", SO_OUT: "9999"}]
        out, _ = consolidate_rows(rows)
        assert out[0][CC_OUT] == "1207"
        assert out[0][SO_OUT] == "2071"

    def test_shuffling_rows_within_a_customer_changes_nothing(self):
        rows = [row("42", f"1{n:03d}", f"2{n:03d}") for n in range(20)]
        expected = _values(consolidate_rows(rows)[0])[0]
        for seed in range(10):
            shuffled = list(rows)
            random.Random(seed).shuffle(shuffled)
            assert _values(consolidate_rows(shuffled)[0])[0] == expected


class TestErrorPolicy:
    @pytest.mark.parametrize("customer", [None, "", "   "])
    def test_blank_customer_counted_passed_through_never_raised(self, customer):
        rows = [row("42", "1140"), row(customer, "1207"), row("42", "1240")]
        out, summary = consolidate_rows(rows)

        assert summary.errors == 1
        assert summary.rows_in == summary.rows_out == 3
        assert out[1]["Company Code"] == "1207"  # passed through unchanged
        assert out[1][CC_OUT] == ""
        assert out[1][SO_OUT] == ""
        # Its company code does not leak into any other customer's list.
        assert out[0][CC_OUT] == "1140,1240"
        assert any("blank Customer" in w for w in summary.warnings)

    def test_missing_company_code_column_is_a_warning_not_an_error(self):
        rows = [{"Customer": "42", "Sales Organization": "2401"}]
        out, summary = consolidate_rows(rows)
        assert summary.errors == 0
        assert out[0][CC_OUT] == ""
        assert out[0][SO_OUT] == "2401"
        assert any("'Company Code' column" in w for w in summary.warnings)
        assert not any("'Sales Organization' column" in w for w in summary.warnings)

    def test_snake_case_aliases_are_accepted(self):
        rows = [
            {"customer": "0042", "company_code": "1140", "sales_organization": "2401"},
            {"customer": "42", "company_code": "1207", "sales_organization": None},
        ]
        out, summary = consolidate_rows(rows)
        assert summary.customers == 1
        assert out[0][CC_OUT] == "1140,1207"
        assert out[0][SO_OUT] == "2401"

    def test_empty_request_is_valid(self):
        out, summary = consolidate_rows([])
        assert out == []
        assert summary.rows_in == summary.rows_out == summary.customers == 0
        assert summary.warnings == []


class TestSummary:
    def test_counts(self):
        rows = [
            row("1", "1140", "2401"), row("1", "1207", "2401"),  # multi cc
            row("2", None, None),                                # no cc, no so
            row("3", "1101", "1021"), row("3", "1101", "2071"),  # multi so
        ]
        _, summary = consolidate_rows(rows)
        assert summary.rows_in == 5
        assert summary.rows_out == 5
        assert summary.customers == 3
        assert summary.customers_with_no_company_code == 1
        assert summary.customers_with_multiple_company_codes == 1
        assert summary.customers_with_no_sales_org == 1
        assert summary.customers_with_multiple_sales_orgs == 1
        assert summary.max_rows_per_customer == 2

    def test_batch_boundary_warning_is_opt_in_and_names_the_customers(self):
        rows = [row("100", "1140"), row("200", "1207"), row("300", "1240")]
        _, quiet = consolidate_rows(rows)
        assert quiet.warnings == []

        _, noisy = consolidate_rows(rows, warn_batch_boundary=True)
        warning = next(w for w in noisy.warnings if "Batch-boundary" in w)
        assert "100" in warning and "300" in warning
        assert "200" not in warning  # interior rows cannot be split off


# ---------------------------------------------------------------------------
# The delimiter contract (§2) — without this the whole stage is silently wrong
# ---------------------------------------------------------------------------

class TestSplitConsolidated:
    def test_both_delimiters_yield_the_same_count(self):
        assert len(split_consolidated("1140;1207")) == 2
        assert len(split_consolidated("1140,1207")) == 2

    def test_whitespace_and_trailing_empties_are_dropped(self):
        assert split_consolidated("1140, 1207 ,,") == ["1140", "1207"]
        assert split_consolidated("1003;1017;") == ["1003", "1017"]
        assert split_consolidated(None) == []
        assert split_consolidated("") == []

    def test_the_writer_and_the_reader_agree(self):
        value = consolidate_values(["1140", "1207", "1240"])
        assert split_consolidated(value) == ["1140", "1207", "1240"]

    def test_derived_counts_over_a_comma_joined_row(self):
        """If this comes back 1, §2 was not done."""
        scoring_row = ScoringRow(
            row_id="13118369",
            Company_Code_Consolidated=(
                "1101,1140,1207,1225,1240,1505,1506,1507,1543,1569,1571"
            ),
            Sales_Org_Consolidated="1021,1401,2071,2251,2401,5431,5711",
        )
        company_codes, sales_orgs, _ = derived_counts(scoring_row)
        assert (company_codes, sales_orgs) == (11, 7)

    def test_semicolon_inputs_keep_their_existing_count(self):
        """Widening is additive — no existing extract changes its score."""
        scoring_row = ScoringRow(
            row_id="1",
            Company_Code_Consolidated="1003;1017;1042",
            Sales_Org_Consolidated="2401",
        )
        assert derived_counts(scoring_row)[:2] == (3, 1)


# ---------------------------------------------------------------------------
# XLSX transport
# ---------------------------------------------------------------------------

def _workbook(headers: list[str], rows: list[list], *, extra_sheet=True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    if extra_sheet:
        other = wb.create_sheet("Weights")
        other.append(["Criterion", "Band", "Points"])
        other.append(["company_code_count", "3+", 25])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _sheet_matrix(contents: bytes) -> dict[str, list[tuple]]:
    wb = load_workbook(io.BytesIO(contents))
    return {ws.title: [tuple(r) for r in ws.iter_rows(values_only=True)]
            for ws in wb.worksheets}


class TestWorkbook:
    HEADERS = ["Customer", "Name 1", "Company Code", "Sales Organization"]

    def test_columns_appended_other_sheets_and_columns_preserved(self):
        contents = _workbook(self.HEADERS, [
            ["0013119338", "Contra Costa County", "1140", None],
            ["13119338", "Contra Costa County", "1207", "2401"],
        ])
        out, summary = consolidate_workbook(contents)
        sheets = _sheet_matrix(out)

        assert list(sheets) == ["Data", "Weights"]
        assert sheets["Weights"][1] == ("company_code_count", "3+", 25)
        assert sheets["Data"][0] == tuple(self.HEADERS) + (CC_OUT, SO_OUT)
        assert len(sheets["Data"]) == 3  # header + 2 rows, none dropped
        for data_row, original in zip(sheets["Data"][1:], contents and [
            ("0013119338", "Contra Costa County", "1140", None),
            ("13119338", "Contra Costa County", "1207", "2401"),
        ]):
            assert data_row[:4] == original
            assert data_row[4:] == ("1140,1207", "2401")
        assert summary.rows_in == summary.rows_out == 2
        assert summary.customers == 1

    def test_rerunning_a_processed_workbook_is_a_no_op(self):
        contents = _workbook(self.HEADERS, [
            ["42", "A", "1207", "2071"],
            ["42", "A", "1140", None],
        ])
        once, _ = consolidate_workbook(contents)
        twice, _ = consolidate_workbook(once)
        # Cell-for-cell identical on every sheet. (The zip bytes themselves
        # carry archive metadata, so the matrix is the meaningful invariant.)
        assert _sheet_matrix(twice) == _sheet_matrix(once)
        assert len(_sheet_matrix(once)["Data"][0]) == len(self.HEADERS) + 2

    def test_blank_customer_row_survives_with_empty_columns(self):
        contents = _workbook(self.HEADERS, [
            ["42", "A", "1140", None],
            [None, "orphan", "1207", None],
        ])
        out, summary = consolidate_workbook(contents)
        data = _sheet_matrix(out)["Data"]
        assert len(data) == 3
        assert data[2][:4] == (None, "orphan", "1207", None)
        # "" is what XLSX calls a blank cell — openpyxl reads it back as None.
        assert data[2][4:] == (None, None)
        assert summary.errors == 1
        assert data[1][4] == "1140"

    def test_fully_blank_trailing_row_is_not_counted(self):
        contents = _workbook(self.HEADERS, [
            ["42", "A", "1140", None],
            [None, None, None, None],
        ])
        _, summary = consolidate_workbook(contents)
        assert summary.rows_in == summary.rows_out == 1
        assert summary.errors == 0

    def test_missing_customer_column_is_a_400(self):
        contents = _workbook(["Name 1", "Company Code"], [["A", "1140"]])
        with pytest.raises(ConsolidateFileError, match="Customer"):
            consolidate_workbook(contents)

    def test_named_sheet_selection(self):
        contents = _workbook(self.HEADERS, [["42", "A", "1140", "2401"]])
        with pytest.raises(ConsolidateFileError, match="No sheet named"):
            consolidate_workbook(contents, "Nope")
        out, summary = consolidate_workbook(contents, "Data")
        assert summary.customers == 1

    def test_missing_sales_org_column_warns_and_writes_empty(self):
        contents = _workbook(["Customer", "Company Code"], [["42", "1140"]])
        out, summary = consolidate_workbook(contents)
        data = _sheet_matrix(out)["Data"]
        assert data[0] == ("Customer", "Company Code", CC_OUT, SO_OUT)
        assert data[1] == ("42", "1140", "1140", None)  # "" -> blank cell
        assert any("'Sales Organization' column" in w for w in summary.warnings)

    def test_the_two_transports_agree(self):
        rows = [
            row("0042", "1140", None, **{"Name 1": "A"}),
            row("42", "1207", "2401", **{"Name 1": "A"}),
            row("43", None, None, **{"Name 1": "B"}),
        ]
        json_rows, json_summary = consolidate_rows(rows)

        contents = _workbook(
            ["Customer", "Company Code", "Sales Organization", "Name 1"],
            [[r["Customer"], r["Company Code"], r["Sales Organization"], r["Name 1"]]
             for r in rows],
            extra_sheet=False,
        )
        out, file_summary = consolidate_workbook(contents)
        data = _sheet_matrix(out)["Data"]

        # Normalised for the one representational difference the formats
        # force: XLSX has no empty-string cell, so "" reads back as None.
        assert [tuple(v or "" for v in d[-2:]) for d in data[1:]] == \
            _values(json_rows)
        assert file_summary.model_dump(exclude={"warnings"}) == \
            json_summary.model_dump(exclude={"warnings"})


# ---------------------------------------------------------------------------
# HTTP contract
# ---------------------------------------------------------------------------

from api.app import app  # noqa: E402


@pytest.fixture
def transport():
    return ASGITransport(app=app)


@pytest_asyncio.fixture
async def client(transport):
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c


class TestEndpoints:
    @pytest.mark.asyncio
    async def test_json_endpoint_round_trip(self, client):
        resp = await client.post("/api/preprocess/consolidate", json={"rows": [
            row("13119338", "1140", None), row("13119338", "1240", "2401"),
        ]})
        assert resp.status_code == 200
        data = resp.json()
        assert len(data["rows"]) == 2
        assert data["rows"][0][CC_OUT] == "1140,1240"
        assert data["rows"][0][SO_OUT] == "2401"
        assert data["summary"]["rows_in"] == data["summary"]["rows_out"] == 2

    @pytest.mark.asyncio
    async def test_json_endpoint_always_warns_about_the_batch_boundary(self, client):
        resp = await client.post("/api/preprocess/consolidate", json={
            "rows": [row("13119338", "1140")]
        })
        warnings = resp.json()["summary"]["warnings"]
        assert any("Batch-boundary" in w for w in warnings)

    @pytest.mark.asyncio
    async def test_empty_rows_200_zeroed_summary(self, client):
        resp = await client.post("/api/preprocess/consolidate", json={"rows": []})
        assert resp.status_code == 200
        assert resp.json()["summary"]["rows_in"] == 0

    @pytest.mark.asyncio
    async def test_blank_customer_is_not_a_500(self, client):
        resp = await client.post("/api/preprocess/consolidate", json={
            "rows": [row("", "1140"), row("42", "1207")]
        })
        assert resp.status_code == 200
        assert resp.json()["summary"]["errors"] == 1
        assert len(resp.json()["rows"]) == 2

    @pytest.mark.asyncio
    async def test_file_endpoint_returns_a_workbook(self, client):
        contents = _workbook(
            ["Customer", "Company Code", "Sales Organization"],
            [["0013119338", "1140", None], ["13119338", "1240", "2401"]],
        )
        resp = await client.post(
            "/api/preprocess/consolidate/file",
            files={"file": ("extract.xlsx", contents,
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert "extract_consolidated.xlsx" in resp.headers["content-disposition"]
        data = _sheet_matrix(resp.content)["Data"]
        assert data[1][-2:] == ("1140,1240", "2401")

    @pytest.mark.asyncio
    async def test_file_endpoint_rejects_a_non_workbook(self, client):
        resp = await client.post(
            "/api/preprocess/consolidate/file",
            files={"file": ("extract.csv", b"Customer,Company Code\n42,1140\n",
                            "text/csv")},
        )
        assert resp.status_code == 400

    @pytest.mark.asyncio
    async def test_file_endpoint_rejects_an_empty_upload(self, client):
        resp = await client.post(
            "/api/preprocess/consolidate/file",
            files={"file": ("extract.xlsx", b"", "application/vnd.ms-excel")},
        )
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
# OpenAPI examples — pinned to the real transform so Swagger cannot drift
# ---------------------------------------------------------------------------

class TestOpenAPIExamples:
    """The documented example must be what the endpoint actually returns.

    A hand-written Swagger example is a second implementation of the contract,
    and the one nobody runs. These tests run it.
    """

    def test_the_documented_example_is_the_real_output(self):
        out, summary = consolidate_rows(EXAMPLE_ROWS_IN, warn_batch_boundary=True)
        assert out == EXAMPLE_ROWS_OUT
        assert summary.model_dump(exclude={"warnings"}) == \
            {k: v for k, v in EXAMPLE_SUMMARY.items() if k != "warnings"}

    def test_the_documented_warning_is_a_real_warning(self):
        _, summary = consolidate_rows(EXAMPLE_ROWS_IN, warn_batch_boundary=True)
        # The example elides the tail of a deliberately long warning; what it
        # does show has to be verbatim.
        prefix = EXAMPLE_SUMMARY["warnings"][0].removesuffix("...")
        assert len(summary.warnings) == 1
        assert summary.warnings[0].startswith(prefix)

    def test_the_examples_validate_against_their_own_models(self):
        ConsolidateRequest.model_validate({"rows": EXAMPLE_ROWS_IN})
        ConsolidateResponse.model_validate(
            {"rows": EXAMPLE_ROWS_OUT, "summary": EXAMPLE_SUMMARY}
        )

    @pytest.mark.asyncio
    async def test_the_example_posted_to_the_endpoint_round_trips(self, client):
        resp = await client.post(
            "/api/preprocess/consolidate", json={"rows": EXAMPLE_ROWS_IN}
        )
        assert resp.status_code == 200
        assert resp.json()["rows"] == EXAMPLE_ROWS_OUT

    def test_openapi_carries_both_examples(self):
        """On the models AND at the media-type level, where Swagger UI always
        looks — this endpoint's body is a list of free-form dicts, which a
        caller cannot guess from the schema alone."""
        spec = app.openapi()
        schemas = spec["components"]["schemas"]
        assert schemas["ConsolidateRequest"]["examples"] == [EXAMPLE_REQUEST]
        assert schemas["ConsolidateResponse"]["examples"] == [EXAMPLE_RESPONSE]

        post = spec["paths"]["/api/preprocess/consolidate"]["post"]
        body = post["requestBody"]["content"]["application/json"]["examples"]
        assert [e["value"] for e in body.values()] == [EXAMPLE_REQUEST]
        assert post["responses"]["200"]["content"]["application/json"]["example"] \
            == EXAMPLE_RESPONSE


# ---------------------------------------------------------------------------
# Acceptance gate (§6) — the two verified customers, when the fixture is present
# ---------------------------------------------------------------------------

STRESS_FIXTURE = (
    Path.home() / "Downloads" / "dedup_STRESS_200_v1-verified_sapgrain.xlsx"
)


@pytest.mark.skipif(
    not STRESS_FIXTURE.exists(),
    reason=f"row-grain stress fixture not present at {STRESS_FIXTURE}",
)
class TestAcceptanceGate:
    """The verified numbers from dedup_STRESS_200_v1-verified_sapgrain.xlsx.

    The fixture lives outside the repository (like the other real extracts),
    so these skip cleanly on a machine that does not have it.
    """

    @staticmethod
    def _consolidated():
        out, summary = consolidate_workbook(STRESS_FIXTURE.read_bytes())
        wb = load_workbook(io.BytesIO(out))
        ws = wb["Data"]
        headers = [c.value for c in ws[1]]
        index = {h: i for i, h in enumerate(headers)}
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return index, rows, summary

    def test_full_file_row_and_customer_counts(self):
        _, rows, summary = self._consolidated()
        assert len(rows) == 839
        assert summary.rows_in == summary.rows_out == 839
        assert summary.customers == 183

    @pytest.mark.parametrize(
        "customer,rows_expected,company_codes,sales_orgs",
        [
            ("13119338", 7,
             "1140,1207,1240,1505,1506,1507,1569",
             "2401"),
            ("13118369", 14,
             "1101,1140,1207,1225,1240,1505,1506,1507,1543,1569,1571",
             "1021,1401,2071,2251,2401,5431,5711"),
        ],
    )
    def test_verified_customer(self, customer, rows_expected, company_codes, sales_orgs):
        index, rows, _ = self._consolidated()
        group = [r for r in rows
                 if customer_key(r[index["Customer"]]) == customer]

        assert len(group) == rows_expected
        # Both strings appear on EVERY row of the group.
        assert {r[index[CC_OUT]] for r in group} == {company_codes}
        assert {r[index[SO_OUT]] for r in group} == {sales_orgs}

        # ... and survive the trip through the scorer's splitter.
        scoring_row = ScoringRow(
            row_id=customer,
            Company_Code_Consolidated=company_codes,
            Sales_Org_Consolidated=sales_orgs,
        )
        cc_count, so_count, _ = derived_counts(scoring_row)
        assert cc_count == len(company_codes.split(","))
        assert so_count == len(sales_orgs.split(","))

    def test_merck_earns_the_combined_presence_bonus(self):
        from dedup.scoring import load_weights, score_row

        index, rows, _ = self._consolidated()
        merck = next(r for r in rows
                     if customer_key(r[index["Customer"]]) == "13118369")
        scoring_row = ScoringRow(
            row_id="13118369",
            Company_Code_Consolidated=merck[index[CC_OUT]],
            Sales_Org_Consolidated=merck[index[SO_OUT]],
        )
        assert derived_counts(scoring_row)[:2] == (11, 7)
        # current_year anchors the two *_last_used ladders only; this row
        # carries neither, and combined_presence_bonus is year-independent, so
        # the value is immaterial here — it is required to keep any caller from
        # silently scoring both ladders zero.
        breakdown, _ = score_row(
            scoring_row, load_weights(), current_year=2026
        )
        assert breakdown["combined_presence_bonus"] > 0
