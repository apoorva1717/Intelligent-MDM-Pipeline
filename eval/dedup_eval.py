"""Offline evaluation harness for Phase 2 dedup + golden-record election.

Scores a *scored* workbook (the output of ``/api/dedup/score/file``, which still
carries the ``expected_cluster`` / ``expected_routing`` fixture columns) against
its ground truth. No network, no LLM — pure arithmetic over the sheet, so it can
run on any scored workbook that kept the fixture columns.

This is the engine behind the thesis evaluation chapter: it reports the standard
pairwise clustering metrics AND the named business-risk counts that make the
result auditable ("precision 0.90, here are the three wrongful blocks").

    python -m eval.dedup_eval path/to/scored.xlsx [--out eval_report.json]

Metrics
-------
Pairwise (vs the fixture):
  ground-truth pairs = same ``expected_cluster`` among rows whose
    ``expected_routing`` is cluster|manual_review;
  predicted pairs    = same ``cluster_id``;
  -> TP / FP / FN, precision, recall, F1.

Business-risk (counts AND offending row_ids):
  wrongful_block_candidates: expected_routing == unique  AND is_golden_record == False
  competing_goldens:         expected_routing == cluster AND routing == unique
  uncertainty_upgrades:      expected_routing == manual_review AND election_status == proposed

Election:
  clusters, elections, manual_review count, and tie-break invocations
  (clusters whose top ``score_final`` is shared by >=2 members, so the winner
  was decided by the tie-break, not the score).
"""

from __future__ import annotations

import json
import sys
from itertools import combinations
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Normalised header -> canonical field name. Same tolerant matching as the
# scoring/route readers (case / whitespace / punctuation insensitive).
_FIELD_BY_HEADER = {
    "customer": "row_id",
    "rowid": "row_id",
    "expectedcluster": "expected_cluster",
    "expectedrouting": "expected_routing",
    "clusterid": "cluster_id",
    "routing": "routing",
    "isgoldenrecord": "is_golden_record",
    "electionstatus": "election_status",
    "scorefinal": "score_final",
    "goldenrecordid": "golden_record_id",
    "proposedgoldenid": "proposed_golden_id",
    "scoredwithweightsversion": "weights_version",
}

_CLUSTER_ROUTINGS = {"cluster", "manual_review"}


def _norm(name: object) -> str:
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _clean(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_bool(value: object) -> Optional[bool]:
    """Excel booleans arrive native; a blanked (manual_review) golden is None."""
    if isinstance(value, bool):
        return value
    text = _clean(value)
    if text is None:
        return None
    low = text.casefold()
    if low in ("true", "1"):
        return True
    if low in ("false", "0"):
        return False
    return None


def _as_float(value: object) -> Optional[float]:
    text = _clean(value)
    if text is None:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_scored_rows(path: str | Path) -> List[dict]:
    """Read a scored workbook into a list of plain dicts (one per data row)."""
    from openpyxl import load_workbook

    wb = load_workbook(Path(path), data_only=True)
    # First sheet carrying a Customer/row_id column is the data sheet.
    ws = None
    for candidate in wb.worksheets:
        header = next(candidate.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if any(_norm(h) in ("customer", "rowid") for h in header if h is not None):
            ws = candidate
            break
    if ws is None:
        raise ValueError("No sheet with a 'Customer'/'row_id' column found.")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    col_field: Dict[int, str] = {}
    for idx, header in enumerate(header_row):
        field = _FIELD_BY_HEADER.get(_norm(header))
        if field and idx not in col_field:
            col_field[idx] = field

    rows: List[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        record = {field: None for field in _FIELD_BY_HEADER.values()}
        for idx, field in col_field.items():
            record[field] = raw[idx] if idx < len(raw) else None
        if _clean(record["row_id"]) is None:
            continue
        record["row_id"] = str(record["row_id"]).strip()
        record["expected_cluster"] = _clean(record["expected_cluster"])
        record["expected_routing"] = (
            (_clean(record["expected_routing"]) or "").casefold() or None
        )
        record["cluster_id"] = _clean(record["cluster_id"])
        record["routing"] = (_clean(record["routing"]) or "").casefold() or None
        record["election_status"] = (
            (_clean(record["election_status"]) or "").casefold() or None
        )
        record["is_golden_record"] = _as_bool(record["is_golden_record"])
        record["score_final"] = _as_float(record["score_final"])
        record["weights_version"] = _clean(record["weights_version"])
        rows.append(record)
    return rows


def _pairs_by_key(rows: List[dict], key: str) -> set:
    """Unordered same-``key`` row_id pairs (blank keys form no pairs)."""
    groups: Dict[str, List[str]] = {}
    for r in rows:
        value = r.get(key)
        if value is None:
            continue
        groups.setdefault(str(value), []).append(r["row_id"])
    pairs = set()
    for members in groups.values():
        for a, b in combinations(sorted(members), 2):
            pairs.add((a, b))
    return pairs


def pairwise_metrics(rows: List[dict]) -> dict:
    """TP/FP/FN + precision/recall/F1 for predicted vs ground-truth pairs."""
    truth_rows = [
        r for r in rows if r["expected_routing"] in _CLUSTER_ROUTINGS
    ]
    gt = _pairs_by_key(truth_rows, "expected_cluster")
    pred = _pairs_by_key(rows, "cluster_id")

    tp = len(gt & pred)
    fp = len(pred - gt)
    fn = len(gt - pred)
    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = (
        2 * precision * recall / (precision + recall)
        if (precision + recall)
        else 0.0
    )
    return {
        "true_positives": tp,
        "false_positives": fp,
        "false_negatives": fn,
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
        "ground_truth_pairs": len(gt),
        "predicted_pairs": len(pred),
    }


def business_risk_metrics(rows: List[dict]) -> dict:
    """The three named, auditable failure counts — each with offending ids."""
    wrongful_block = [
        r["row_id"] for r in rows
        if r["expected_routing"] == "unique" and r["is_golden_record"] is False
    ]
    competing_goldens = [
        r["row_id"] for r in rows
        if r["expected_routing"] == "cluster" and r["routing"] == "unique"
    ]
    uncertainty_upgrades = [
        r["row_id"] for r in rows
        if r["expected_routing"] == "manual_review"
        and r["election_status"] == "proposed"
    ]

    def _entry(ids: List[str]) -> dict:
        return {"count": len(ids), "row_ids": sorted(ids)}

    return {
        "wrongful_block_candidates": _entry(wrongful_block),
        "competing_goldens": _entry(competing_goldens),
        "uncertainty_upgrades": _entry(uncertainty_upgrades),
    }


def election_metrics(rows: List[dict]) -> dict:
    """Cluster / election counts and tie-break invocations."""
    clusters: Dict[str, List[dict]] = {}
    for r in rows:
        if r["cluster_id"] is not None:
            clusters.setdefault(r["cluster_id"], []).append(r)

    elections = sum(1 for r in rows if r["is_golden_record"] is True and r["cluster_id"])
    manual_review = sum(1 for r in rows if r["election_status"] == "manual_review")

    tiebreak_clusters = []
    for cid, members in clusters.items():
        scores = [m["score_final"] for m in members if m["score_final"] is not None]
        if len(scores) >= 2 and scores.count(max(scores)) >= 2:
            tiebreak_clusters.append(cid)

    return {
        "clusters": len(clusters),
        "elections": elections,
        "manual_review_rows": manual_review,
        "tiebreak_decided_clusters": {
            "count": len(tiebreak_clusters),
            "cluster_ids": sorted(tiebreak_clusters),
        },
    }


def evaluate(path: str | Path) -> dict:
    """Full evaluation report for one scored workbook."""
    rows = load_scored_rows(path)
    versions = sorted({r["weights_version"] for r in rows if r["weights_version"]})
    return {
        "source": str(path),
        "rows_evaluated": len(rows),
        # >1 version means the workbook mixes rows scored under different
        # weights (score drift) — flagged for the reader.
        "weights_versions": versions,
        "pairwise": pairwise_metrics(rows),
        "business_risk": business_risk_metrics(rows),
        "election": election_metrics(rows),
    }


def _format_report(report: dict) -> str:
    """Human-readable console table."""
    p = report["pairwise"]
    b = report["business_risk"]
    e = report["election"]
    lines = [
        f"Dedup evaluation — {report['source']}",
        f"  rows evaluated: {report['rows_evaluated']}",
        "",
        "  Pairwise clustering",
        f"    precision {p['precision']:.2f}   recall {p['recall']:.2f}   F1 {p['f1']:.2f}",
        f"    TP {p['true_positives']}  FP {p['false_positives']}  FN {p['false_negatives']}"
        f"   (GT pairs {p['ground_truth_pairs']}, predicted {p['predicted_pairs']})",
        "",
        "  Business risk (named)",
    ]
    for name in ("wrongful_block_candidates", "competing_goldens", "uncertainty_upgrades"):
        entry = b[name]
        ids = ", ".join(entry["row_ids"]) if entry["row_ids"] else "-"
        lines.append(f"    {name}: {entry['count']}  [{ids}]")
    tb = e["tiebreak_decided_clusters"]
    lines += [
        "",
        "  Election",
        f"    clusters {e['clusters']}  elections {e['elections']}"
        f"  manual_review {e['manual_review_rows']}  tie-break-decided {tb['count']}",
    ]
    return "\n".join(lines)


def main(argv: Optional[List[str]] = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    if not argv or argv[0] in ("-h", "--help"):
        print(__doc__)
        return 0 if argv else 2

    path = argv[0]
    out = "eval_report.json"
    if "--out" in argv:
        out = argv[argv.index("--out") + 1]

    report = evaluate(path)
    print(_format_report(report))
    Path(out).write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"\nWrote {out}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
