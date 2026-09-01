"""Name-block evaluation against the solved reference (§5).

Scores an enriched workbook — the output of ``scripts/run_batch.py`` or
``POST /enrich/file`` — against ``testall100_SOLVED_REFERENCE_v1.xlsx``, whose
``Reference`` sheet holds an ``INPUT`` row and an ``EXPECTED`` row per record,
paired on ``Customer``.

The reference is not ground truth and does not claim to be: it was authored by
one reviewer from the thesis's documented rules plus general knowledge, and it
says so on its ``Method`` sheet. Two mechanisms keep the score honest about
that, and both are read from the workbook rather than hard-coded here:

``Match Rules``
    Per column: ``exact`` (string equality after trim), ``exact_ci``
    (case-insensitive — the thesis does not specify street or city casing), or
    ``skip`` (registry- and run-dependent columns: ROR ID, LEI ID, Domain,
    Operating Name, Search Terms, Record Type, every Flag and Provenance
    column). An empty expected cell in a skip column means "no claim", not
    "must be empty".

``Cell Notes``
    Per (Customer, column): overrides the column rule for one cell.
    ``any_of: a | b`` accepts either form, with ``(empty)`` in the list
    meaning a blank cell is acceptable; ``skip`` drops the cell entirely.
    68 such notes exist, each one a place the reviewer could not certify a
    single correct value.

No value from the test set is embedded in this file: every expectation, rule
and note is read from the reference workbook at run time.

Usage::

    python -m eval.name_eval enriched.xlsx \\
        --reference testall100_SOLVED_REFERENCE_v1.xlsx \\
        [--baseline baseline_enriched.xlsx] [--out report.json]

Reports, in order:

* the name-block mismatch count — the headline number, Name 1-4 across all
  records;
* the "left exactly as supplied" count, with the records carrying it;
* every name write, with its source, verdict, confidence and flag state;
* the non-name columns (Domain, Contact, Email, Operating Name) diffed against
  ``--baseline``, which is what shows the change was scoped to names.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))

#: The columns this evaluation is about. Everything else is either scored by
#: its Match Rule or reported as a scope check.
NAME_COLUMNS = ("Name 1", "Name 2", "Name 3", "Name 4")

#: Columns that must not move when a name-only change ships (§5).
SCOPE_COLUMNS = ("Domain", "Contact", "Email", "Operating Name")

#: The sentence §1 is written against. Counted, and its records listed.
LEFT_AS_SUPPLIED = "left exactly as supplied"

_EMPTY_TOKEN = "(empty)"


def _norm(value: Any) -> str:
    """A cell as the comparison sees it: text, trimmed, blanks unified."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _load_sheet(path: Path, sheet: str | None = None) -> tuple[list[str], list[tuple]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [_norm(c) for c in rows[0]]
    return header, rows[1:]


def _by_customer(header: list[str], rows: list[tuple]) -> dict[str, dict[str, str]]:
    """Rows keyed by ``Customer``, each a column -> normalised value map."""
    try:
        key = header.index("Customer")
    except ValueError:
        raise SystemExit("Workbook has no 'Customer' column.")
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        if key >= len(row) or row[key] in (None, ""):
            continue
        out[_norm(row[key])] = {
            h: _norm(v) for h, v in zip(header, row) if h
        }
    return out


def load_reference(path: Path) -> dict[str, Any]:
    """The reference's expectations, rules and per-cell notes."""
    header, rows = _load_sheet(path, "Reference")
    kind = header.index("row_kind")
    cust = header.index("Customer")

    expected: dict[str, dict[str, str]] = {}
    supplied: dict[str, dict[str, str]] = {}
    for row in rows:
        if cust >= len(row) or row[cust] in (None, ""):
            continue
        record = {h: _norm(v) for h, v in zip(header, row) if h}
        target = expected if _norm(row[kind]) == "EXPECTED" else supplied
        target[_norm(row[cust])] = record

    rules: dict[str, str] = {}
    rh, rr = _load_sheet(path, "Match Rules")
    ci, ri = rh.index("column"), rh.index("rule")
    for row in rr:
        if ci < len(row) and row[ci]:
            rules[_norm(row[ci])] = _norm(row[ri]).lower()

    notes: dict[tuple[str, str], str] = {}
    nh, nr = _load_sheet(path, "Cell Notes")
    nc, ncol, nrule = nh.index("Customer"), nh.index("column"), nh.index("rule")
    for row in nr:
        if nc < len(row) and row[nc]:
            notes[(_norm(row[nc]), _norm(row[ncol]))] = _norm(row[nrule])

    return {
        "expected": expected, "input": supplied,
        "rules": rules, "notes": notes,
    }


def _cell_matches(
    actual: str, expected: str, rule: str, note: str | None,
) -> tuple[bool, str]:
    """``(matched, why)`` for one cell under its rule and note.

    The note overrides the column rule, which is what makes a widened cell a
    widened cell rather than a second opinion the scorer has to reconcile.
    """
    if note:
        low = note.lower()
        if low == "skip":
            return True, "skip (cell note)"
        if low.startswith("any_of:"):
            options = [o.strip() for o in note.split(":", 1)[1].split("|")]
            for option in options:
                if option == _EMPTY_TOKEN and not actual:
                    return True, "any_of (empty)"
                if _norm(option).casefold() == actual.casefold():
                    return True, "any_of"
            return False, f"any_of {options!r}"

    if rule == "skip":
        return True, "skip (column rule)"
    if rule == "exact_ci":
        return actual.casefold() == expected.casefold(), "exact_ci"
    return actual == expected, "exact"


def score(
    enriched: dict[str, dict[str, str]], reference: dict[str, Any],
) -> dict[str, Any]:
    """Score the name block, and report every cell that missed."""
    expected, rules, notes = (
        reference["expected"], reference["rules"], reference["notes"],
    )

    mismatches: list[dict[str, str]] = []
    scored = 0
    per_column: Counter[str] = Counter()
    missing: list[str] = []

    for customer, want in sorted(expected.items()):
        got = enriched.get(customer)
        if got is None:
            missing.append(customer)
            continue
        for column in NAME_COLUMNS:
            if column not in want:
                continue
            rule = rules.get(column, "exact")
            note = notes.get((customer, column))
            actual, target = got.get(column, ""), want.get(column, "")
            if rule == "skip" and not note:
                continue
            scored += 1
            ok, why = _cell_matches(actual, target, rule, note)
            if not ok:
                per_column[column] += 1
                mismatches.append({
                    "customer": customer, "column": column,
                    "expected": target, "actual": actual, "rule": why,
                    "input": reference["input"].get(customer, {}).get(column, ""),
                })

    return {
        "records": len(expected),
        "cells_scored": scored,
        "name_block_mismatches": len(mismatches),
        "mismatches_by_column": dict(per_column),
        "records_not_in_output": missing,
        "mismatches": mismatches,
    }


def left_as_supplied(enriched: dict[str, dict[str, str]]) -> list[str]:
    """Records whose Flag Reason still says the name was left as supplied."""
    return sorted(
        customer for customer, row in enriched.items()
        if LEFT_AS_SUPPLIED in row.get("Flag Reason", "").lower()
    )


def name_writes(
    enriched: dict[str, dict[str, str]], reference: dict[str, Any],
) -> list[dict[str, str]]:
    """Every record whose Name 1 or Name 2 differs from what it arrived with.

    Reported with the provenance the record carries, so "source, verdict,
    confidence, flag state" is read off the shipped row rather than from a
    log the reader cannot check.
    """
    out: list[dict[str, str]] = []
    for customer, row in sorted(enriched.items()):
        supplied = reference["input"].get(customer, {})
        for column, provenance in (
            ("Name 1", "Name 1 Provenance"), ("Name 2", "Name 2 Provenance"),
        ):
            before, after = supplied.get(column, ""), row.get(column, "")
            if not after or before == after:
                continue
            out.append({
                "customer": customer, "column": column,
                "was": before, "now": after,
                "provenance": row.get(provenance, ""),
                "flag_codes": row.get("Flag Codes", ""),
                "flagged": row.get("Flag for Review", ""),
            })
    return out


def scope_diff(
    enriched: dict[str, dict[str, str]],
    baseline: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    """Non-name columns that moved. §5 requires this list to be empty."""
    out: list[dict[str, str]] = []
    for customer, row in sorted(enriched.items()):
        was = baseline.get(customer)
        if was is None:
            continue
        for column in SCOPE_COLUMNS:
            if row.get(column, "") != was.get(column, ""):
                out.append({
                    "customer": customer, "column": column,
                    "baseline": was.get(column, ""), "now": row.get(column, ""),
                })
    return out


def _render(report: dict[str, Any]) -> None:
    s = report["score"]
    print("=" * 72)
    print("NAME-BLOCK EVALUATION")
    print("=" * 72)
    print(f"records scored              : {s['records']}")
    print(f"name cells compared         : {s['cells_scored']}")
    print(f"NAME-BLOCK MISMATCHES       : {s['name_block_mismatches']}")
    for column, count in sorted(s["mismatches_by_column"].items()):
        print(f"    {column:<10}            : {count}")
    if s["records_not_in_output"]:
        print(f"!! not in output            : {s['records_not_in_output']}")

    left = report["left_as_supplied"]
    print(f"\n'{LEFT_AS_SUPPLIED}'  : {len(left)} records")
    if left:
        print(f"    {', '.join(left)}")

    writes = report["name_writes"]
    print(f"\nNAME WRITES                 : {len(writes)}")
    for w in writes:
        flag = w["flag_codes"] or "-"
        print(f"  {w['customer']}  {w['column']}: {w['was']!r} -> {w['now']!r}")
        print(f"      provenance={w['provenance'] or '-'}  flags={flag}")

    if report.get("scope_diff") is not None:
        diff = report["scope_diff"]
        print(f"\nNON-NAME COLUMNS CHANGED    : {len(diff)}"
              f"  {'(scope held)' if not diff else '(SCOPE BREACH)'}")
        for d in diff:
            print(f"  {d['customer']} {d['column']}: "
                  f"{d['baseline']!r} -> {d['now']!r}")

    if s["mismatches"]:
        print(f"\nMISMATCHES ({len(s['mismatches'])})")
        for m in s["mismatches"]:
            print(f"  {m['customer']} {m['column']} [{m['rule']}]")
            print(f"      input    : {m['input']!r}")
            print(f"      expected : {m['expected']!r}")
            print(f"      actual   : {m['actual']!r}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("enriched", help="Enriched workbook to score.")
    ap.add_argument(
        "--reference", default="testall100_SOLVED_REFERENCE_v1.xlsx",
        help="The solved reference workbook.",
    )
    ap.add_argument(
        "--baseline", default=None,
        help="A prior enriched workbook, for the non-name scope diff.",
    )
    ap.add_argument("--out", default=None, help="Write the report as JSON.")
    args = ap.parse_args()

    reference = load_reference(Path(args.reference))
    enriched = _by_customer(*_load_sheet(Path(args.enriched)))

    report: dict[str, Any] = {
        "enriched": args.enriched,
        "reference": args.reference,
        "score": score(enriched, reference),
        "left_as_supplied": left_as_supplied(enriched),
        "name_writes": name_writes(enriched, reference),
    }
    if args.baseline:
        report["baseline"] = args.baseline
        report["scope_diff"] = scope_diff(
            enriched, _by_customer(*_load_sheet(Path(args.baseline))),
        )

    _render(report)
    if args.out:
        Path(args.out).write_text(
            json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8",
        )
        print(f"\nreport written to {args.out}")


if __name__ == "__main__":
    main()
